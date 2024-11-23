import tkinter as tk
from tkinter import messagebox,ttk
from ttkthemes import ThemedTk
import datetime
import openpyxl
from PIL import Image, ImageTk
# Load the workbook
WORKBOOK_PATH = "TTMS.xlsx"

def authenticate_user(username, password):
    # Load the workbook and Users sheet
    try:
        workbook = openpyxl.load_workbook(WORKBOOK_PATH)
        sheet = workbook["Users"]

        # Search for user
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                return row[2]  # Return the role if authentication succeeds
        return None
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load workbook: {e}")
        return None


def login_window(theme_color="#2196F3", window_size=(400, 500)):
    def login():
        username = username_entry.get()
        password = password_entry.get()

        if username and password:
            role = authenticate_user(username, password)
            if role:
                messagebox.showinfo("Success", f"Welcome {username}! ✨")
                root.destroy()
                redirect_user(role)
            else:
                messagebox.showerror("Error", "Invalid credentials ❌")
        else:
            messagebox.showwarning("Warning", "Please fill all fields ⚠️")

    root = tk.Tk()
    root.title("Login")

    # Make window size adjustable
    width, height = window_size
    root.geometry(f"{width}x{height}")
    root.configure(bg="white")

    # Center window
    root.eval('tk::PlaceWindow . center')

    # Style configuration
    style = {
        'bg': "white",
        'fg': theme_color,
        'font_title': ("Helvetica", 22, "bold"),
        'font_normal': ("Helvetica", 11),
        'padding': 20
    }

    # Main container
    main_frame = tk.Frame(root, bg=style['bg'])
    main_frame.place(relx=0.5, rely=0.5, anchor="center")

    # Title
    tk.Label(main_frame,
             text="Welcome Back!",
             font=style['font_title'],
             bg=style['bg'],
             fg=style['fg']).pack(pady=style['padding'])

    # Entry fields with modern look
    def create_entry(placeholder, show=None):
        frame = tk.Frame(main_frame, bg=style['bg'])
        frame.pack(pady=10, padx=style['padding'])

        entry = tk.Entry(frame,
                         font=style['font_normal'],
                         show=show,
                         width=25,
                         bd=0,
                         bg="#f0f0f0")
        entry.insert(0, placeholder)
        entry.pack(ipady=8, pady=5)

        # Add underline
        tk.Frame(frame, height=2, bg=theme_color).pack(fill="x")

        return entry

    username_entry = create_entry("Username")
    password_entry = create_entry("Password", show="●")

    # Login button
    login_btn = tk.Button(main_frame,
                          text="LOGIN",
                          command=login,
                          font=style['font_normal'],
                          bg=theme_color,
                          fg="white",
                          width=20,
                          bd=0,
                          cursor="hand2")
    login_btn.pack(pady=style['padding'])

    # Bind Enter key
    root.bind('<Return>', lambda event: login())

    root.mainloop()



# Role-based redirection (Placeholder for dashboards)
def redirect_user(role):
    if role == "Admin":
        messagebox.showinfo("Admin Dashboard", "Redirecting to Admin Dashboard...")
        admin_dashboard()
        # Call Admin Dashboard
    elif role == "Dispatcher":
        messagebox.showinfo("Dispatcher Dashboard", "Redirecting to Dispatcher Dashboard...")
        dispatcher_dashboard()
        # Call Dispatcher Dashboard
    elif role == "Accountant":
        messagebox.showinfo("Accountant Dashboard", "Redirecting to Accountant Dashboard...")
        accountant_dashboard()
        # Call Accountant Dashboard
    elif role == "Manager":
        messagebox.showinfo("Manager Dashboard", "Redirecting to Manager Dashboard...")
        manager_dashboard()
        # Call Manager Dashboard
    elif role == "Driver":
        messagebox.showinfo("Driver Dashboard", "Redirecting to Driver Dashboard...")
        # Call Driver Dashboard
    else:
        messagebox.showerror("Access Denied", "Role not recognized!")





def admin_dashboard(user_type="Admin", login_window=None):
    def update_time():
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_label.config(text=current_time)
        root.after(1000, update_time)

    def logout():
        root.destroy()
        if login_window:
            login_window.deiconify()
    def open_driver_management():
        root.destroy()
        driver_management_gui(user_role="Admin")

    def open_truck_management():
        root.destroy()
        truck_management_gui(user_role="Admin")

    def open_order_management():
        root.destroy()
        order_management_gui(user_role="Admin")

    def open_dispatch_management():
        root.destroy()
        dispatch_management_gui(user_role="Admin")

    def open_account_management():
        root.destroy()
        accounts_management_gui(user_role="Admin")

    def open_report_management():
        root.destroy()
        reports_analytics_gui(user_role="Admin")

    def create_image_button(parent, image_path, text, command):
        # Create a frame to hold the button contents
        frame = ttk.Frame(parent)

        try:
            # Load and resize image
            img = Image.open(image_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            # Create button with image and text
            btn = ttk.Button(frame, text=text, compound=tk.TOP, image=photo, command=command)
            btn.image = photo  # Keep a reference!
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

            return frame
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            # Fallback to text-only button
            btn = ttk.Button(frame, text=text, command=command)
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
            return frame

    root = tk.Tk()
    root.title("Admin Dashboard")
    root.geometry("1024x768")

    # Configure grid weight
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Style configuration
    style = ttk.Style()
    style.configure("Header.TFrame", background="#2c3e50")
    style.configure("Content.TFrame", background="#ecf0f1")

    # Header frame
    header_frame = ttk.Frame(root, style="Header.TFrame")
    header_frame.grid(row=0, column=0, sticky="ew")

    user_label = tk.Label(header_frame,
                          text=f"User Type: {user_type}",
                          font=("Arial", 12, "bold"),
                          bg="#2c3e50",
                          fg="white")
    user_label.pack(side=tk.LEFT, padx=20, pady=10)

    time_label = tk.Label(header_frame,
                          font=("Arial", 12),
                          bg="#2c3e50",
                          fg="white")
    time_label.pack(side=tk.LEFT, padx=20, pady=10)

    logout_btn = ttk.Button(header_frame,
                            text="Logout",
                            command=logout)
    logout_btn.pack(side=tk.RIGHT, padx=20, pady=10)

    # Content frame
    content_frame = ttk.Frame(root, style="Content.TFrame")
    content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)

    # Configure content grid
    for i in range(3):
        content_frame.grid_columnconfigure(i, weight=1, pad=10)
    for i in range(3):
        content_frame.grid_rowconfigure(i, weight=1, pad=10)

    # Dashboard buttons with images
    buttons = [
        ("Manage Drivers", "icons/drivers.png", open_driver_management, 0, 0),
        ("Manage Trucks", "icons/trucks.png", open_truck_management, 0, 1),
        ("Manage Orders", "icons/orders.png", open_order_management, 0, 2),
        ("Manage Accounts", "icons/accounts.png", open_account_management, 1, 0),
        ("Manage Dispatch", "icons/dispatch1.png", open_dispatch_management, 1, 1),
        ("View Reports", "icons/reports.png", open_report_management, 1, 2),
        ("User Management", "icons/users.png", None, 2, 1)
    ]

    for text, icon_path, command, row, col in buttons:
        btn_frame = create_image_button(content_frame, icon_path, text, command)
        btn_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    # Start time update
    update_time()

    root.mainloop()

def manager_dashboard(user_type="Manager", login_window=None):
    def update_time():
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_label.config(text=current_time)
        root.after(1000, update_time)

    def logout():
        root.destroy()
        if login_window:
            login_window.deiconify()
    def open_driver_management():
        root.destroy()
        driver_management_gui(user_role="Manager")

    def open_truck_management():
        root.destroy()
        truck_management_gui(user_role="Manager")

    def open_order_management():
        root.destroy()
        order_management_gui(user_role="Manager")

    def open_dispatch_management():
        root.destroy()
        dispatch_management_gui(user_role="Manager")

    def open_account_management():
        root.destroy()
        accounts_management_gui(user_role="Manager")

    def open_report_management():
        root.destroy()
        reports_analytics_gui(user_role="Manager")

    def create_image_button(parent, image_path, text, command):
        # Create a frame to hold the button contents
        frame = ttk.Frame(parent)

        try:
            # Load and resize image
            img = Image.open(image_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            # Create button with image and text
            btn = ttk.Button(frame, text=text, compound=tk.TOP, image=photo, command=command)
            btn.image = photo  # Keep a reference!
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

            return frame
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            # Fallback to text-only button
            btn = ttk.Button(frame, text=text, command=command)
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
            return frame

    root = tk.Tk()
    root.title("Manager Dashboard")
    root.geometry("1024x768")

    # Configure grid weight
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Style configuration
    style = ttk.Style()
    style.configure("Header.TFrame", background="#2c3e50")
    style.configure("Content.TFrame", background="#ecf0f1")

    # Header frame
    header_frame = ttk.Frame(root, style="Header.TFrame")
    header_frame.grid(row=0, column=0, sticky="ew")

    user_label = tk.Label(header_frame,
                          text=f"User Type: {user_type}",
                          font=("Arial", 12, "bold"),
                          bg="#2c3e50",
                          fg="white")
    user_label.pack(side=tk.LEFT, padx=20, pady=10)

    time_label = tk.Label(header_frame,
                          font=("Arial", 12),
                          bg="#2c3e50",
                          fg="white")
    time_label.pack(side=tk.LEFT, padx=20, pady=10)

    logout_btn = ttk.Button(header_frame,
                            text="Logout",
                            command=logout)
    logout_btn.pack(side=tk.RIGHT, padx=20, pady=10)

    # Content frame
    content_frame = ttk.Frame(root, style="Content.TFrame")
    content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)

    # Configure content grid
    for i in range(3):
        content_frame.grid_columnconfigure(i, weight=1, pad=10)
    for i in range(3):
        content_frame.grid_rowconfigure(i, weight=1, pad=10)

    # Dashboard buttons with images
    buttons = [
        ("Manage Drivers", "icons/drivers.png", open_driver_management, 0, 0),
        ("Manage Trucks", "icons/trucks.png", open_truck_management, 0, 1),
        ("Manage Orders", "icons/orders.png", open_order_management, 0, 2),
        ("Manage Accounts", "icons/accounts.png", open_account_management, 1, 0),
        ("Manage Dispatch", "icons/dispatch1.png", open_dispatch_management, 1, 1),
        ("View Reports", "icons/reports.png", open_report_management, 1, 2),
        # ("User Management", "icons/users.png", None, 2, 1)
    ]

    for text, icon_path, command, row, col in buttons:
        btn_frame = create_image_button(content_frame, icon_path, text, command)
        btn_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    # Start time update
    update_time()

    root.mainloop()




def accountant_dashboard(user_type="Accountant", login_window=None):
    def update_time():
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_label.config(text=current_time)
        root.after(1000, update_time)

    def logout():
        root.destroy()
        if login_window:
            login_window.deiconify()

    def open_account_management():
        root.destroy()
        accounts_management_gui(user_role="Accountant")

    def create_image_button(parent, image_path, text, command):
        # Create a frame to hold the button contents
        frame = ttk.Frame(parent)

        try:
            # Load and resize image
            img = Image.open(image_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            # Create button with image and text
            btn = ttk.Button(frame, text=text, compound=tk.TOP, image=photo, command=command)
            btn.image = photo  # Keep a reference!
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

            return frame
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            # Fallback to text-only button
            btn = ttk.Button(frame, text=text, command=command)
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
            return frame

    root = tk.Tk()
    root.title("Accountant Dashboard")
    root.geometry("1024x768")

    # Configure grid weight
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Style configuration
    style = ttk.Style()
    style.configure("Header.TFrame", background="#2c3e50")
    style.configure("Content.TFrame", background="#ecf0f1")
    style.configure("MenuButton.TButton", padding=10, width=20)

    # Header frame
    header_frame = ttk.Frame(root, style="Header.TFrame")
    header_frame.grid(row=0, column=0, sticky="ew")

    # User type and time display
    user_label = tk.Label(header_frame,
                          text=f"User Type: {user_type}",
                          font=("Arial", 12, "bold"),
                          bg="#2c3e50",
                          fg="white")
    user_label.pack(side=tk.LEFT, padx=20, pady=10)

    time_label = tk.Label(header_frame,
                          font=("Arial", 12),
                          bg="#2c3e50",
                          fg="white")
    time_label.pack(side=tk.LEFT, padx=20, pady=10)

    logout_btn = ttk.Button(header_frame,
                            text="Logout",
                            command=logout)
    logout_btn.pack(side=tk.RIGHT, padx=20, pady=10)

    # Content frame
    content_frame = ttk.Frame(root, style="Content.TFrame")
    content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)

    # Configure content grid
    for i in range(2):
        content_frame.grid_columnconfigure(i, weight=1, pad=10)
    for i in range(2):
        content_frame.grid_rowconfigure(i, weight=1, pad=10)

    # Dashboard buttons with grid layout - limited options for dispatcher
    buttons = [
        ("Manage Accounts","icons/accounts.png", open_account_management, 0, 0),

    ]

    for text, icon_path, command, row, col in buttons:
        btn_frame = create_image_button(content_frame, icon_path, text, command)
        btn_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    # Start time update
    update_time()

    root.mainloop()

def dispatcher_dashboard(user_type="Dispatcher", login_window=None):
    def update_time():
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_label.config(text=current_time)
        root.after(1000, update_time)

    def logout():
        root.destroy()
        if login_window:
            login_window.deiconify()

    def open_dispatch_management():
        root.destroy()
        dispatch_management_gui(user_role="Dispatcher")

    def open_order_management():
        root.destroy()
        order_management_gui(user_role="Dispatcher")
    def create_image_button(parent, image_path, text, command):
        # Create a frame to hold the button contents
        frame = ttk.Frame(parent)

        try:
            # Load and resize image
            img = Image.open(image_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            # Create button with image and text
            btn = ttk.Button(frame, text=text, compound=tk.TOP, image=photo, command=command)
            btn.image = photo  # Keep a reference!
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

            return frame
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            # Fallback to text-only button
            btn = ttk.Button(frame, text=text, command=command)
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
            return frame

    root = tk.Tk()
    root.title("Dispatcher Dashboard")
    root.geometry("1024x768")

    # Configure grid weight
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Style configuration
    style = ttk.Style()
    style.configure("Header.TFrame", background="#2c3e50")
    style.configure("Content.TFrame", background="#ecf0f1")
    style.configure("MenuButton.TButton", padding=10, width=20)

    # Header frame
    header_frame = ttk.Frame(root, style="Header.TFrame")
    header_frame.grid(row=0, column=0, sticky="ew")

    # User type and time display
    user_label = tk.Label(header_frame,
                          text=f"User Type: {user_type}",
                          font=("Arial", 12, "bold"),
                          bg="#2c3e50",
                          fg="white")
    user_label.pack(side=tk.LEFT, padx=20, pady=10)

    time_label = tk.Label(header_frame,
                          font=("Arial", 12),
                          bg="#2c3e50",
                          fg="white")
    time_label.pack(side=tk.LEFT, padx=20, pady=10)

    logout_btn = ttk.Button(header_frame,
                            text="Logout",
                            command=logout)
    logout_btn.pack(side=tk.RIGHT, padx=20, pady=10)

    # Content frame
    content_frame = ttk.Frame(root, style="Content.TFrame")
    content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=20)

    # Configure content grid
    for i in range(2):
        content_frame.grid_columnconfigure(i, weight=1, pad=10)
    for i in range(2):
        content_frame.grid_rowconfigure(i, weight=1, pad=10)

    # Dashboard buttons with grid layout - limited options for dispatcher
    buttons = [
        ("Manage Dispatch","icons/dispatch1.png", open_dispatch_management, 0, 0),
        ("Manage Orders","icons/orders.png", open_order_management, 0, 1),
        ("View Active Drivers", None,None, 1, 0),
        ("View Reports", None, 1,None, 1)
    ]

    for text, icon_path, command, row, col in buttons:
        btn_frame = create_image_button(content_frame, icon_path, text, command)
        btn_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    # Start time update
    update_time()

    root.mainloop()

def driver_management_gui(user_role=""):
    def go_back():
        """Handle the back button based on user role."""
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()

    def load_drivers():
        """Load driver data from the Excel sheet."""
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Drivers"]
            drivers = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
            return drivers
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load drivers: {e}")
            return []
    def save_driver(driver_data):
        """Save new driver to the Excel sheet."""
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Drivers"]
            sheet.append(driver_data)
            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Driver added successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save driver: {e}")


    def update_driver(row_num, updated_data):
        """Update existing driver data."""
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Drivers"]

            for col_num, value in enumerate(updated_data, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)

            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Driver updated successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update driver: {e}")

    def update_trip_count(row_num):
        """Increment trip count when trip status changes to 'On Trip'."""
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Drivers"]
            current_count = sheet.cell(row=row_num, column=9).value  # Assuming column 9 is Trip Count
            sheet.cell(row=row_num, column=9, value=(current_count or 0) + 1)
            workbook.save(WORKBOOK_PATH)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update trip count: {e}")

    def delete_driver(row_num):
        """Delete a driver from the Excel sheet."""
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Drivers"]
            sheet.delete_rows(row_num)
            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Driver deleted successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete driver: {e}")

    def check_license_expiry():
        """Check for upcoming license expirations."""
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Drivers"]
            today = datetime.now()
            notifications = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                license_expiry = datetime.strptime(row[3], "%Y-%m-%d")  # Assuming column 3 is License Expiry
                days_to_expiry = (license_expiry - today).days
                if days_to_expiry <= 30:
                    notifications.append(f"Driver {row[1]} (ID: {row[0]}) - License expires in {days_to_expiry} days")

            if notifications:
                messagebox.showwarning("License Expiry Notifications", "\n".join(notifications))
            else:
                messagebox.showinfo("Compliance Check", "No upcoming license expirations.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to check license expiry: {e}")

    def load_data_into_table(filtered_data=None):
        """Load all driver data or filtered data into the table."""
        driver_table.delete(*driver_table.get_children())  # Clear existing data
        drivers = filtered_data if filtered_data else load_drivers()
        for idx, driver in enumerate(drivers, start=1):  # Row indexing starts at 1 for clarity
            driver_table.insert("", "end", values=(idx, *driver))

    def search_drivers():
        """Filter and display drivers based on search criteria."""
        criteria = search_criteria.get()
        search_value = search_entry.get().lower()
        if not search_value:
            messagebox.showerror("Error", "Search field cannot be empty!")
            return

        criteria_map = {
            "ID": 0,
            "Name": 1,
            "Contact": 5,
            "License": 3,
        }

        column_index = criteria_map.get(criteria, 1)
        drivers = load_drivers()

        # Filter drivers based on the search value
        filtered_drivers = [
            driver for driver in drivers
            if search_value in str(driver[column_index]).lower()
        ]

        if not filtered_drivers:
            messagebox.showinfo("No Results", "No matching drivers found.")
        load_data_into_table(filtered_drivers)

    def add_driver():
        """Add a new driver to the database."""
        driver_data = [
            id_entry.get(),
            name_entry.get(),
            cnic_entry.get(),
            license_entry.get(),
            address_entry.get(),
            contact_entry.get(),
            salary_entry.get(),
            salary_status_combo.get(),
            joining_date_entry.get(),
            resigning_date_entry.get(),
            trip_count_entry.get(),
            trip_status_combo.get()
        ]

        # Validate ID uniqueness
        drivers = load_drivers()
        existing_ids = [driver[0] for driver in drivers]
        if driver_data[0] in existing_ids:
            messagebox.showerror("Error", "Driver ID already exists!")
            return

        save_driver(driver_data)
        load_data_into_table()

    def update_selected_driver():
        """Update the selected driver's details."""
        selected_item = driver_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No driver selected!")
            return

        selected_row_index = driver_table.item(selected_item)["values"][0]
        updated_data = [
            id_entry.get(),
            name_entry.get(),
            cnic_entry.get(),
            license_entry.get(),
            address_entry.get(),
            contact_entry.get(),
            salary_entry.get(),
            salary_status_combo.get(),
            joining_date_entry.get(),
            resigning_date_entry.get(),
            trip_count_entry.get(),
            trip_status_combo.get()
        ]
        if trip_status_combo.get() == "On Trip":
            update_trip_count(selected_row_index + 1)

        update_driver(selected_row_index + 1, updated_data)
        load_data_into_table()

    def delete_selected_driver():
        """Delete the selected driver."""
        selected_item = driver_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No driver selected!")
            return

        selected_row_index = driver_table.item(selected_item)["values"][0]
        delete_driver(selected_row_index)
        load_data_into_table()

    def on_driver_select(event):
        """Populate form fields when a driver is selected."""
        selected_item = driver_table.selection()
        if selected_item:
            selected_driver = driver_table.item(selected_item)["values"]
            id_entry.delete(0, tk.END)
            id_entry.insert(0, selected_driver[1])
            name_entry.delete(0, tk.END)
            name_entry.insert(0, selected_driver[2])
            cnic_entry.delete(0, tk.END)
            cnic_entry.insert(0, selected_driver[3])
            license_entry.delete(0, tk.END)
            license_entry.insert(0, selected_driver[4])
            address_entry.delete(0, tk.END)
            address_entry.insert(0, selected_driver[5])
            contact_entry.delete(0, tk.END)
            contact_entry.insert(0, selected_driver[6])
            salary_entry.delete(0, tk.END)
            salary_entry.insert(0, selected_driver[7])
            trip_status_combo.set(selected_driver[8])

    # Create themed root window
    root = ThemedTk(theme="arc")
    root.title("Driver Management System")
    root.geometry("1200x800")
    root.configure(bg='#f0f0f0')

    # Main container
    main_container = ttk.Frame(root, padding="10")
    main_container.pack(fill=tk.BOTH, expand=True)

    # Header frame
    header_frame = ttk.Frame(main_container)
    header_frame.pack(fill=tk.X, pady=(0, 10))
    ttk.Button(header_frame, text="← Back", command=go_back).pack(side=tk.LEFT)
    ttk.Label(header_frame, text="Driver Management System", font=('Helvetica', 16, 'bold')).pack(side=tk.LEFT, padx=20)

    # Search frame
    search_frame = ttk.LabelFrame(main_container, text="Search", padding="5")
    search_frame.pack(fill=tk.X, pady=(0, 10))
    search_criteria = ttk.Combobox(search_frame, values=["ID", "Name", "Contact", "License"], width=15)
    search_criteria.set("Name")
    search_criteria.pack(side=tk.LEFT, padx=5)
    search_entry = ttk.Entry(search_frame, width=30)
    search_entry.pack(side=tk.LEFT, padx=5)
    ttk.Button(search_frame, text="Search", command=search_drivers).pack(side=tk.LEFT, padx=5)

    # Form frame
    form_frame = ttk.LabelFrame(main_container, text="Driver Information", padding="10")
    form_frame.pack(fill=tk.X, pady=(0, 10))
    form_fields = [
        ("ID", id_entry := tk.Entry(form_frame)),
        ("Name", name_entry := tk.Entry(form_frame)),
        ("CNIC", cnic_entry := tk.Entry(form_frame)),
        ("License Expiry", license_entry := tk.Entry(form_frame)),
        ("Address", address_entry := tk.Entry(form_frame)),
        ("Contact", contact_entry := tk.Entry(form_frame)),
        ("Salary", salary_entry := tk.Entry(form_frame)),
        ("Salary Status", salary_status_combo := ttk.Combobox(form_frame, values=["Paid", "Unpaid"])),
        ("Joining Date", joining_date_entry := tk.Entry(form_frame)),
        ("Resigning Date", resigning_date_entry := tk.Entry(form_frame)),
        ("Trip Count", trip_count_entry := tk.Entry(form_frame)),
        ("Trip Status", trip_status_combo := ttk.Combobox(form_frame, values=["Available", "On Trip", "Off Duty"]))
    ]
    for i, (label_text, entry_widget) in enumerate(form_fields):
        ttk.Label(form_frame, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky='e')
        entry_widget.grid(row=i, column=1, padx=5, pady=5, sticky='w')

    # Buttons frame
    buttons_frame = ttk.Frame(main_container)
    buttons_frame.pack(fill=tk.X, pady=(0, 10))
    buttons = [
        ("Add Driver", add_driver),
        ("Update Driver", update_selected_driver),
        ("Delete Driver", delete_selected_driver),
        ("Check License Expiry", check_license_expiry)
    ]
    for text, command in buttons:
        ttk.Button(buttons_frame, text=text, command=command).pack(side=tk.LEFT, padx=5)

    # Table frame
    table_frame = ttk.Frame(main_container)
    table_frame.pack(fill=tk.BOTH, expand=True)
    columns = ("Index", "ID", "Name", "CNIC", "License Expiry", "Address", "Contact", "Salary", "Salary Status",
               "Joining Date", "Resigning Date", "Trip Count", "Trip Status")
    driver_table = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
    for col in columns:
        driver_table.heading(col, text=col)
        driver_table.column(col, width=100)


    driver_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    driver_table.bind("<<TreeviewSelect>>", on_driver_select)

    # Scrollbars
    scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=driver_table.yview)
    driver_table.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Load initial data
    load_data_into_table()
    root.mainloop()


def truck_management_gui(user_role=""):
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()

    def load_trucks():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Trucks"]
            trucks = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                trucks.append(row)
            return trucks
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load trucks: {e}")
            return []

    def save_truck(truck_data):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Trucks"]

            # Check for duplicate truck ID
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == truck_data[0]:
                    messagebox.showerror("Error", "Truck ID already exists!")
                    return False

            sheet.append(truck_data)
            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Truck added successfully!")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save truck: {e}")
            return False

    def search_trucks():
        search_term = search_entry.get().lower()
        search_by = search_criteria.get()

        if not search_term:
            load_data_into_table()
            return

        trucks = load_trucks()
        filtered_trucks = []

        criteria_map = {
            "ID": 0,
            "Model": 1,
            "Status": 2,
            "Permit": 3
        }

        col_idx = criteria_map.get(search_by, 0)

        for truck in trucks:
            if str(truck[col_idx]).lower().find(search_term) != -1:
                filtered_trucks.append(truck)

        load_data_into_table(filtered_trucks)

    root = ThemedTk(theme="arc")  # Using themed Tk for better appearance
    root.title("Truck Management System")
    root.geometry("1200x800")

    # Main container with padding
    main_container = ttk.Frame(root, padding="10")
    main_container.pack(fill=tk.BOTH, expand=True)

    # Header frame
    header_frame = ttk.Frame(main_container)
    header_frame.pack(fill=tk.X, pady=(0, 10))

    ttk.Button(header_frame, text="← Back", command=go_back).pack(side=tk.LEFT)
    ttk.Label(header_frame, text="Truck Management System",
              font=('Helvetica', 16, 'bold')).pack(side=tk.LEFT, padx=20)

    # Search frame
    search_frame = ttk.LabelFrame(main_container, text="Search", padding="5")
    search_frame.pack(fill=tk.X, pady=(0, 10))

    search_criteria = ttk.Combobox(search_frame,
                                   values=["ID", "Model", "Status", "Permit"],
                                   width=15)
    search_criteria.set("ID")
    search_criteria.pack(side=tk.LEFT, padx=5)

    search_entry = ttk.Entry(search_frame, width=30)
    search_entry.pack(side=tk.LEFT, padx=5)

    ttk.Button(search_frame, text="Search", command=search_trucks).pack(side=tk.LEFT, padx=5)

    # Input form frame
    form_frame = ttk.LabelFrame(main_container, text="Truck Information", padding="10")
    form_frame.pack(fill=tk.X, pady=(0, 10))

    # Grid layout for form fields
    form_fields = [
        ("Truck ID:", id_entry := ttk.Entry(form_frame)),
        ("Model:", model_entry := ttk.Entry(form_frame)),
        ("Status:", status_combo := ttk.Combobox(form_frame,
                                                 values=["Operational", "Under Maintenance", "Retired"])),
        ("Permit Number:", permit_entry := ttk.Entry(form_frame)),
        ("Maintenance Schedule:", maintenance_entry := ttk.Entry(form_frame))
    ]

    for i, (label, widget) in enumerate(form_fields):
        ttk.Label(form_frame, text=label).grid(row=i // 2, column=i % 2 * 2, padx=5, pady=5, sticky='e')
        widget.grid(row=i // 2, column=i % 2 * 2 + 1, padx=5, pady=5, sticky='w')

    # Buttons frame
    buttons_frame = ttk.Frame(main_container)
    buttons_frame.pack(fill=tk.X, pady=(0, 10))

    def add_truck():
        truck_data = [
            id_entry.get(),
            model_entry.get(),
            status_combo.get(),
            permit_entry.get(),
            maintenance_entry.get()
        ]
        if save_truck(truck_data):
            load_data_into_table()
            # Clear form fields
            for widget in (id_entry, model_entry, status_combo, permit_entry, maintenance_entry):
                widget.set('') if isinstance(widget, ttk.Combobox) else widget.delete(0, tk.END)

    for text, command in [
        ("Add Truck", add_truck),
        ("Update Truck", lambda: update_selected_truck()),
        ("Delete Truck", lambda: delete_selected_truck()),
        ("Clear Form", lambda: clear_form())
    ]:
        ttk.Button(buttons_frame, text=text, command=command).pack(side=tk.LEFT, padx=5)

    # Table frame with scrollbar
    table_frame = ttk.Frame(main_container)
    table_frame.pack(fill=tk.BOTH, expand=True)

    # Create scrollbar
    scrollbar = ttk.Scrollbar(table_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Create treeview with scrollbar
    columns = ("ID", "Model", "Status", "Permit", "Maintenance Schedule")
    truck_table = ttk.Treeview(table_frame, columns=columns, show="headings",
                               height=15, yscrollcommand=scrollbar.set)

    for col in columns:
        truck_table.heading(col, text=col)
        truck_table.column(col, width=150)

    truck_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=truck_table.yview)

    # Bind table selection event
    def on_select(event):
        selected = truck_table.selection()
        if selected:
            values = truck_table.item(selected[0])['values']
            id_entry.delete(0, tk.END)
            id_entry.insert(0, values[0])
            model_entry.delete(0, tk.END)
            model_entry.insert(0, values[1])
            status_combo.set(values[2])
            permit_entry.delete(0, tk.END)
            permit_entry.insert(0, values[3])
            maintenance_entry.delete(0, tk.END)
            maintenance_entry.insert(0, values[4])

    truck_table.bind('<<TreeviewSelect>>', on_select)

    def clear_form():
        # Clear all entry widgets and reset comboboxes in the form
        for widget in (id_entry, model_entry, status_combo, permit_entry, maintenance_entry):
            if isinstance(widget, ttk.Combobox):
                widget.set('')
            else:
                widget.delete(0, tk.END)
    def load_data_into_table(data=None):
        for item in truck_table.get_children():
            truck_table.delete(item)
        trucks = data if data is not None else load_trucks()
        for truck in trucks:
            truck_table.insert("", "end", values=truck)
    def update_truck(row_num, updated_data):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Trucks"]

            for col_num, value in enumerate(updated_data, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)

            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Truck updated successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update truck: {e}")

    def delete_truck(row_num):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Trucks"]
            sheet.delete_rows(row_num)
            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Truck deleted successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete truck: {e}")


    def update_selected_truck():
        selected_item = truck_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No truck selected!")
            return

        selected_row = truck_table.item(selected_item)["values"]
        updated_data = [
            id_entry.get(),
            model_entry.get(),
            status_combo.get(),
            maintenance_entry.get()
        ]
        update_truck(int(selected_row[0]) + 1, updated_data)
        load_data_into_table()

    def delete_selected_truck():
        selected_item = truck_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No truck selected!")
            return

        selected_row = truck_table.item(selected_item)["values"]
        delete_truck(int(selected_row[0]) + 1)
        load_data_into_table()
    root.mainloop()



def order_management_gui(user_role=""):
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()
        elif user_role == "Dispatcher":
            dispatcher_dashboard()

    def load_orders():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Orders"]
            orders = []
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                orders.append((i, *row))  # Include row index for tracking
            return orders
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load orders: {e}")
            return []

    def save_order(order_data):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Orders"]

            # Check if order ID already exists
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == order_data[0]:  # Compare Order IDs
                    messagebox.showerror("Error", "Order ID already exists!")
                    return False

            # Calculate remaining amount
            total_amount = float(order_data[7])  # Changed index to match the order_data list
            paid_amount = float(order_data[8])  # Changed index to match the order_data list
            remaining = total_amount - paid_amount

            # Append all data including remaining amount
            sheet.append(order_data)
            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Order added successfully!")
            return True
        except ValueError as ve:
            messagebox.showerror("Error", "Please enter valid numbers for amounts")
            return False
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save order: {e}")
            return False

    def delete_order(row_id):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Orders"]
            sheet.delete_rows(row_id)
            workbook.save(WORKBOOK_PATH)
            load_data_into_table()
            messagebox.showinfo("Success", "Order deleted successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete order: {e}")

    def update_order(row_id, updated_data):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Orders"]

            # Update fields in the row
            for col, value in enumerate(updated_data, start=1):
                sheet.cell(row=row_id, column=col, value=value)

            # Calculate and update the remaining amount
            total_amount = float(updated_data[7])  # Total Amount is 7th index (0-based)
            paid_amount = float(updated_data[8])  # Paid Amount is 8th index (0-based)
            remaining = total_amount - paid_amount
            sheet.cell(row=row_id, column=len(updated_data) + 1, value=remaining)

            workbook.save(WORKBOOK_PATH)
            load_data_into_table()  # Refresh data in the Treeview
            messagebox.showinfo("Success", "Order updated successfully!")
        except ValueError as ve:
            messagebox.showerror("Input Error", f"Invalid numeric value: {ve}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update order: {e}")

    def add_order():
        try:
            # Validate required fields
            if not all([order_id_entry.get(), customer_entry.get(),
                        total_amount_entry.get(), paid_amount_entry.get()]):
                raise ValueError("Please fill all required fields")

            total_amount = float(total_amount_entry.get())
            paid_amount = float(paid_amount_entry.get())
            remaining_amount=float(total_amount-paid_amount)
            # Validate amounts
            if total_amount <= 0:
                raise ValueError("Total amount must be greater than 0")
            if paid_amount < 0:
                raise ValueError("Paid amount cannot be negative")
            if paid_amount > total_amount:
                raise ValueError("Paid amount cannot exceed total amount!")

            order_data = [
                order_id_entry.get(),
                customer_entry.get(),
                contact_entry.get(),
                pickup_entry.get(),
                destination_entry.get(),
                distance_entry.get(),
                status_combo.get() or "Pending",  # Default value if empty
                total_amount,
                paid_amount,
                payment_combo.get() or "Pending",  # Default value if empty
                remaining_amount
            ]

            if save_order(order_data):
                load_data_into_table()
                clear_form()
        except ValueError as ve:
            messagebox.showerror("Input Error", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add order: {e}")

    def delete_selected_order():
        selected = order_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select an order to delete.")
            return

        row_id = int(order_table.item(selected[0])["values"][0])
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this order?"):
            delete_order(row_id)

    def update_selected_order():
        selected = order_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select an order to update.")
            return

        try:
            total_amount = float(total_amount_entry.get())
            paid_amount = float(paid_amount_entry.get())
            if paid_amount > total_amount:
                raise ValueError("Paid amount cannot exceed total amount!")

            updated_data = [
                order_id_entry.get(),
                customer_entry.get(),
                contact_entry.get(),
                pickup_entry.get(),
                destination_entry.get(),
                distance_entry.get(),
                status_combo.get(),
                total_amount,
                paid_amount,
                payment_combo.get()
            ]

            row_id = int(order_table.item(selected[0])["values"][0])
            update_order(row_id, updated_data)
        except ValueError as ve:
            messagebox.showerror("Input Error", f"Invalid input: {ve}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update order: {e}")

    def search_orders():
        search_term = search_entry.get().lower()
        search_by = search_criteria.get()

        if not search_term:
            load_data_into_table()
            return

        orders = load_orders()
        criteria_indices = {
            "Order ID": 1,
            "Customer Name": 2,
            "Destination": 5,
            "Status": 6,
            "Payment Status": 10
        }

        idx = criteria_indices.get(search_by, 1)
        filtered_orders = [order for order in orders if search_term in str(order[idx]).lower()]
        load_data_into_table(filtered_orders)

    def load_data_into_table(data=None):
        order_table.delete(*order_table.get_children())
        orders = data if data else load_orders()
        for order in orders:
            try:
                # Calculate remaining amount properly
                total_amount = float(order[8])  # Total Amount column
                paid_amount = float(order[9])  # Paid Amount column
                remaining = total_amount - paid_amount

                # Format remaining amount to 2 decimal places
                remaining = "{:.2f}".format(remaining)

                # Insert all values including remaining
                order_table.insert('', 'end', values=(*order, remaining))
            except (ValueError, IndexError) as e:
                print(f"Error processing order: {e}")
                continue

    def clear_form():
        for widget in (order_id_entry, customer_entry, contact_entry,
                       pickup_entry, destination_entry, distance_entry,
                       total_amount_entry, paid_amount_entry):
            widget.delete(0, tk.END)
        status_combo.set("")
        payment_combo.set("")

    # Root window
    root = ThemedTk(theme="arc")
    root.title("Order Management System")
    root.geometry("1200x800")

    # Header
    header = ttk.Frame(root)
    header.pack(fill=tk.X, pady=10)
    ttk.Button(header, text="← Back", command=go_back).pack(side=tk.LEFT)
    ttk.Label(header, text="Order Management System", font=("Helvetica", 16, "bold")).pack(side=tk.LEFT, padx=20)

    # Search
    search_frame = ttk.LabelFrame(root, text="Search Orders")
    search_frame.pack(fill=tk.X, padx=10, pady=10)

    search_criteria = ttk.Combobox(search_frame, values=["Order ID", "Customer Name", "Destination", "Status", "Payment Status"], width=15)
    search_criteria.set("Order ID")
    search_criteria.pack(side=tk.LEFT, padx=5)

    search_entry = ttk.Entry(search_frame)
    search_entry.pack(side=tk.LEFT, padx=5)
    ttk.Button(search_frame, text="Search", command=search_orders).pack(side=tk.LEFT, padx=5)

    # Form
    form_frame = ttk.LabelFrame(root, text="Order Details")
    form_frame.pack(fill=tk.X, padx=10, pady=10)

    fields = [
        ("Order ID", order_id_entry := ttk.Entry(form_frame)),
        ("Customer Name", customer_entry := ttk.Entry(form_frame)),
        ("Contact", contact_entry := ttk.Entry(form_frame)),
        ("Pickup Location", pickup_entry := ttk.Entry(form_frame)),
        ("Destination", destination_entry := ttk.Entry(form_frame)),
        ("Distance", distance_entry := ttk.Entry(form_frame)),
        ("Status", status_combo := ttk.Combobox(form_frame, values=["Pending", "In Transit", "Delivered", "Cancelled"])),
        ("Total Amount", total_amount_entry := ttk.Entry(form_frame)),
        ("Paid Amount", paid_amount_entry := ttk.Entry(form_frame)),
        ("Payment Status", payment_combo := ttk.Combobox(form_frame, values=["Pending", "Partial", "Completed"]))
    ]

    for i, (label, widget) in enumerate(fields):
        ttk.Label(form_frame, text=label).grid(row=i // 2, column=i % 2 * 2, padx=5, pady=5, sticky='e')
        widget.grid(row=i // 2, column=i % 2 * 2 + 1, padx=5, pady=5, sticky='w')

    # Buttons
    buttons_frame = ttk.Frame(root)
    buttons_frame.pack(fill=tk.X, padx=10, pady=10)

    ttk.Button(buttons_frame, text="Add Order", command=add_order).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Update Order", command=update_selected_order).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Delete Order", command=delete_selected_order).pack(side=tk.LEFT, padx=5)

    # Table
    table_frame = ttk.Frame(root)
    table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    table_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)
    table_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    order_columns = [
        "Row", "Order ID", "Customer Name", "Contact",
        "Pickup Location", "Destination", "Distance (km)",
        "Status", "Total Amount", "Paid Amount",
        "Payment Status", "Remaining"
    ]

    # Add horizontal scrollbar
    h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
    h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    order_table = ttk.Treeview(table_frame, columns=order_columns, show="headings",
                               yscrollcommand=table_scrollbar.set,
                               xscrollcommand=h_scrollbar.set, height=15)

    # Configure both scrollbars
    table_scrollbar.config(command=order_table.yview)
    h_scrollbar.config(command=order_table.xview)

    # Adjust column widths and make them resizable
    for col in order_columns:
        order_table.heading(col, text=col)
        order_table.column(col, width=100, minwidth=50)  # Set minimum width

    order_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    table_scrollbar.config(command=order_table.yview)

    # Bind table events
    def on_select_order(event):
        selected_item = order_table.selection()
        if selected_item:
            values = order_table.item(selected_item[0])['values']
            order_id_entry.delete(0, tk.END)
            order_id_entry.insert(0, values[1])
            customer_entry.delete(0, tk.END)
            customer_entry.insert(0, values[2])
            contact_entry.delete(0, tk.END)
            contact_entry.insert(0, values[3])
            pickup_entry.delete(0, tk.END)
            pickup_entry.insert(0, values[4])
            destination_entry.delete(0, tk.END)
            destination_entry.insert(0, values[5])
            distance_entry.delete(0, tk.END)
            distance_entry.insert(0, values[6])
            status_combo.set(values[7])
            total_amount_entry.delete(0, tk.END)
            total_amount_entry.insert(0, values[8])
            paid_amount_entry.delete(0, tk.END)
            paid_amount_entry.insert(0, values[9])
            payment_combo.set(values[10])

    order_table.bind("<<TreeviewSelect>>", on_select_order)

    # Load initial data
    load_data_into_table()

    root.mainloop()



def dispatch_management_gui(user_role=""):
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role=="Manager":
            manager_dashboard()
        elif user_role == "Dispatcher":
            dispatcher_dashboard()

    def load_dispatch_data():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Dispatch"]
            dispatches = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                dispatches.append(row)
            return dispatches
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load dispatch data: {e}")
            return []

    def save_dispatch(dispatch_data):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Dispatch"]
            sheet.append(dispatch_data)
            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Dispatch assignment saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save dispatch: {e}")

    def get_available_drivers():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Drivers"]
            available_drivers = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[7] == "Available":  # Assuming column 7 is Trip Status
                    available_drivers.append(f"{row[0]} - {row[1]}")  # ID - Name
            return available_drivers
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load available drivers: {e}")
            return []

    def get_available_trucks():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Trucks"]
            available_trucks = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[2] == "Operational":  # Assuming column 2 is Status
                    available_trucks.append(f"{row[0]} - {row[1]}")  # ID - Model
            return available_trucks
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load available trucks: {e}")
            return []

    def get_pending_orders():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Orders"]
            pending_orders = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[5] == "Pending":  # Assuming column 5 is Status
                    pending_orders.append(f"{row[0]} - {row[1]}")  # Order ID - Customer
            return pending_orders
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load pending orders: {e}")
            return []

    def assign_dispatch():
        if not order_combo.get() or not driver_combo.get() or not truck_combo.get():
            messagebox.showerror("Error", "Please select all required fields!")
            return

        dispatch_data = [
            order_combo.get().split(" - ")[0],  # Order ID
            driver_combo.get().split(" - ")[0],  # Driver ID
            truck_combo.get().split(" - ")[0],  # Truck ID
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),  # Dispatch Time
            "In Transit",  # Status
            estimated_entry.get()  # Estimated Delivery Time
        ]

        save_dispatch(dispatch_data)
        update_order_status(dispatch_data[0], "In Transit")
        update_driver_status(dispatch_data[1], "On Trip")
        load_data_into_table()
        refresh_combos()

    def update_order_status(order_id, new_status):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Orders"]

            # Find and update the order status
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == order_id:  # First column is Order ID
                    row[5].value = new_status  # Column 5 is Status
                    break

            workbook.save(WORKBOOK_PATH)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update order status: {e}")

    def update_driver_status(driver_id, new_status):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Drivers"]

            # Find and update the driver's trip status
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == driver_id:  # First column is Driver ID
                    row[7].value = new_status  # Column 7 is Trip Status
                    break

            workbook.save(WORKBOOK_PATH)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update driver status: {e}")

    root = tk.Tk()
    root.title("Dispatch Management")
    root.geometry("1000x600")
    back_button = tk.Button(root, text="Back", command=go_back)
    back_button.pack(anchor='nw', padx=10, pady=5)
    # Assignment Form
    form_frame = tk.Frame(root)
    form_frame.pack(pady=10)
    tk.Label(form_frame, text="Select Order:").grid(row=0, column=0, padx=5)
    order_combo = ttk.Combobox(form_frame, values=get_pending_orders(), width=30)
    order_combo.grid(row=0, column=1, padx=5)

    tk.Label(form_frame, text="Assign Driver:").grid(row=1, column=0, padx=5)
    driver_combo = ttk.Combobox(form_frame, values=get_available_drivers(), width=30)
    driver_combo.grid(row=1, column=1, padx=5)

    tk.Label(form_frame, text="Assign Truck:").grid(row=2, column=0, padx=5)
    truck_combo = ttk.Combobox(form_frame, values=get_available_trucks(), width=30)
    truck_combo.grid(row=2, column=1, padx=5)

    tk.Label(form_frame, text="Estimated Delivery:").grid(row=3, column=0, padx=5)
    estimated_entry = tk.Entry(form_frame)
    estimated_entry.grid(row=3, column=1, padx=5)

    tk.Button(form_frame, text="Assign Dispatch", command=assign_dispatch).grid(row=4, column=0, columnspan=2, pady=10)

    # Dispatch Table
    table_frame = tk.Frame(root)
    table_frame.pack(pady=10)

    columns = ("Order ID", "Driver ID", "Truck ID", "Dispatch Time", "Status", "Est. Delivery")
    dispatch_table = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)

    for col in columns:
        dispatch_table.heading(col, text=col)
        dispatch_table.column(col, width=150)

    dispatch_table.pack()

    def load_data_into_table():
        for item in dispatch_table.get_children():
            dispatch_table.delete(item)
        dispatches = load_dispatch_data()
        for dispatch in dispatches:
            dispatch_table.insert("", "end", values=dispatch)

    def refresh_combos():
        order_combo['values'] = get_pending_orders()
        driver_combo['values'] = get_available_drivers()
        truck_combo['values'] = get_available_trucks()

    load_data_into_table()
    root.mainloop()
def accounts_management_gui(user_role=""):
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role=="Manager":
            manager_dashboard()
        elif user_role == "Accountant":
            accountant_dashboard()
    def load_financial_data():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Financials"]
            transactions = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                transactions.append(row)
            return transactions
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load")
            return []

    def save_transaction(transaction_data):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            sheet = workbook["Financials"]
            sheet.append(transaction_data)
            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Transaction recorded successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save transaction: {e}")

    def generate_invoice(order_id):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            orders_sheet = workbook["Orders"]

            # Find order details
            order_details = None
            for row in orders_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == order_id:
                    order_details = row
                    break

            if order_details:
                # Generate invoice logic here
                invoice_data = {
                    "order_id": order_id,
                    "customer": order_details[1],
                    "amount": order_details[6],
                    "date": datetime.datetime.now().strftime("%Y-%m-%d"),
                    "status": order_details[7]
                }
                # Save invoice to Excel and/or generate PDF
                save_invoice(invoice_data)
            else:
                messagebox.showerror("Error", "Order not found!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate invoice: {e}")

    def save_invoice(invoice_data):
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            if "Invoices" not in workbook.sheetnames:
                invoice_sheet = workbook.create_sheet("Invoices")
                headers = ["Order ID", "Customer", "Amount", "Date", "Status"]
                invoice_sheet.append(headers)

            sheet = workbook["Invoices"]
            invoice_row = [
                invoice_data["order_id"],
                invoice_data["customer"],
                invoice_data["amount"],
                invoice_data["date"],
                invoice_data["status"]
            ]
            sheet.append(invoice_row)
            workbook.save(WORKBOOK_PATH)
            messagebox.showinfo("Success", "Invoice saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save invoice: {e}")

    root = tk.Tk()
    root.title("Accounts & Financial Management")
    root.geometry("1000x600")

    back_button = tk.Button(root, text="Back", command=go_back)
    back_button.pack(anchor='nw', padx=10, pady=5)
    # Create notebook for tabbed interface
    notebook = ttk.Notebook(root)
    notebook.pack(pady=10, expand=True)
    # Transactions tab
    transactions_frame = ttk.Frame(notebook)
    notebook.add(transactions_frame, text="Transactions")

    # Transaction form
    form_frame = tk.Frame(transactions_frame)
    form_frame.pack(pady=10)

    tk.Label(form_frame, text="Transaction Type").grid(row=0, column=0, padx=5)
    transaction_type = ttk.Combobox(form_frame,
                                   values=["Order Payment", "Salary Payment",
                                          "Fuel Expense", "Maintenance Expense"])
    transaction_type.grid(row=0, column=1, padx=5)

    tk.Label(form_frame, text="Amount").grid(row=0, column=2, padx=5)
    amount_entry = tk.Entry(form_frame)
    amount_entry.grid(row=0, column=3, padx=5)

    tk.Label(form_frame, text="Reference ID").grid(row=1, column=0, padx=5)
    reference_entry = tk.Entry(form_frame)
    reference_entry.grid(row=1, column=1, padx=5)

    tk.Label(form_frame, text="Payment Mode").grid(row=1, column=2, padx=5)
    payment_mode = ttk.Combobox(form_frame, values=["Cash", "Credit", "Bank Transfer"])
    payment_mode.grid(row=1, column=3, padx=5)

    def record_transaction():
        transaction_data = [
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            transaction_type.get(),
            amount_entry.get(),
            reference_entry.get(),
            payment_mode.get()
        ]
        save_transaction(transaction_data)
        load_transactions_table()

    tk.Button(form_frame, text="Record Transaction",
              command=record_transaction).grid(row=2, column=0, columnspan=4, pady=10)

    # Transactions table
    table_frame = tk.Frame(transactions_frame)
    table_frame.pack(pady=10)

    columns = ("Date", "Type", "Amount", "Reference", "Payment Mode")
    transactions_table = ttk.Treeview(table_frame, columns=columns,
                                    show="headings", height=10)

    for col in columns:
        transactions_table.heading(col, text=col)
        transactions_table.column(col, width=150)

    transactions_table.pack()

    def load_transactions_table():
        for item in transactions_table.get_children():
            transactions_table.delete(item)
        transactions = load_financial_data()
        for transaction in transactions:
            transactions_table.insert("", "end", values=transaction)

    # Invoicing tab
    invoicing_frame = ttk.Frame(notebook)
    notebook.add(invoicing_frame, text="Invoicing")

    # Add invoicing interface here
    # Similar structure to transactions but for invoice generation

    def initialize_financial_sheet():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            if "Financials" not in workbook.sheetnames:
                financial_sheet = workbook.create_sheet("Financials")
                headers = ["Date", "Type", "Amount", "Reference", "Payment Mode"]
                financial_sheet.append(headers)
                workbook.save(WORKBOOK_PATH)
                messagebox.showinfo("Success", "Financial sheet created successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to initialize financial sheet: {e}")

    # Initialize the financial sheet if it doesn't exist
    initialize_financial_sheet()

    load_transactions_table()
    root.mainloop()
def reports_analytics_gui(user_role=""):
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()
    def generate_fleet_report():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            trucks_sheet = workbook["Trucks"]
            dispatch_sheet = workbook["Dispatch"]

            # Calculate fleet metrics
            total_trucks = 0
            operational = 0
            maintenance = 0
            dispatched = set()

            for row in trucks_sheet.iter_rows(min_row=2, values_only=True):
                total_trucks += 1
                if row[2] == "Operational":
                    operational += 1
                elif row[2] == "Under Maintenance":
                    maintenance += 1

            for row in dispatch_sheet.iter_rows(min_row=2, values_only=True):
                if row[4] == "In Transit":  # Status column
                    dispatched.add(row[2])  # Truck ID

            # Update fleet metrics display
            fleet_metrics.delete(1.0, tk.END)
            fleet_metrics.insert(tk.END,
                f"Total Trucks: {total_trucks}\n"
                f"Operational: {operational}\n"
                f"Under Maintenance: {maintenance}\n"
                f"Currently Dispatched: {len(dispatched)}\n"
                f"Fleet Utilization: {(len(dispatched)/operational)*100:.2f}%")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate fleet report: {e}")

    def generate_driver_performance():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            dispatch_sheet = workbook["Dispatch"]
            drivers_sheet = workbook["Drivers"]

            driver_stats = {}

            # Calculate completed deliveries per driver
            for row in dispatch_sheet.iter_rows(min_row=2, values_only=True):
                driver_id = row[1]  # Driver ID
                status = row[4]     # Status

                if driver_id not in driver_stats:
                    driver_stats[driver_id] = {"completed": 0, "in_transit": 0}

                if status == "Delivered":
                    driver_stats[driver_id]["completed"] += 1
                elif status == "In Transit":
                    driver_stats[driver_id]["in_transit"] += 1

            # Update driver performance display
            performance_metrics.delete(1.0, tk.END)
            for driver_id, stats in driver_stats.items():
                driver_name = ""
                for row in drivers_sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] == driver_id:
                        driver_name = row[1]
                        break

                performance_metrics.insert(tk.END,
                    f"Driver: {driver_name} (ID: {driver_id})\n"
                    f"Completed Deliveries: {stats['completed']}\n"
                    f"Active Deliveries: {stats['in_transit']}\n"
                    f"------------------------\n")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate driver performance report: {e}")

    def generate_financial_report():
        try:
            workbook = openpyxl.load_workbook(WORKBOOK_PATH)
            financial_sheet = workbook["Financials"]

            total_revenue = 0
            total_expenses = 0
            expense_categories = {}

            for row in financial_sheet.iter_rows(min_row=2, values_only=True):
                transaction_type = row[1]
                amount = float(row[2])

                if transaction_type == "Order Payment":
                    total_revenue += amount
                else:
                    total_expenses += amount
                    expense_categories[transaction_type] = expense_categories.get(transaction_type, 0) + amount

            # Update financial metrics display
            financial_metrics.delete(1.0, tk.END)
            financial_metrics.insert(tk.END,
                f"Total Revenue: ${total_revenue:,.2f}\n"
                f"Total Expenses: ${total_expenses:,.2f}\n"
                f"Net Profit: ${(total_revenue - total_expenses):,.2f}\n\n"
                f"Expense Breakdown:\n")

            for category, amount in expense_categories.items():
                financial_metrics.insert(tk.END,
                    f"{category}: ${amount:,.2f}\n")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate financial report: {e}")

    def export_reports():
        try:
            # Create a new workbook for reports
            report_wb = openpyxl.Workbook()

            # Export fleet metrics
            fleet_sheet = report_wb.active
            fleet_sheet.title = "Fleet Metrics"
            fleet_sheet.append(["Fleet Usage Report"])
            fleet_sheet.append(fleet_metrics.get(1.0, tk.END).split("\n"))

            # Export driver performance
            perf_sheet = report_wb.create_sheet("Driver Performance")
            perf_sheet.append(["Driver Performance Report"])
            perf_sheet.append(performance_metrics.get(1.0, tk.END).split("\n"))

            # Export financial metrics
            fin_sheet = report_wb.create_sheet("Financial Metrics")
            fin_sheet.append(["Financial Report"])
            fin_sheet.append(financial_metrics.get(1.0, tk.END).split("\n"))

            # Save the report
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            report_wb.save(f"TTMS_Report_{timestamp}.xlsx")
            messagebox.showinfo("Success", "Reports exported successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export reports: {e}")

    root = tk.Tk()
    root.title("Reports & Analytics")
    root.geometry("1200x800")
    back_button = tk.Button(root, text="Back", command=go_back)
    back_button.pack(anchor='nw', padx=10, pady=5)
    # Create notebook for tabbed interface
    notebook = ttk.Notebook(root)
    notebook.pack(pady=10, expand=True)
    # Fleet Usage Tab
    fleet_frame = ttk.Frame(notebook)
    notebook.add(fleet_frame, text="Fleet Usage")

    tk.Button(fleet_frame, text="Generate Fleet Report",
              command=generate_fleet_report).pack(pady=10)
    fleet_metrics = tk.Text(fleet_frame, height=10, width=50)
    fleet_metrics.pack(pady=10)

    # Driver Performance Tab
    performance_frame = ttk.Frame(notebook)
    notebook.add(performance_frame, text="Driver Performance")

    tk.Button(performance_frame, text="Generate Performance Report",
              command=generate_driver_performance).pack(pady=10)
    performance_metrics = tk.Text(performance_frame, height=10, width=50)
    performance_metrics.pack(pady=10)

    # Financial Insights Tab
    financial_frame = ttk.Frame(notebook)
    notebook.add(financial_frame, text="Financial Insights")

    tk.Button(financial_frame, text="Generate Financial Report",
              command=generate_financial_report).pack(pady=10)
    financial_metrics = tk.Text(financial_frame, height=50)
    financial_metrics.pack(pady=10)

    # Export Button

    tk.Button(root, text="Export All Reports",
              command=export_reports).pack(pady=20)

    root.mainloop()

    root.mainloop()
if __name__ == "__main__":
    login_window(window_size=(300, 450))  # Larger window

