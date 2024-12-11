import sys,csv,sqlite3
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table,SimpleDocTemplate, TableStyle
from datetime import datetime, date
from reportlab.lib import colors
import tempfile
import tkinter as tk
from tkinter import messagebox,ttk,filedialog
from tkcalendar import DateEntry
from fpdf import FPDF
import re
import shutil
import pandas as pd
import os
from matplotlib import pyplot as plt
from ttkthemes import ThemedTk
import datetime
from datetime import datetime
import openpyxl
from tkinter import *
import logging
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from PIL import Image, ImageTk
import matplotlib.pyplot as plt
import seaborn as sns

# Database configuration
DATABASE_PATH = "TTMS.db"

def initialize_database():
    """
    Initialize the database by creating it if it doesn't exist
    and ensuring all required tables are present.
    """
    try:
        # Connect to database (creates it if it doesn't exist)
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        # Create tables with appropriate schema
        tables = {
            "Users": '''
                CREATE TABLE IF NOT EXISTS Users (
                    UserID INTEGER PRIMARY KEY AUTOINCREMENT,
                    Username TEXT UNIQUE NOT NULL,
                    Password TEXT NOT NULL,
                    Role TEXT NOT NULL,
                    FullName TEXT,
                    Contact TEXT,
                    Address TEXT,
                    CNIC TEXT UNIQUE,
                    Email TEXT
                )
            ''',
            "Drivers": '''
                CREATE TABLE IF NOT EXISTS Drivers (
                    DriverID INTEGER PRIMARY KEY AUTOINCREMENT,
                    Name TEXT NOT NULL,
                    CNIC TEXT UNIQUE,
                    LicenseExp DATE,
                    Address TEXT,
                    Contact TEXT,
                    Salary REAL,
                    Salary_Status TEXT,
                    DOJ DATE,
                    DOR DATE,
                    Trip INTEGER,
                    Status TEXT
                )
            ''',
            "Trucks": '''
                CREATE TABLE IF NOT EXISTS Trucks (
                    TruckID INTEGER PRIMARY KEY AUTOINCREMENT,
                    Model TEXT,
                    Status TEXT,
                    Permit TEXT,
                    WeightCapacity REAL,
                    MaintenanceSchedule DATE,
                    Odometer INTEGER
                )
            ''',
            "Orders": '''
                CREATE TABLE IF NOT EXISTS Orders (
                    OrderID INTEGER PRIMARY KEY AUTOINCREMENT,
                    OrderName TEXT,
                    CustomerName TEXT,
                    Contact TEXT,
                    Pickup TEXT,
                    Destination TEXT,
                    Region TEXT,
                    Distance REAL,
                    Status TEXT,
                    Weight REAL,
                    GST REAL,
                    TotalAmount REAL,
                    PaidAmount REAL,
                    AmountStatus TEXT,
                    RemainingAmount REAL,
                    OrderDate DATE
                )
            ''',
            "Dispatch": '''
                CREATE TABLE IF NOT EXISTS Dispatch (
                    DispatchID INTEGER PRIMARY KEY AUTOINCREMENT,
                    OrderID INTEGER,
                    DriverID INTEGER,
                    TruckID INTEGER,
                    DispatchTime DATETIME,
                    Status TEXT,
                    EstimatedDeliveryTime DATETIME,
                    FOREIGN KEY (OrderID) REFERENCES Orders(OrderID),
                    FOREIGN KEY (DriverID) REFERENCES Drivers(DriverID),
                    FOREIGN KEY (TruckID) REFERENCES Trucks(TruckID)
                )
            ''',
            "Financials": '''
                CREATE TABLE IF NOT EXISTS Financials (
                    FinancialID INTEGER PRIMARY KEY AUTOINCREMENT,
                    Date DATE,
                    Type TEXT,
                    Amount REAL,
                    Description TEXT,
                    PaymentMode TEXT
                )
            ''',
            "MaintenanceHistory": '''
                CREATE TABLE IF NOT EXISTS MaintenanceHistory (
                    MaintenanceID INTEGER PRIMARY KEY AUTOINCREMENT,
                    TruckID INTEGER,
                    Amount REAL,
                    Description TEXT,
                    Date DATE,
                    Odometer INTEGER,
                    FOREIGN KEY (TruckID) REFERENCES Trucks(TruckID)
                )
            ''',
            "FuelHistory": '''
                CREATE TABLE IF NOT EXISTS FuelHistory (
                    FuelID INTEGER PRIMARY KEY AUTOINCREMENT,
                    TruckID INTEGER,
                    Amount REAL,
                    Liters REAL,
                    Date DATE,
                    FOREIGN KEY (TruckID) REFERENCES Trucks(TruckID)
                )
            ''',
            "SalaryHistory": '''
                CREATE TABLE IF NOT EXISTS SalaryHistory (
                    SalaryID INTEGER PRIMARY KEY AUTOINCREMENT,
                    DriverID INTEGER,
                    Amount REAL,
                    PaymentDate DATE,
                    Status TEXT,
                    FOREIGN KEY (DriverID) REFERENCES Drivers(DriverID)
                )
            ''',
            "LeaveManagement": '''
                CREATE TABLE IF NOT EXISTS LeaveManagement (
                    LeaveID INTEGER PRIMARY KEY AUTOINCREMENT,
                    DriverID INTEGER NOT NULL,
                    StartDate DATE NOT NULL,
                    EndDate DATE NOT NULL,
                    LeaveType TEXT NOT NULL,
                    Status TEXT NOT NULL,
                    Reason TEXT,
                    FOREIGN KEY (DriverID) REFERENCES Drivers(DriverID)
                )
            '''

        }

        # Create each table
        for table_name, create_table_sql in tables.items():
            cursor.execute(create_table_sql)

        # Insert default admin user if not exists
        cursor.execute('''
            INSERT OR IGNORE INTO Users (Username, Password, Role)
            VALUES (?, ?, ?)
        ''', ('admin', 'admin123', 'Admin'))

        # Commit changes and close connection
        conn.commit()
        conn.close()
        return True

    except Exception as e:
        messagebox.showerror("Error", f"Failed to initialize database: {e}")
        return False

# Initialize database
if not initialize_database():
    sys.exit(1)  # Exit if database initialization fails
if not os.path.exists('exports'):
    os.makedirs('exports')


def authenticate_user(username, password):
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()

        # Modified query to select all columns to see proper order
        cursor.execute('''
            SELECT Role, FullName 
            FROM Users 
            WHERE Username = ? AND Password = ?
        ''', (username, password))

        result = cursor.fetchone()
        conn.close()

        if result:
            return result[0], result[1]  # Return tuple of (role, fullname)
        return None, None

    except Exception as e:
        messagebox.showerror("Error", f"Database authentication error: {e}")
        return None, None

def login_window(theme_color="#1a237e", window_size=(400, 500)):
    def login():
        username = username_entry.get()
        password = password_entry.get()

        if username and password:
            role, fullname = authenticate_user(username, password)  # Get both role and fullname

            if role:
                messagebox.showinfo("Success", f"Welcome {username}! ✨")
                root.destroy()
                redirect_user(role, fullname)  # Pass role and username
            else:
                messagebox.showerror("Error", "Invalid credentials ❌")
        else:
            messagebox.showwarning("Warning", "Please fill all fields ⚠️")

    root = tk.Tk()
    root.title("Login")
    # Load and set window icon
    def resource_path(relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Make window size adjustable
    width, height = window_size
    root.geometry(f"{width}x{height}")
    root.configure(bg="white")

    # Center window
    root.eval('tk::PlaceWindow . center')

    # Style configuration
    style = {
        'bg': "white",
        'fg': theme_color,
        'font_title': ("Helvetica", 22, "bold"),
        'font_normal': ("Helvetica", 11),
        'padding': 20
    }

    # Main container
    main_frame = tk.Frame(root, bg=style['bg'])
    main_frame.place(relx=0.5, rely=0.5, anchor="center")

    # Title
    tk.Label(main_frame,
             text="Welcome Back!",
             font=style['font_title'],
             bg=style['bg'],
             fg=style['fg']).pack(pady=style['padding'])

    # Entry fields with modern look
    def create_entry(placeholder, show=None,bg="white"):
        frame = tk.Frame(main_frame, bg=style['bg'])
        frame.pack(pady=10, padx=style['padding'])

        entry = tk.Entry(frame,
                         font=style['font_normal'],
                         show=show,
                         width=25,
                         bd=0,
                         bg="#f0f0f0")
        entry.insert(0, placeholder)
        entry.pack(ipady=8, pady=5)

        # Add underline
        tk.Frame(frame, height=2, bg=theme_color).pack(fill="x")

        return entry

    username_entry = create_entry("Username")
    username_entry.configure(bg="white")  # or bg="SystemWindow" for system default

    password_entry = create_entry("Password", show="●")
    password_entry.configure(bg="white")  # or bg="SystemWindow" for system default

    # Login button
    login_btn = tk.Button(main_frame,
                          text="LOGIN",
                          command=login,
                          font=style['font_normal'],
                          bg=theme_color,
                          fg="white",
                          width=20,
                          bd=0,
                          cursor="hand2")
    login_btn.pack(pady=style['padding'])

    # Bind Enter key
    root.bind('<Return>', lambda event: login())

    root.mainloop()



# Role-based redirection (Placeholder for dashboards)
def redirect_user(role, username):  # Add username parameter
    if role == "Admin":
        messagebox.showinfo("Admin Dashboard", "Redirecting to Admin Dashboard...")
        admin_dashboard(user_type="Admin", username=username)  # Pass username
        # Call Admin Dashboard
    elif role == "Dispatcher":
        messagebox.showinfo("Dispatcher Dashboard", "Redirecting to Dispatcher Dashboard...")
        dispatcher_dashboard(user_type="Dispatcher", username=username)
        # Call Dispatcher Dashboard
    elif role == "Accountant":
        messagebox.showinfo("Accountant Dashboard", "Redirecting to Accountant Dashboard...")
        accountant_dashboard(user_type="Accountant", username=username)
        # Call Accountant Dashboard
    elif role == "Manager":
        messagebox.showinfo("Manager Dashboard", "Redirecting to Manager Dashboard...")
        manager_dashboard(user_type="Manager", username=username)
        # Call Manager Dashboard
    elif role == "Driver":
        messagebox.showinfo("Driver Dashboard", "Redirecting to Driver Dashboard...")
        # Call Driver Dashboard
    else:
        messagebox.showerror("Access Denied", "Role not recognized!")

def resource_path(relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

def admin_dashboard(user_type="Admin",username=None, login_window=None):
    import datetime
    def update_time():
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_label.config(text=current_time)
        root.after(1000, update_time)

    def logout():
        root.destroy()
        if login_window:
            login_window.deiconify()

    def open_driver_management():
        root.destroy()
        driver_management_gui(user_role="Admin")

    def open_truck_management():
        root.destroy()
        truck_management_gui(user_role="Admin")

    def open_order_management():
        root.destroy()
        order_management_gui(user_role="Admin")

    def open_dispatch_management():
        root.destroy()
        dispatch_management_gui(user_role="Admin")

    def open_account_management():
        root.destroy()
        accounts_management_gui(user_role="Admin")

    def open_report_management():
        root.destroy()
        reports_analytics_gui(user_role="Admin")

    def open_user_management():
        root.destroy()
        user_management_gui()

    def create_image_button(parent, image_path, text, command):
        frame = ttk.Frame(parent)
        try:
            img = Image.open(image_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            btn = ttk.Button(frame, text=text, compound=tk.TOP, image=photo, command=command)
            btn.image = photo  # Keep a reference to avoid garbage collection
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            btn = ttk.Button(frame, text=text, command=command)
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        return frame

    root = tk.Tk()
    root.title("TruckFlow Solutions - Admin Dashboard")
    root.geometry("1000x600")
    root.minsize(1024, 600)

    # Create a main frame to hold all sections
    main_frame = tk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=True)

    # Header: Create a frame for company name, address, and contact details
    header_frame = tk.Frame(main_frame, bg="#1a237e", height=150)
    header_frame.pack(fill=tk.X, side=tk.TOP)



    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a centered frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    # Create a frame for Logout, User Type, and Current Time in the same row
    user_info_frame = tk.Frame(header_frame, bg="#1a237e")
    user_info_frame.pack(side=tk.TOP, fill=tk.X)

    # User Type (left)
    user_label = tk.Label(
        user_info_frame,
        text=f"User Type: {user_type} | User: {username}",  # Modified this line
        font=("Arial", 12, "bold"),
        bg="#1a237e",
        fg="white",
    )
    user_label.pack(side=tk.LEFT, padx=20, pady=10)
    # Logout Button (with styling)
    logout_btn = ttk.Button(user_info_frame, text="Logout", command=logout, style="Logout.TButton")
    logout_btn.pack(side=tk.RIGHT, padx=20, pady=10)
    # Time (right)
    time_label = tk.Label(user_info_frame, font=("Helvetica", 12), bg="#1a237e", fg="white")
    time_label.pack(side=tk.LEFT, padx=20, pady=10)



    # Content Area
    content_frame = tk.Frame(main_frame, bg="#f5f5f5")
    content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    content_frame.grid_columnconfigure((0, 1, 2), weight=1)
    content_frame.grid_rowconfigure((0, 1, 2), weight=1)

    buttons = [
        ("Manage Drivers", resource_path("icons/drivers.png"), open_driver_management, 0, 0),
        ("Manage Trucks", resource_path("icons/trucks.png"), open_truck_management, 0, 1),
        ("Manage Orders", resource_path("icons/orders.png"), open_order_management, 0, 2),
        ("Manage Accounts", resource_path("icons/accounts.png"), open_account_management, 1, 0),
        ("Manage Dispatch", resource_path("icons/dispatch1.png"), open_dispatch_management, 1, 1),
        ("View Reports", resource_path("icons/reports.png"), open_report_management, 1, 2),
        ("User Management", resource_path("icons/users.png"), open_user_management, 2, 1),
    ]

    for text, icon_path, command, row, col in buttons:
        btn_frame = create_image_button(content_frame, icon_path, text, command)
        btn_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    # Footer
    footer_frame = tk.Frame(main_frame, bg="#1a237e", height=50)
    footer_frame.pack(fill=tk.X, side=tk.BOTTOM)

    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.BOTTOM, pady=10)

    # Update the time every second
    update_time()
    root.mainloop()

def manager_dashboard(user_type="Manager",username=None, login_window=None):
    import datetime
    def update_time():
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_label.config(text=current_time)
        root.after(1000, update_time)

    def logout():
        root.destroy()
        if login_window:
            login_window.deiconify()

    def open_driver_management():
        root.destroy()
        driver_management_gui(user_role="Manager")

    def open_truck_management():
        root.destroy()
        truck_management_gui(user_role="Manager")

    def open_order_management():
        root.destroy()
        order_management_gui(user_role="Manager")

    def open_dispatch_management():
        root.destroy()
        dispatch_management_gui(user_role="Manager")

    def open_account_management():
        root.destroy()
        accounts_management_gui(user_role="Manager")

    def open_report_management():
        root.destroy()
        reports_analytics_gui(user_role="Manager")
    #
    # def open_user_management():
    #     root.destroy()
    #     user_management_gui()

    def create_image_button(parent, image_path, text, command):
        frame = ttk.Frame(parent)
        try:
            img = Image.open(image_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            btn = ttk.Button(frame, text=text, compound=tk.TOP, image=photo, command=command)
            btn.image = photo  # Keep a reference to avoid garbage collection
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            btn = ttk.Button(frame, text=text, command=command)
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        return frame

    root = tk.Tk()
    root.title("TruckFlow Solutions - Manager Dashboard")
    root.geometry("1000x600")
    root.minsize(1024, 600)

    # Create a main frame to hold all sections
    main_frame = tk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=True)

    header_frame = tk.Frame(main_frame, bg="#1a237e", height=150)
    header_frame.pack(fill=tk.X, side=tk.TOP)



    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a centered frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    # Create a frame for Logout, User Type, and Current Time in the same row
    user_info_frame = tk.Frame(header_frame, bg="#1a237e")
    user_info_frame.pack(side=tk.TOP, fill=tk.X)

    # User Type (left)
    user_label = tk.Label(
        user_info_frame,
        text=f"User Type: {user_type} | User: {username}",
        font=("Arial", 12, "bold"),
        bg="#1a237e",
        fg="white",
    )
    user_label.pack(side=tk.LEFT, padx=20, pady=10)
    # Logout Button (with styling)
    logout_btn = ttk.Button(user_info_frame, text="Logout", command=logout, style="Logout.TButton")
    logout_btn.pack(side=tk.RIGHT, padx=20, pady=10)
    # Time (right)
    time_label = tk.Label(user_info_frame, font=("Helvetica", 12), bg="#1a237e", fg="white")
    time_label.pack(side=tk.LEFT, padx=20, pady=10)



    # Content Area
    content_frame = tk.Frame(main_frame, bg="#f5f5f5")
    content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    content_frame.grid_columnconfigure((0, 1, 2), weight=1)
    content_frame.grid_rowconfigure((0, 1, 2), weight=1)

    buttons = [
        ("Manage Drivers", resource_path("icons/drivers.png"), open_driver_management, 0, 0),
        ("Manage Trucks", resource_path("icons/trucks.png"), open_truck_management, 0, 1),
        ("Manage Orders", resource_path("icons/orders.png"), open_order_management, 0, 2),
        ("Manage Accounts", resource_path("icons/accounts.png"), open_account_management, 1, 0),
        ("Manage Dispatch", resource_path("icons/dispatch1.png"), open_dispatch_management, 1, 1),
        ("View Reports", resource_path("icons/reports.png"), open_report_management, 1, 2),
        # ("User Management", resource_path("icons/users.png"), open_user_management, 2, 1),
    ]

    for text, icon_path, command, row, col in buttons:
        btn_frame = create_image_button(content_frame, icon_path, text, command)
        btn_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    # Footer
    footer_frame = tk.Frame(main_frame, bg="#1a237e", height=50)
    footer_frame.pack(fill=tk.X, side=tk.BOTTOM)

    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.BOTTOM, pady=10)

    # Update the time every second
    update_time()
    root.mainloop()


def accountant_dashboard(user_type="Accountant",username=None, login_window=None):
    import datetime
    def update_time():
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_label.config(text=current_time)
        root.after(1000, update_time)

    def logout():
        root.destroy()
        if login_window:
            login_window.deiconify()

    def open_account_management():
        root.destroy()
        accounts_management_gui(user_role="Accountant")

    def create_image_button(parent, image_path, text, command):
        frame = ttk.Frame(parent)
        try:
            img = Image.open(image_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            btn = ttk.Button(frame, text=text, compound=tk.TOP, image=photo, command=command)
            btn.image = photo  # Keep a reference to avoid garbage collection
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            btn = ttk.Button(frame, text=text, command=command)
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        return frame

    root = tk.Tk()
    root.title("TruckFlow Solutions - Accountant Dashboard")
    root.geometry("1000x600")
    root.minsize(1024, 600)
    # Create a main frame to hold all sections
    main_frame = tk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=True)

    header_frame = tk.Frame(main_frame, bg="#1a237e", height=150)
    header_frame.pack(fill=tk.X, side=tk.TOP)



    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")

    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a centered frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    # Create a frame for Logout, User Type, and Current Time in the same row
    user_info_frame = tk.Frame(header_frame, bg="#1a237e")
    user_info_frame.pack(side=tk.TOP, fill=tk.X)

    # User Type (left)
    user_label = tk.Label(
        user_info_frame,
        text=f"User Type: {user_type} | User: {username}",
        font=("Arial", 12, "bold"),
        bg="#1a237e",
        fg="white",
    )
    user_label.pack(side=tk.LEFT, padx=20, pady=10)
    # Logout Button (with styling)
    logout_btn = ttk.Button(user_info_frame, text="Logout", command=logout, style="Logout.TButton")
    logout_btn.pack(side=tk.RIGHT, padx=20, pady=10)
    # Time (right)
    time_label = tk.Label(user_info_frame, font=("Helvetica", 12), bg="#1a237e", fg="white")
    time_label.pack(side=tk.LEFT, padx=20, pady=10)

    # Content Area
    content_frame = tk.Frame(main_frame, bg="#f5f5f5")
    content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    content_frame.grid_columnconfigure((0,1,2), weight=1)
    content_frame.grid_rowconfigure((0,1,2), weight=1)

    buttons = [
        ("Manage Accounts", resource_path("icons/accounts.png"), open_account_management, 0, 0),

    ]

    for text, icon_path, command, row, col in buttons:
        btn_frame = create_image_button(content_frame, icon_path, text, command)
        btn_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    # Footer
    footer_frame = tk.Frame(main_frame, bg="#1a237e", height=50)
    footer_frame.pack(fill=tk.X, side=tk.BOTTOM)

    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.BOTTOM, pady=10)

    # Update the time every second
    update_time()
    root.mainloop()


def dispatcher_dashboard(user_type="Dispatcher",username=None, login_window=None):
    import datetime
    def update_time():
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        time_label.config(text=current_time)
        root.after(1000, update_time)

    def logout():
        root.destroy()
        if login_window:
            login_window.deiconify()

    def open_dispatch_management():
        root.destroy()
        dispatch_management_gui(user_role="Dispatcher")

    def open_order_management():
        root.destroy()
        order_management_gui(user_role="Dispatcher")

    def create_image_button(parent, image_path, text, command):
        frame = ttk.Frame(parent)
        try:
            img = Image.open(image_path)
            img = img.resize((64, 64), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            btn = ttk.Button(frame, text=text, compound=tk.TOP, image=photo, command=command)
            btn.image = photo  # Keep a reference to avoid garbage collection
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            btn = ttk.Button(frame, text=text, command=command)
            btn.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        return frame

    root = tk.Tk()
    root.title("TruckFlow Solutions - Dispatcher Dashboard")
    root.geometry("1000x600")
    root.minsize(1024, 600)
    # Create a main frame to hold all sections
    main_frame = tk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=True)

    header_frame = tk.Frame(main_frame, bg="#1a237e", height=150)
    header_frame.pack(fill=tk.X, side=tk.TOP)



    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a centered frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    # Create a frame for Logout, User Type, and Current Time in the same row
    user_info_frame = tk.Frame(header_frame, bg="#1a237e")
    user_info_frame.pack(side=tk.TOP, fill=tk.X)

    # User Type (left)
    user_label = tk.Label(
        user_info_frame,
        text=f"User Type: {user_type} | User: {username}",
        font=("Arial", 12, "bold"),
        bg="#1a237e",
        fg="white",
    )
    user_label.pack(side=tk.LEFT, padx=20, pady=10)
    # Logout Button (with styling)
    logout_btn = ttk.Button(user_info_frame, text="Logout", command=logout, style="Logout.TButton")
    logout_btn.pack(side=tk.RIGHT, padx=20, pady=10)
    # Time (right)
    time_label = tk.Label(user_info_frame, font=("Helvetica", 12), bg="#1a237e", fg="white")
    time_label.pack(side=tk.LEFT, padx=20, pady=10)

    # Content Area
    content_frame = tk.Frame(main_frame, bg="#f5f5f5")
    content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    content_frame.grid_columnconfigure((0, 1, 2), weight=1)
    content_frame.grid_rowconfigure((0, 1, 2), weight=1)

    buttons = [
        ("Manage Dispatch",  resource_path("icons/dispatch1.png"), open_dispatch_management, 0, 0),
        ("Manage Orders",  resource_path("icons/orders.png"), open_order_management, 0, 1),

    ]

    for text, icon_path, command, row, col in buttons:
        btn_frame = create_image_button(content_frame, icon_path, text, command)
        btn_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")

    # Footer
    footer_frame = tk.Frame(main_frame, bg="#1a237e", height=50)
    footer_frame.pack(fill=tk.X, side=tk.BOTTOM)

    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.BOTTOM, pady=10)

    # Update the time every second
    update_time()
    root.mainloop()

def driver_management_gui(user_role=""):
    def go_back():
        """Handle the back button based on user role."""
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()

    def load_drivers():
        """Load driver data from the Excel sheet."""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM Drivers")
            drivers = cursor.fetchall()
            conn.close()
            return drivers
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load drivers: {e}")
            return []

    def save_driver(driver_data):
        """Save new driver to the Excel sheet."""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute("""
                    INSERT INTO Drivers (Name, CNIC, LicenseExp, Address, Contact,
                                       Salary, Salary_Status, DOJ, DOR, Trip, Status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, driver_data[1:])  # Exclude ID as it's autoincrement
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Driver added successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save driver: {e}")

    def track_salary_payment(driver_id, amount):
        """Record salary payment for a driver"""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Check current salary status
            cursor.execute("SELECT Salary_Status FROM Drivers WHERE DriverID=?", (driver_id,))
            salary_status = cursor.fetchone()[0]

            if salary_status == "Unpaid":
                # Update salary status
                cursor.execute("""
                    UPDATE Drivers
                    SET Salary_Status='Paid'
                    WHERE DriverID=?
                """, (driver_id,))

                # Record payment in salary history
                cursor.execute("""
                    INSERT INTO SalaryHistory (DriverID, Amount, PaymentDate, Status)
                    VALUES (?, ?, date('now'), 'Paid')
                """, (driver_id, amount))

                conn.commit()
                messagebox.showinfo("Success", "Salary payment recorded successfully!")
            else:
                messagebox.showinfo("Notice", "Salary already paid. No action required.")

            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to record salary payment: {e}")

    def generate_salary_report():
        """Generate monthly salary report with enhanced GUI"""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Get current month's data
            current_month = datetime.now().strftime("%Y-%m")

            # Get salary payments for current month
            cursor.execute("""
                SELECT Amount FROM SalaryHistory
                WHERE strftime('%Y-%m', PaymentDate) = ?
            """, (current_month,))
            monthly_data = cursor.fetchall()

            # Get unpaid salaries
            cursor.execute("""
                SELECT DriverID, Name FROM Drivers
                WHERE Salary_Status = 'Unpaid'
            """)
            unpaid_reminders = cursor.fetchall()

            # Close the database connection
            conn.close()

            # Enhanced visualization
            report_window = ThemedTk(theme="arc")  # Modern theme
            report_window.title(f"Salary Report - {current_month}")
            report_window.geometry("600x400")

            # Main frame
            main_frame = ttk.Frame(report_window, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Header
            ttk.Label(main_frame, text=f"Monthly Salary Report ({current_month})", font=("Helvetica", 14, "bold")).pack(
                pady=10)

            # Summary frame
            summary_frame = ttk.LabelFrame(main_frame, text="Summary", padding="5")
            summary_frame.pack(fill=tk.X, pady=5)

            ttk.Label(summary_frame, text=f"Total Payments: {len(monthly_data)}", font=("Helvetica", 10)).pack()
            # Then modify the sum calculation to:
            ttk.Label(summary_frame, text=f"Total Amount: /-{sum(payment[0] for payment in monthly_data)}",
                      font=("Helvetica", 10, "bold")).pack()

            # Unpaid reminders
            if unpaid_reminders:
                reminder_frame = ttk.LabelFrame(main_frame, text="Unpaid Salary Reminders", padding="5")
                reminder_frame.pack(fill=tk.BOTH, expand=True, pady=5)

                # Create scrollable text widget for reminders
                reminder_text = tk.Text(reminder_frame, height=10, wrap=tk.WORD)
                scrollbar = ttk.Scrollbar(reminder_frame, orient=tk.VERTICAL, command=reminder_text.yview)
                reminder_text.configure(yscrollcommand=scrollbar.set)

                reminder_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

                for reminder in unpaid_reminders:
                    driver_id, driver_name = reminder  # Unpack the tuple
                    reminder_text.insert(tk.END, f"• Unpaid salary for Driver {driver_name} (ID: {driver_id},Name: {driver_name})\n")
                reminder_text.configure(state='disabled')

            ttk.Button(main_frame, text="Close", command=report_window.destroy).pack(pady=10)

            report_window.mainloop()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate salary report: {e}")


    def track_driver_status():
        """Track driver availability and trip status"""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Get current status counts
            cursor.execute("""
                SELECT Status, COUNT(*)
                FROM Drivers
                GROUP BY Status
            """)
            status_counts = dict(cursor.fetchall())

            # Get individual driver statistics
            cursor.execute("""
                SELECT DriverID, Name, Status,
                    COUNT(CASE WHEN Status = 'Available' THEN 1 END) as available_days,
                    COUNT(CASE WHEN Status = 'On Trip' THEN 1 END) as trip_days,
                    COUNT(CASE WHEN Status IN ('Off Duty','On Leave') THEN 1 END) as off_days
                FROM Drivers               
                GROUP BY DriverID, Name

            """)
            driver_stats = cursor.fetchall()

            status_window = ThemedTk(theme="arc")
            status_window.title("Driver Status Report")
            status_window.geometry("700x500")

            main_frame = ttk.Frame(status_window, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Header
            ttk.Label(main_frame, text="Driver Status Report", font=("Helvetica", 14, "bold")).pack(pady=10)

            # Status summary frame
            summary_frame = ttk.LabelFrame(main_frame, text="Current Status Summary", padding="5")
            summary_frame.pack(fill=tk.X, pady=5)

            # Create status bars
            for status, count in status_counts.items():
                frame = ttk.Frame(summary_frame)
                frame.pack(fill=tk.X, pady=2)
                ttk.Label(frame, text=f"{status}:", width=15).pack(side=tk.LEFT)
                progress = ttk.Progressbar(frame, length=200, mode='determinate')
                progress.pack(side=tk.LEFT, padx=5)
                total_drivers = sum(status_counts.values())
                progress['value'] = (count / total_drivers) * 100
                ttk.Label(frame, text=f"{count} drivers").pack(side=tk.LEFT)

            # Driver details section
            details_frame = ttk.LabelFrame(main_frame, text="Detailed Driver Statistics", padding="5")
            details_frame.pack(fill=tk.BOTH, expand=True, pady=5)

            # Add scrollbars
            tree_scroll_y = ttk.Scrollbar(details_frame, orient="vertical")
            tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

            # Create treeview for driver details
            tree = ttk.Treeview(details_frame, columns=("ID", "Name", "Available", "Trip", "Off"), show="headings",
                                yscrollcommand=tree_scroll_y.set)
            tree_scroll_y.config(command=tree.yview)

            # Configure column headings and widths
            column_widths = {"ID": 100, "Name": 150, "Available": 100, "Trip": 100, "Off": 100}
            for col in tree["columns"]:
                tree.heading(col, text=col)
                tree.column(col, width=column_widths[col], anchor="center")  # Center align all columns

            # Insert data
            # Insert data
            for driver in driver_stats:
                tree.insert("", tk.END, values=(
                    driver[0],  # DriverID
                    driver[1],  # Name
                    driver[3],  # available_days
                    driver[4],  # trip_days
                    driver[5]  # off_days
                ))

            # Make the tree expand with window resizing
            tree.pack(fill=tk.BOTH, expand=True)

            # Configure weight of the details_frame columns
            details_frame.columnconfigure(0, weight=1)
            details_frame.rowconfigure(0, weight=1)

            # Bind resize event to adjust column widths
            def on_window_resize(event):
                window_width = event.width
                for col in tree["columns"]:
                    tree.column(col, width=int(window_width * column_widths[col] / sum(column_widths.values())))

            status_window.bind("<Configure>", on_window_resize)

            ttk.Button(main_frame, text="Close", command=status_window.destroy).pack(pady=10)

            status_window.mainloop()
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate status report: {e}")

    def manage_driver_leave():
        """Manage driver leave requests and tracking"""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()


            # Create leave management window
            leave_window = ThemedTk(theme="arc")
            leave_window.title("Leave Management")
            leave_window.geometry("800x600")

            main_frame = ttk.Frame(leave_window, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Leave request form
            form_frame = ttk.LabelFrame(main_frame, text="Leave Request Form", padding="10")
            form_frame.pack(fill=tk.X, pady=5)

            # Form fields
            ttk.Label(form_frame, text="Driver ID:").grid(row=0, column=0, padx=5, pady=5)
            driver_id_entry = ttk.Entry(form_frame)
            driver_id_entry.grid(row=0, column=1, padx=5, pady=5)

            ttk.Label(form_frame, text="Start Date:").grid(row=0, column=2, padx=5, pady=5)
            start_date = DateEntry(form_frame, width=19, background='darkblue',foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
            start_date.grid(row=0, column=3, padx=5, pady=5)

            ttk.Label(form_frame, text="End Date:").grid(row=1, column=2, padx=5, pady=5)
            end_date = DateEntry(form_frame, width=19, background='darkblue',foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
            end_date.grid(row=1, column=3, padx=5, pady=5)

            ttk.Label(form_frame, text="Leave Type:").grid(row=1, column=0, padx=5, pady=5)
            leave_type = ttk.Combobox(form_frame, values=["Off Duty","Sick Leave", "Annual Leave", "Emergency Leave"])
            leave_type.grid(row=1, column=1, padx=5, pady=5)

            ttk.Label(form_frame, text="Reason:").grid(row=2, column=0, padx=5, pady=5)
            reason_entry = ttk.Entry(form_frame, width=50)
            reason_entry.grid(row=2, column=1, columnspan=3, padx=5, pady=5)

            def submit_leave_request():
                """Submit new leave request"""
                try:
                    cursor.execute('''
                                    INSERT INTO LeaveManagement 
                                    (DriverID, StartDate, EndDate, LeaveType, Status, Reason)
                                    VALUES (?, ?, ?, ?, ?, ?)
                                ''', (
                        driver_id_entry.get(),
                        start_date.get(),
                        end_date.get(),
                        leave_type.get(),
                        "Pending",
                        reason_entry.get()
                    ))
                    conn.commit()
                    messagebox.showinfo("Success", "Leave request submitted successfully!")
                    refresh_leave_table()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to submit leave request: {e}")
            ttk.Button(form_frame, text="Submit Request", command=submit_leave_request).grid(row=3, column=1, columnspan=2, pady=10)

            # Leave requests table
            table_frame = ttk.LabelFrame(main_frame, text="Leave Requests", padding="5")
            table_frame.pack(fill=tk.BOTH, expand=True, pady=5)

            columns = ("Driver ID", "Start Date", "End Date", "Leave Type", "Status", "Reason")
            leave_table = ttk.Treeview(table_frame, columns=columns, show='headings')

            for col in columns:
                leave_table.heading(col, text=col,anchor='center')
                leave_table.column(col, width=100,anchor='center')

            def refresh_leave_table():
                """Refresh the leave requests table"""
                leave_table.delete(*leave_table.get_children())
                cursor.execute(
                    'SELECT DriverID, StartDate, EndDate, LeaveType, Status, Reason FROM LeaveManagement')
                for row in cursor.fetchall():
                    leave_table.insert("", "end", values=row)

            def approve_leave():
                """Approve selected leave request"""
                selected = leave_table.selection()
                if not selected:
                    messagebox.showerror("Error", "No request selected!")
                    return

                item = leave_table.item(selected[0])
                driver_id = item['values'][0]

                # Update leave status
                cursor.execute('''
                                UPDATE LeaveManagement 
                                SET Status = 'Approved' 
                                WHERE DriverID = ?
                            ''', (driver_id,))

                # Update driver status
                cursor.execute('''
                                UPDATE drivers 
                                SET Status = 'On Leave' 
                                WHERE DriverID = ?
                            ''', (driver_id,))

                conn.commit()
                refresh_leave_table()
                messagebox.showinfo("Success", "Leave request approved!")

            def reject_leave():
                """Reject selected leave request"""
                selected = leave_table.selection()
                if not selected:
                    messagebox.showerror("Error", "No request selected!")
                    return
                item = leave_table.item(selected[0])
                driver_id = item['values'][0]

                cursor.execute('''
                                                UPDATE LeaveManagement 
                                                SET Status = 'Rejected' 
                                                WHERE DriverID = ?
                                            ''', (driver_id,))
                conn.commit()
                refresh_leave_table()
                messagebox.showinfo("Success", "Leave request rejected!")

            # Action buttons
            button_frame = ttk.Frame(table_frame)
            button_frame.pack(fill=tk.X, pady=5)
            ttk.Button(button_frame, text="Approve", command=approve_leave).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Reject", command=reject_leave ).pack(side=tk.LEFT, padx=5)
            ttk.Button(button_frame, text="Refresh", command=refresh_leave_table).pack(side=tk.LEFT, padx=5)

            # Add scrollbars
            y_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=leave_table.yview)
            leave_table.configure(yscrollcommand=y_scrollbar.set)

            # Pack table and scrollbar
            leave_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            # Initial load of leave requests
            refresh_leave_table()

            leave_window.mainloop()

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open leave management: {e}")

    def update_driver(driver_id, updated_data):
        try:
            conn = sqlite3.connect('TTMS.db')  # Make sure this matches your actual database name
            cursor = conn.cursor()
            cursor.execute("""
                    UPDATE Drivers
                    SET Name=?, CNIC=?, LicenseExp=?, Address=?, Contact=?,
                        Salary=?, Salary_Status=?, DOJ=?, DOR=?, Trip=?, Status=?
                    WHERE DriverID=?
                """, (updated_data[1], updated_data[2], updated_data[3],
                      updated_data[4], updated_data[5], updated_data[6],
                      updated_data[7], updated_data[8], updated_data[9],
                      updated_data[10], updated_data[11], driver_id))
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Driver updated successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update driver: {e}")

    def update_trip_count(driver_id):
        """Increment trip count when trip status changes to 'On Trip'"""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE Drivers
                SET Trip = Trip + 1
                WHERE DriverID = ?
            """, (driver_id,))

            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update trip count: {e}")

    def delete_driver(driver_id):
        """Delete a driver from the Excel sheet."""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM Drivers WHERE DriverID=?", (driver_id,))
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Driver deleted successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete driver: {e}")

    def check_license_expiry():
        """Check for upcoming license expirations."""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            cursor.execute("""
                    SELECT DriverID, Name, LicenseExp
                    FROM Drivers
                    WHERE julianday(LicenseExp) - julianday('now') <= 30
                """)
            expiring_licenses = cursor.fetchall()

            if expiring_licenses:
                notifications = []
                for driver in expiring_licenses:
                    days_to_expiry = (datetime.strptime(driver[2], "%Y-%m-%d") - datetime.now()).days
                    notifications.append(
                        f"Driver {driver[1]} (ID: {driver[0]}) - License expires in {days_to_expiry} days")
                messagebox.showwarning("License Expiry Notifications", "\n".join(notifications))
            else:
                messagebox.showinfo("Compliance Check", "No upcoming license expirations.")

            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to check license expiry: {e}")

    def load_data_into_table(filtered_data=None):
        """Load all driver data or filtered data into the table."""
        driver_table.delete(*driver_table.get_children())  # Clear existing data
        drivers = filtered_data if filtered_data else load_drivers()
        for idx, driver in enumerate(drivers, start=1):  # Row indexing starts at 1 for clarity
            driver_table.insert("", "end", values=(idx, *driver))

    def search_drivers():
        """Filter and display drivers based on search criteria."""
        criteria = search_criteria.get()
        search_value = search_entry.get().lower()
        if not search_value:
            messagebox.showerror("Error", "Search field cannot be empty!")
            return

        criteria_map = {
            "ID": "DriverID",
            "Name": "Name",
            "Contact": "Contact",
            "License": "LicenseExp",
        }

        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            column = criteria_map.get(criteria, "Name")
            cursor.execute(f"""
                SELECT * FROM Drivers
                WHERE lower({column}) LIKE ?
            """, (f'%{search_value}%',))

            filtered_drivers = cursor.fetchall()
            conn.close()

            if not filtered_drivers:
                messagebox.showinfo("No Results", "No matching drivers found.")
            load_data_into_table(filtered_drivers)
        except Exception as e:
            messagebox.showerror("Error", f"Search failed: {e}")

    def add_driver():
        """Add a new driver to the database."""
        driver_data = [
            id_entry.get(),
            name_entry.get(),
            cnic_entry.get(),
            license_entry.get(),
            address_entry.get(),
            contact_entry.get(),
            salary_entry.get(),
            salary_status_combo.get(),
            joining_date_entry.get(),
            resigning_date_entry.get(),
            trip_count_entry.get(),
            trip_status_combo.get()
        ]

        # Validate ID uniqueness
        drivers = load_drivers()
        existing_ids = [driver[0] for driver in drivers]
        if driver_data[0] in existing_ids:
            messagebox.showerror("Error", "Driver ID already exists!")
            return

        save_driver(driver_data)
        load_data_into_table()

    def update_selected_driver():
        """Update the selected driver's details."""
        selected_item = driver_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No driver selected!")
            return

        selected_row_index = driver_table.item(selected_item)["values"][0]
        updated_data = [
            id_entry.get(),
            name_entry.get(),
            cnic_entry.get(),
            license_entry.get(),
            address_entry.get(),
            contact_entry.get(),
            salary_entry.get(),
            salary_status_combo.get(),
            joining_date_entry.get(),
            resigning_date_entry.get(),
            trip_count_entry.get(),
            trip_status_combo.get()
        ]
        print(f"Updating driver ")  # Debug print

        if trip_status_combo.get() == "On Trip":
            update_trip_count(selected_row_index)

        update_driver(selected_row_index, updated_data)
        load_data_into_table()

    def delete_selected_driver():
        """Delete the selected driver."""
        selected_item = driver_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No driver selected!")
            return

        selected_row_index = driver_table.item(selected_item)["values"][0]
        delete_driver(selected_row_index)
        load_data_into_table()

    def on_driver_select(event):
        """Populate form fields when a driver is selected."""
        selected_item = driver_table.selection()
        if selected_item:
            selected_driver = driver_table.item(selected_item)["values"]
            id_entry.delete(0, tk.END)
            id_entry.insert(0, selected_driver[1])
            name_entry.delete(0, tk.END)
            name_entry.insert(0, selected_driver[2])
            cnic_entry.delete(0, tk.END)
            cnic_entry.insert(0, selected_driver[3])
            license_entry.delete(0,tk.END)
            license_entry.insert(0,selected_driver[4])
            address_entry.delete(0, tk.END)
            address_entry.insert(0, selected_driver[5])
            contact_entry.delete(0, tk.END)
            contact_entry.insert(0, selected_driver[6])
            salary_entry.delete(0, tk.END)
            salary_entry.insert(0, selected_driver[7])
            salary_status_combo.set(selected_driver[8])
            joining_date_entry.delete(0,tk.END)
            joining_date_entry.insert(0,selected_driver[9])
            resigning_date_entry.delete(0, tk.END)
            resigning_date_entry.insert(0, selected_driver[10])
            trip_count_entry.delete(0, tk.END)
            trip_count_entry.insert(0, selected_driver[11])
            trip_status_combo.set(selected_driver[12])
    def refresh_treeview():
        """Refresh the treeview with latest data from the Excel file"""
        try:
            load_data_into_table()
            messagebox.showinfo("Success", "Data refreshed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh data: {e}")

    def export_reports(report_type):
        """Export reports to Excel"""
        try:
            # Create exports directory if it doesn't exist
            if not os.path.exists('exports'):
                os.makedirs('exports')

            conn = sqlite3.connect('TTMS.db')

            if report_type == "salary":
                # Get salary data directly from database
                query = """
                    SELECT d.DriverID, d.Name, d.Salary, d.Salary_Status, 
                           sh.PaymentDate, sh.Amount
                    FROM Drivers d
                    LEFT JOIN SalaryHistory sh ON d.DriverID = sh.DriverID
                    WHERE strftime('%Y-%m', sh.PaymentDate) = strftime('%Y-%m', 'now')
                """
                df = pd.read_sql_query(query, conn)
                filename = f"exports/salary_report_{datetime.now().strftime('%Y%m')}.xlsx"

            elif report_type == "status":
                # Get status data directly from database
                query = """
                    SELECT DriverID, Name, Status, Trip
                    FROM Drivers
                """
                df = pd.read_sql_query(query, conn)
                filename = f"exports/status_report_{datetime.now().strftime('%Y%m%d')}.xlsx"

            conn.close()
            df.to_excel(filename, index=False)
            messagebox.showinfo("Success", f"Report exported to {filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export report: {e}")

    def backup_database():
        """Create backup of the driver database"""
        try:
            # Create backup directory if it doesn't exist
            if not os.path.exists('backup'):
                os.makedirs('backup')

            # Use the correct database path and keep SQLite format
            database_path = 'TTMS.db'
            backup_path = f"backup/TTMS_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy2(database_path, backup_path)
            messagebox.showinfo("Success", f"Database backed up to {backup_path}")
        except Exception as e:
            messagebox.showerror("Backup Error", f"Failed to backup database: {e}")

    def calculate_driver_performance():
        """Calculate driver performance metrics and display them"""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            performance_data = {}
            on_time_deliveries=0

            cursor.execute("SELECT DriverID,Name,Trip FROM Drivers")
            rows = cursor.fetchall()
            for row in rows:
                driver_id = row[0]
                name = row[1]
                trips = row[2]
                performance_score = (trips * 0.6) + (on_time_deliveries * 0.4)
                performance_data[driver_id] = {
                    "name": name,
                    "trips": trips,
                    "on_time": on_time_deliveries,
                    "score": performance_score
                }


            # Create a new window to display performance data
            performance_window = tk.Toplevel()
            performance_window.title("Driver Performance Report")
            performance_window.geometry("600x400")

            # Create a frame to contain the treeview and scrollbar
            frame = ttk.Frame(performance_window)
            frame.pack(fill="both", expand=True)

            # Create a treeview to display the data
            tree = ttk.Treeview(frame, columns=("ID", "Name", "Trips", "On-Time", "Score"),show="headings")

            # Define column headings and center alignment
            columns = ["ID", "Name", "Trips", "On-Time", "Score"]
            column_widths = [80, 150, 100, 150, 120]  # Adjust these values as needed

            for col, width in zip(columns, column_widths):
                tree.heading(col, text=col)
                tree.column(col, width=width, anchor="center")

            # Insert data into treeview
            for driver_id, data in performance_data.items():
                tree.insert("", "end", values=(
                    driver_id,
                    data["name"],
                    data["trips"],
                    data["on_time"],
                    f"{data['score']:.2f}"
                ))

            # Add scrollbar
            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)

            # Configure weight for auto-resize
            frame.columnconfigure(0, weight=1)
            frame.rowconfigure(0, weight=1)

            # Grid layout for better resizing
            tree.grid(row=0, column=0, sticky="nsew")
            scrollbar.grid(row=0, column=1, sticky="ns")

            # Bind resize event to window
            def on_resize(event):
                # Calculate new column widths
                total_width = tree.winfo_width()
                for col, base_width in zip(columns, column_widths):
                    ratio = base_width / sum(column_widths)
                    new_width = int(total_width * ratio)
                    tree.column(col, width=new_width)

            performance_window.bind("<Configure>", on_resize)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate performance: {e}")

    def clear_form():
        for widget in (id_entry,name_entry,cnic_entry,license_entry,contact_entry,address_entry,joining_date_entry,resigning_date_entry,trip_count_entry,salary_entry):
            widget.delete(0, tk.END)
        salary_status_combo.set("")
        trip_status_combo.set("")

    # Create themed root window
    root = ThemedTk(theme="arc")
    root.title("Driver Management System")
    root.geometry("1000x700")
    root.minsize(1024, 700)
    root.configure(bg='#f0f0f0')

    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Header frame
    header_frame = tk.Frame(root, bg="#1a237e", height=100)
    header_frame.grid(row=0, column=0, sticky="nsew")
    header_frame.grid_columnconfigure(0, weight=1)

    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a separate frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)
    # Navigation bar
    nav_frame = tk.Frame(header_frame, bg="#1a237e")
    nav_frame.pack(fill=tk.X, pady=5)
    ttk.Button(nav_frame, text="← Back", command=go_back).pack(side=tk.LEFT, padx=20)

    tk.Label(
        nav_frame,
        text="Driver Management System",
        font=("Helvetica", 16, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.LEFT, padx=20)

    # Main container
    main_container = ttk.Frame(root, padding="10")
    main_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

    # Search frame
    search_frame = ttk.LabelFrame(main_container, text="Search", padding="5")
    search_frame.pack(fill=tk.X, pady=(0, 10))
    search_criteria = ttk.Combobox(search_frame, values=["ID", "Name", "Contact", "License"], width=15)
    search_criteria.set("Name")
    search_criteria.pack(side=tk.LEFT, padx=5)
    search_entry = ttk.Entry(search_frame, width=30)
    search_entry.pack(side=tk.LEFT, padx=5)
    ttk.Button(search_frame, text="Search", command=search_drivers).pack(side=tk.LEFT, padx=5)

    # Form frame
    # Form frame with grid layout
    form_frame = ttk.LabelFrame(main_container, text="Driver Information", padding="10")
    form_frame.pack(fill=tk.X, pady=(0, 10))

    # Define form fields in 4 columns (to fit all 12 fields with 3 per column)
    form_fields = [
        # First column
        [("Driver ID:", id_entry := ttk.Entry(form_frame)),
         ("Name:", name_entry := ttk.Entry(form_frame)),
         ("CNIC:", cnic_entry := ttk.Entry(form_frame))],

        # Second column
        [("License Expiry:",
          license_entry := DateEntry(form_frame, width=18, background='darkblue', foreground='white', borderwidth=2,
                                     date_pattern='yyyy-mm-dd')),
         ("Contact:", contact_entry := ttk.Entry(form_frame)),
         ("Address:", address_entry := ttk.Entry(form_frame))],

        # Third column
        [("Salary:", salary_entry := ttk.Entry(form_frame)),
         ("Salary Status:", salary_status_combo := ttk.Combobox(form_frame, width=18, values=["Paid", "Unpaid"])),
         ("Joining Date:",
          joining_date_entry := DateEntry(form_frame, width=18, background='darkblue', foreground='white',
                                          borderwidth=2, date_pattern='yyyy-mm-dd'))],

        # Fourth column
        [("Resigning Date:", resigning_date_entry := ttk.Entry(form_frame)),
         ("Trip Count:", trip_count_entry := ttk.Entry(form_frame)),
         ("Trip Status:",
          trip_status_combo := ttk.Combobox(form_frame, width=18, values=["Available", "On Trip", "Off Duty"]))]
    ]

    # Create the grid layout
    for col, column_fields in enumerate(form_fields):
        for row, (label, widget) in enumerate(column_fields):
            # Label
            ttk.Label(form_frame, text=label).grid(
                row=row,
                column=col * 2,
                padx=5,
                pady=5,
                sticky='e'
            )
            # Entry/Combobox
            widget.grid(
                row=row,
                column=col * 2 + 1,
                padx=5,
                pady=5,
                sticky='w'
            )

    buttons_frame = ttk.Frame(main_container)
    buttons_frame.pack(fill=tk.X, pady=(0, 10))

    # Create sub-frames for each row (4 buttons per row)
    buttons_per_row = 9
    buttons = [
        ("Add Driver", add_driver),
        ("Update Driver", update_selected_driver),
        ("Delete Driver", delete_selected_driver),
        ("Check License Expiry", check_license_expiry),
        ("Salary Report", generate_salary_report),
        ("Manage Leave", manage_driver_leave),
        ("Track Status", track_driver_status),
        ("Record Payment", lambda: track_salary_payment(id_entry.get(), float(salary_entry.get()))),
        ("Refresh", refresh_treeview),
        ("Clear", clear_form),
        ("Export Reports", lambda: export_reports("salary")),
        ("Backup Database", backup_database),
        ("Performance Report", calculate_driver_performance)
    ]

    # Calculate number of rows needed
    num_rows = (len(buttons) + buttons_per_row - 1) // buttons_per_row

    # Create buttons row by row
    for row in range(num_rows):
        row_frame = ttk.Frame(buttons_frame)
        row_frame.pack(fill=tk.X, pady=2)

        # Calculate start and end indices for current row
        start_idx = row * buttons_per_row
        end_idx = min(start_idx + buttons_per_row, len(buttons))

        # Create buttons for current row
        for text, command in buttons[start_idx:end_idx]:
            btn = ttk.Button(row_frame, text=text, command=command)
            btn.pack(side=tk.LEFT, padx=3, fill=tk.X)

    # First, create and pack the footer frame
    # Table frame
    # Then create and pack the table frame
    table_frame = ttk.Frame(main_container)
    table_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

    # Create a frame for the table and scrollbars
    table_scroll_frame = ttk.Frame(table_frame)
    table_scroll_frame.pack(fill=tk.BOTH, expand=True)
    # Treeview and scrollbars configuration remain the same
    columns = (
    "Index", "ID", "Name", "CNIC", "License Expiry", "Address", "Contact", "Salary", "Salary Status", "Joining Date",
    "Resigning Date", "Trip Count", "Trip Status")
    driver_table = ttk.Treeview(table_scroll_frame, columns=columns, show='headings', height=15)

    # Configure column headings and widths
    for col in columns:
        driver_table.heading(col, text=col, anchor='center')
        driver_table.column(col, width=100, anchor='center')
    driver_table.bind('<<TreeviewSelect>>', on_driver_select)

    # Add vertical and horizontal scrollbars
    y_scrollbar = ttk.Scrollbar(table_scroll_frame, orient=tk.VERTICAL, command=driver_table.yview)
    driver_table.configure(yscrollcommand=y_scrollbar.set)

    x_scrollbar = ttk.Scrollbar(table_scroll_frame, orient=tk.HORIZONTAL, command=driver_table.xview)
    driver_table.configure(xscrollcommand=x_scrollbar.set)

    driver_table.grid(row=0, column=0, sticky='nsew')
    y_scrollbar.grid(row=0, column=1, sticky='ns')
    x_scrollbar.grid(row=1, column=0, sticky='ew')

    # Configure grid weights
    table_scroll_frame.grid_rowconfigure(0, weight=1)
    table_scroll_frame.grid_columnconfigure(0, weight=1)

    # Load initial data
    load_data_into_table()
    footer_frame = tk.Frame(root, bg="#1a237e", height=30)
    footer_frame.grid(row=2, column=0, sticky="nsew")
    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(pady=10)

    root.mainloop()



def truck_management_gui(user_role=""):
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()

    def load_trucks():
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM Trucks")
            trucks = cursor.fetchall()
            conn.close()
            return trucks
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load trucks: {e}")
            return []

    def save_truck(truck_data):
        try:
            with sqlite3.connect('TTMS.db', timeout=20) as conn:
                cursor = conn.cursor()

                # Check for duplicate truck ID
                cursor.execute("SELECT TruckID FROM Trucks WHERE TruckID = ?", (truck_data[0],))
                if cursor.fetchone():
                    messagebox.showerror("Error", "Truck ID already exists!")
                    return False

                cursor.execute("""
                    INSERT INTO Trucks (TruckID, Model, Status, Permit, MaintenanceSchedule, WeightCapacity)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, truck_data)

                conn.commit()
                messagebox.showinfo("Success", "Truck added successfully!")
                return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save truck: {e}")
            return False

    def search_trucks():
        search_term = search_entry.get().lower()
        search_by = search_criteria.get()

        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            if not search_term:
                load_data_into_table()
                return

            criteria_map = {
                "ID": "TruckID",
                "Model": "Model",
                "Status": "Status",
                "Permit": "Permit",
                "Capacity": "WeightCapacity"
            }

            column = criteria_map.get(search_by, "TruckID")
            cursor.execute(f"SELECT * FROM Trucks WHERE LOWER({column}) LIKE ?", (f'%{search_term}%',))
            filtered_trucks = cursor.fetchall()
            conn.close()

            load_data_into_table(filtered_trucks)

        except Exception as e:
            messagebox.showerror("Error", f"Search failed: {e}")

    def track_maintenance_expense(truck_id, amount, description, date=None):
        """Record maintenance expense for a truck"""
        conn = None
        try:
            conn = sqlite3.connect('TTMS.db', timeout=30)  # Increased timeout
            cursor = conn.cursor()

            date = date or datetime.now().strftime("%Y-%m-%d")

            # Get current odometer reading
            cursor.execute("SELECT Odometer FROM Trucks WHERE TruckID = ?", (truck_id,))
            result = cursor.fetchone()
            current_odometer = result[0] if result else 0

            # Insert maintenance record
            cursor.execute("""
                INSERT INTO MaintenanceHistory (TruckID, Amount, Description, Date, Odometer)
                VALUES (?, ?, ?, ?, ?)
            """, (truck_id, amount, description, date, current_odometer))

            conn.commit()
            messagebox.showinfo("Success", "Maintenance expense recorded successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to record maintenance expense: {e}")
        finally:
            if conn:
                conn.close()



    def track_fuel_expense(truck_id, amount, liters, date=None):
        """Record fuel expense for a truck"""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Create FuelHistory table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS FuelHistory (
                    TruckID TEXT,
                    Amount REAL,
                    Liters REAL,
                    Date TEXT
                )
            """)

            date = date or datetime.now().strftime("%Y-%m-%d")

            cursor.execute("""
                INSERT INTO FuelHistory (TruckID, Amount, Liters, Date)
                VALUES (?, ?, ?, ?)
            """, (truck_id, amount, liters, date))

            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Fuel expense recorded successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to record fuel expense: {e}")

    def update_odometer(truck_id, distance):
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Get current reading
            cursor.execute("SELECT Odometer FROM Trucks WHERE TruckID = ?", (truck_id,))
            current_reading = cursor.fetchone()[0] or 0
            new_reading = current_reading + distance

            # Update odometer
            cursor.execute("UPDATE Trucks SET Odometer = ? WHERE TruckID = ?",
                           (new_reading, truck_id))

            if new_reading >= 2000:
                messagebox.showwarning("Maintenance Required",
                                       f"Truck {truck_id} has reached {new_reading}km. Maintenance is required!")

            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update odometer: {e}")

    def reset_odometer(truck_id):
        """Reset odometer after maintenance"""
        conn = None
        try:
            conn = sqlite3.connect('TTMS.db', timeout=20)  # Add timeout
            cursor = conn.cursor()
            cursor.execute("UPDATE Trucks SET Odometer = 0 WHERE TruckID = ?", (truck_id,))
            messagebox.showinfo("Success",f"Odometer has been reset for truck:,{truck_id}")
            conn.commit()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to reset odometer: {e}")
        finally:
            if conn:
                conn.close()

    def get_current_odometer(truck_id):
        """Get current odometer reading for a truck"""
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            cursor.execute("SELECT Odometer FROM Trucks WHERE TruckID = ?", (truck_id,))
            result = cursor.fetchone()

            conn.close()
            return result[0] if result else 0
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get odometer reading: {e}")
            return 0

    def add_truck():
        truck_data = [
            id_entry.get(),
            model_entry.get(),
            status_combo.get(),
            permit_entry.get(),
            maintenance_entry.get(),
            capacity_entry.get(),
        ]
        if save_truck(truck_data):
            load_data_into_table()
            # Clear form fields
            for widget in (id_entry, model_entry, status_combo, permit_entry, maintenance_entry):
                widget.set('') if isinstance(widget, ttk.Combobox) else widget.delete(0, tk.END)

    def refresh_treeview():
        """Refresh the treeview with latest data from database"""
        try:
            load_data_into_table()
            messagebox.showinfo("Success", "Data refreshed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh data: {e}")

    def on_select(event):
        selected = truck_table.selection()
        if selected:
            values = truck_table.item(selected[0])['values']
            id_entry.delete(0, tk.END)
            id_entry.insert(0, values[0])
            model_entry.delete(0, tk.END)
            model_entry.insert(0, values[1])
            status_combo.set(values[2])
            permit_entry.delete(0, tk.END)
            permit_entry.insert(0, values[3])
            capacity_entry.delete(0,tk.END)
            capacity_entry.insert(0,values[4])
            maintenance_entry.delete(0, tk.END)
            maintenance_entry.insert(0, values[5])



    def clear_form():
        # Clear all entry widgets and reset comboboxes in the form
        for widget in (id_entry, model_entry, status_combo, permit_entry,capacity_entry, maintenance_entry,
                       maintenance_amount_entry, maintenance_desc_entry, maintenance_date_entry,
                       fuel_amount_entry, fuel_liters_entry, fuel_date_entry):
            if isinstance(widget, ttk.Combobox):
                widget.set('')
            else:
                widget.delete(0, tk.END)

    def load_data_into_table(data=None):
        for item in truck_table.get_children():
            truck_table.delete(item)
        trucks = data if data is not None else load_trucks()
        for truck in trucks:
            truck_table.insert("", "end", values=truck)

    def update_selected_truck():
        selected_item = truck_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No truck selected!")
            return

        if not messagebox.askyesno("Confirm", "Are you sure you want to update this truck?"):
            return

        truck_id = truck_table.item(selected_item)["values"][0]
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            cursor.execute("""
                UPDATE Trucks
                SET Model=?, Status=?, Permit=?, WeightCapacity=?, MaintenanceSchedule=?
                WHERE TruckID=?
            """, (model_entry.get(), status_combo.get(), permit_entry.get(),
                  capacity_entry.get(), maintenance_entry.get(), truck_id))

            conn.commit()
            conn.close()

            load_data_into_table()
            clear_form()
            messagebox.showinfo("Success", "Truck updated successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to update truck: {e}")

    def delete_selected_truck():
        selected_item = truck_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "No truck selected!")
            return

        if not messagebox.askyesno("Confirm", "Are you sure you want to delete this truck?"):
            return

        truck_id = truck_table.item(selected_item)["values"][0]
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            cursor.execute("DELETE FROM Trucks WHERE TruckID=?", (truck_id,))

            conn.commit()
            conn.close()

            load_data_into_table()
            clear_form()
            messagebox.showinfo("Success", "Truck deleted successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete truck: {e}")


    def generate_maintenance_report():
        """Generate a maintenance expense report for a date range using database"""
        try:
            # Connect to the SQLite database
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Prepare SQL query to fetch maintenance data
            query = """
            SELECT MaintenanceID, Amount, Description, Date, Odometer
            FROM MaintenanceHistory
            """
            cursor.execute(query)

            # Fetch all records within the date range
            report_data = cursor.fetchall()
            total_expense = sum(item[1] for item in report_data)  # Calculate total expense

            # Create report window
            report_window = tk.Toplevel()
            report_window.title("Maintenance Report")
            report_window.geometry("800x600")

            # Create a frame to hold the tree and scrollbars
            frame = ttk.Frame(report_window)
            frame.pack(fill=tk.BOTH, expand=True)

            # Create the scrollbars
            y_scroll = ttk.Scrollbar(frame, orient=tk.VERTICAL)
            x_scroll = ttk.Scrollbar(frame, orient=tk.HORIZONTAL)

            # Add report content
            report_tree = ttk.Treeview(frame, columns=("ID", "Amount", "Description", "Date", "Odometer"),
                                       yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

            # Configure scrollbars
            y_scroll.config(command=report_tree.yview)
            x_scroll.config(command=report_tree.xview)

            # Hide the default tree column
            report_tree['show'] = 'headings'

            # Configure columns and headings with center alignment
            columns = ["ID", "Amount", "Description", "Date", "Odometer"]
            for col in columns:
                report_tree.heading(col, text=col)
                report_tree.column(col, anchor='center', width=100)  # Set default width to 100

            # Insert data
            for item in report_data:
                report_tree.insert("", "end", values=item)

            # Layout with grid
            report_tree.grid(row=0, column=0, sticky='nsew')
            y_scroll.grid(row=0, column=1, sticky='ns')
            x_scroll.grid(row=1, column=0, sticky='ew')

            # Configure grid weights
            frame.grid_rowconfigure(0, weight=1)
            frame.grid_columnconfigure(0, weight=1)

            tk.Label(report_window,
                     text=f"Total Maintenance Expense: ${total_expense:.2f}",
                     font=("Helvetica", 12, "bold")).pack(pady=10)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {e}")
        finally:
            # Close the database connection
            conn.close()

    def export_to_csv():
        """Export truck data to CSV file"""
        try:
            trucks = load_trucks()
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")]
            )

            if filename:
                with open(filename, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(["ID", "Model", "Status", "Permit","Weight Capacity(KG)", "Maintenance Schedule", "Odometer"])
                    writer.writerows(trucks)
                messagebox.showinfo("Success", "Data exported successfully!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {e}")

    def schedule_maintenance_check():
        """Check for upcoming maintenance based on schedule"""
        try:
            trucks = load_trucks()
            upcoming_maintenance = []

            for truck in trucks:
                try:
                    maintenance_date = datetime.strptime(truck[5], "%Y-%m-%d")
                    current_date = datetime.now()
                    days_until = (maintenance_date - current_date).days

                    if 0 <= days_until <= 7:
                        upcoming_maintenance.append((truck[0], days_until))

                except Exception as e:
                    messagebox.showerror("Error", f"Error processing truck {truck[0]}: {str(e)}")

            if upcoming_maintenance:
                message = "Upcoming Maintenance:\n\n"
                for truck_id, days in upcoming_maintenance:
                    message += f"Truck {truck_id}: {days} days remaining\n"
                messagebox.showwarning("Maintenance Alert", message)
            else:
                messagebox.showinfo("Maintenance Status", "No upcoming maintenance scheduled for next 7 days")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to check maintenance schedule: {e}")

    root = ThemedTk(theme="arc")  # Using themed Tk for better appearance
    root.title("Truck Management System")
    root.geometry("1000x700")
    root.minsize(1024, 700)
    root.configure(bg='#f0f0f0')

    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Header frame
    header_frame = tk.Frame(root, bg="#1a237e", height=100)
    header_frame.grid(row=0, column=0, sticky="nsew")
    header_frame.grid_columnconfigure(0, weight=1)

    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a separate frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)
    # Navigation bar
    nav_frame = tk.Frame(header_frame, bg="#1a237e")
    nav_frame.pack(fill=tk.X, pady=5)
    ttk.Button(nav_frame, text="← Back", command=go_back).pack(side=tk.LEFT, padx=20)

    tk.Label(
        nav_frame,
        text="Truck Management System",
        font=("Helvetica", 16, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.LEFT, padx=20)

    # Main container
    main_container = ttk.Frame(root, padding="10")
    main_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

    # Search frame
    search_frame = ttk.LabelFrame(main_container, text="Search", padding="5")
    search_frame.pack(fill=tk.X, pady=(0, 10))
    search_criteria = ttk.Combobox(search_frame, values=["ID", "Model", "Status", "Permit", "Capacity"], width=15)
    search_criteria.set("ID")
    search_criteria.pack(side=tk.LEFT, padx=5)

    search_entry = ttk.Entry(search_frame, width=30)
    search_entry.pack(side=tk.LEFT, padx=5)

    ttk.Button(search_frame, text="Search", command=search_trucks).pack(side=tk.LEFT, padx=5)

    # Modify the form_frame section in truck_management_gui:

    # Input form frame
    form_frame = ttk.LabelFrame(main_container, text="Truck Information", padding="10")
    form_frame.pack(fill=tk.X, pady=(0, 10))

    # Create separate frames for different sections
    truck_info_frame = ttk.Frame(form_frame)
    truck_info_frame.pack(side=tk.LEFT, padx=10)

    maintenance_frame = ttk.LabelFrame(form_frame, text="Maintenance Records", padding="5")
    maintenance_frame.pack(side=tk.LEFT, padx=10)

    fuel_frame = ttk.LabelFrame(form_frame, text="Fuel Records", padding="5")
    fuel_frame.pack(side=tk.LEFT, padx=10)

    # Truck Information fields
    form_fields = [
        ("Truck ID:", id_entry := ttk.Entry(truck_info_frame)),
        ("Model:", model_entry := ttk.Entry(truck_info_frame)),
        ("Status:", status_combo := ttk.Combobox(truck_info_frame, width=18,
                                                 values=["Operational", "Under Maintenance", "Retired"])),
        ("Permit Number:", permit_entry := ttk.Entry(truck_info_frame)),
        ("Maintenance Schedule:",
         maintenance_entry := DateEntry(truck_info_frame, width=18, background='darkblue', foreground='white',
                                        borderwidth=2, date_pattern='yyyy-mm-dd')),
        ("Weight Capacity(KG):", capacity_entry := ttk.Entry(truck_info_frame))
    ]

    for i, (label, widget) in enumerate(form_fields):
        ttk.Label(truck_info_frame, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
        widget.grid(row=i, column=1, padx=5, pady=5, sticky='w')

    # Maintenance Record fields
    maintenance_fields = [
        ("Amount:", maintenance_amount_entry := ttk.Entry(maintenance_frame)),
        ("Description:", maintenance_desc_entry := ttk.Entry(maintenance_frame)),
        ("Date:",
         maintenance_date_entry := DateEntry(maintenance_frame, width=18, background='darkblue', foreground='white',
                                             borderwidth=2, date_pattern='yyyy-mm-dd')),
    ]

    for i, (label, widget) in enumerate(maintenance_fields):
        ttk.Label(maintenance_frame, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
        widget.grid(row=i, column=1, padx=5, pady=5, sticky='w')

    # Add maintenance record button
    ttk.Button(maintenance_frame,
               text="Record Maintenance",
               command=lambda: track_maintenance_expense(
                   id_entry.get(),
                   float(maintenance_amount_entry.get()),
                   maintenance_desc_entry.get(),
                   maintenance_date_entry.get() or None
               )).grid(row=len(maintenance_fields), column=0, columnspan=1,padx=1 ,pady=5)

    ttk.Button(maintenance_frame,
               text="Reset Odometer",
               command=lambda: reset_odometer(id_entry.get())
               ).grid(row=len(maintenance_fields), column=1, columnspan=2,padx=2, pady=5)

    # Fuel Record fields
    fuel_fields = [
        ("Amount:", fuel_amount_entry := ttk.Entry(fuel_frame)),
        ("Liters:", fuel_liters_entry := ttk.Entry(fuel_frame)),
        ("Date:",
         fuel_date_entry := DateEntry(fuel_frame, width=18, background='darkblue', foreground='white', borderwidth=2,
                                      date_pattern='yyyy-mm-dd')),
    ]

    for i, (label, widget) in enumerate(fuel_fields):
        ttk.Label(fuel_frame, text=label).grid(row=i, column=0, padx=5, pady=5, sticky='e')
        widget.grid(row=i, column=1, padx=5, pady=5, sticky='w')

    # Add fuel record button
    ttk.Button(fuel_frame,
               text="Record Fuel",
               command=lambda: track_fuel_expense(
                   id_entry.get(),
                   float(fuel_amount_entry.get()),
                   float(fuel_liters_entry.get()),
                   fuel_date_entry.get() or None
               )).grid(row=len(fuel_fields), column=0, columnspan=2, pady=5)
    # Modify the buttons list to remove redundant buttons
    buttons_frame = ttk.Frame(main_container)
    buttons_frame.pack(fill=tk.X, pady=(0, 10))
    for text, command in [
        ("Add Truck", add_truck),
        ("Update Truck", lambda: update_selected_truck()),
        ("Delete Truck", lambda: delete_selected_truck()),
        ("Clear Form", lambda: clear_form()),
        ("Update Odometer", lambda: update_odometer(id_entry.get(), 0)),
        ("Refresh", refresh_treeview),
        ("Maintenance Report", lambda: generate_maintenance_report()),
        ("Export Data", export_to_csv),
        ("Check Maintenance", schedule_maintenance_check),
    ]:
        ttk.Button(buttons_frame, text=text, command=command).pack(side=tk.LEFT, padx=5)

    # Table frame with scrollbar
    table_frame = ttk.Frame(main_container)
    table_frame.pack(fill=tk.BOTH, expand=True)

    # Create a frame for the table and scrollbars
    table_scroll_frame = ttk.Frame(table_frame)
    table_scroll_frame.pack(fill=tk.BOTH, expand=True)

    # Create the treeview
    columns = ("ID", "Model", "Status", "Permit", "Weight Capacity(KG)", "Maintenance Schedule", "Odometer")
    truck_table = ttk.Treeview(table_scroll_frame, columns=columns, show='headings', height=15)

    # Configure column headings and widths
    for col in columns:
        truck_table.heading(col, text=col, anchor='center')
        truck_table.column(col, width=150, anchor='center')
    truck_table.bind('<<TreeviewSelect>>', on_select)
    # Add vertical scrollbar
    y_scrollbar = ttk.Scrollbar(table_scroll_frame, orient=tk.VERTICAL, command=truck_table.yview)
    truck_table.configure(yscrollcommand=y_scrollbar.set)
    # add horizontal scrollbar
    x_scrollbar = ttk.Scrollbar(table_scroll_frame, orient=tk.HORIZONTAL, command=truck_table.xview)
    truck_table.configure(xscrollcommand=x_scrollbar.set)

    # Grid layout for table and scrollbars
    truck_table.grid(row=0, column=0, sticky='nsew')
    y_scrollbar.grid(row=0, column=1, sticky='ns')
    x_scrollbar.grid(row=1, column=0, sticky='ew')

    # Configure grid weights
    table_scroll_frame.grid_rowconfigure(0, weight=1)
    table_scroll_frame.grid_columnconfigure(0, weight=1)

    # Bind table selection event

    load_data_into_table()
    footer_frame = tk.Frame(root, bg="#1a237e", height=30)
    footer_frame.grid(row=2, column=0, sticky="nsew")
    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(pady=10)

    root.mainloop()



def order_management_gui(user_role=""):
    import datetime
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()
        elif user_role == "Dispatcher":
            dispatcher_dashboard()

    def refresh_treeview():
        """Refresh the treeview with latest data from the database"""
        try:
            load_data_into_table()
            messagebox.showinfo("Success", "Data refreshed successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to refresh data: {e}")

    def load_orders():
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute("""
                SELECT rowid, * FROM Orders
            """)
            orders = cursor.fetchall()
            conn.close()
            return orders
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load orders: {e}")
            return []

    def save_order(order_data):
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Check if order ID already exists
            cursor.execute("SELECT OrderID FROM Orders WHERE OrderID = ?", (order_data[0],))
            if cursor.fetchone():
                messagebox.showerror("Error", "Order ID already exists!")
                conn.close()
                return False

            # Calculate remaining amount
            total_amount = float(order_data[11])
            paid_amount = float(order_data[12])
            remaining = total_amount - paid_amount

            cursor.execute("""
                INSERT INTO Orders (
                    OrderID, OrderName, CustomerName, Contact, Pickup,
                    Destination, Region, Distance, Status, Weight, GST,
                    TotalAmount, PaidAmount, AmountStatus, RemainingAmount, OrderDate
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, order_data)

            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Order added successfully!")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save order: {e}")
            return False

    def delete_order(order_id):
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute("DELETE FROM Orders WHERE rowid = ?", (order_id,))
            conn.commit()
            conn.close()
            load_data_into_table()
            messagebox.showinfo("Success", "Order deleted successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete order: {e}")

    def update_order(order_id, updated_data):
        try:
            conn = sqlite3.connect('TTMS.db')  # Changed from 'truckflow.db' to 'TTMS.db'
            cursor = conn.cursor()

            # Calculate remaining amount
            total_amount = float(updated_data[11])
            paid_amount = float(updated_data[12])
            remaining = total_amount - paid_amount

            cursor.execute("""
                UPDATE Orders SET
                    OrderID=?, OrderName=?, CustomerName=?, Contact=?,
                    Pickup=?, Destination=?, Region=?, Distance=?,
                    Status=?, Weight=?, GST=?, TotalAmount=?,
                    PaidAmount=?, AmountStatus=?, RemainingAmount=?, OrderDate=?
                WHERE rowid=?
            """, (*updated_data, order_id))

            conn.commit()
            conn.close()
            load_data_into_table()
            messagebox.showinfo("Success", "Order updated successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update order: {e}")

    def add_order():
        try:
            # Validate required fields
            if not all([order_id_entry.get(), customer_entry.get(),
                        total_amount_entry.get(), paid_amount_entry.get()]):
                raise ValueError("Please fill all required fields")

            total_amount = float(total_amount_entry.get())
            paid_amount = float(paid_amount_entry.get())
            remaining_amount = float(total_amount - paid_amount)

            # Validate amounts
            if total_amount <= 0:
                raise ValueError("Total amount must be greater than 0")
            if paid_amount < 0:
                raise ValueError("Paid amount cannot be negative")
            if paid_amount > total_amount:
                raise ValueError("Paid amount cannot exceed total amount!")

            order_data = [
                order_id_entry.get(),  # OrderID
                order_name_entry.get(),  # OrderName
                customer_entry.get(),  # CustomerName
                contact_entry.get(),  # Contact
                pickup_entry.get(),  # Pickup
                destination_entry.get(),  # Destination
                region_entry.get(),
                distance_entry.get(),  # Distance
                status_combo.get(),  # Status
                weight_entry.get(),  # Weight(KG)
                gst_entry.get(),  # GST%
                total_amount,  # TotalAmount
                paid_amount,  # PaidAmount
                payment_combo.get(),  # AmountStatus
                remaining_amount,  # RemainingAmount
                order_date_entry.get()

            ]

            if save_order(order_data):
                load_data_into_table()
                clear_form()
        except ValueError as ve:
            messagebox.showerror("Input Error", str(ve))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add order: {e}")

    def delete_selected_order():
        selected = order_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select an order to delete.")
            return

        row_id = int(order_table.item(selected[0])["values"][0])
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this order?"):
            delete_order(row_id)

    def update_selected_order():
        selected = order_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select an order to update.")
            return

        try:
            total_amount = float(total_amount_entry.get())
            paid_amount = float(paid_amount_entry.get())
            remaining_amount = float(total_amount - paid_amount)
            if paid_amount > total_amount:
                raise ValueError("Paid amount cannot exceed total amount!")

            updated_data = [
                order_id_entry.get(),
                order_name_entry.get(),
                customer_entry.get(),
                contact_entry.get(),
                pickup_entry.get(),
                destination_entry.get(),
                region_entry.get(),
                distance_entry.get(),
                status_combo.get(),
                weight_entry.get(),
                gst_entry.get(),
                total_amount,
                paid_amount,
                payment_combo.get(),
                remaining_amount,
                order_date_entry.get()


            ]

            row_id = int(order_table.item(selected[0])["values"][0])
            update_order(row_id, updated_data)
        except ValueError as ve:
            messagebox.showerror("Input Error", f"Invalid input: {ve}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update order: {e}")

    def search_orders():
        search_term = search_entry.get().lower()
        search_by = search_criteria.get()

        if not search_term:
            load_data_into_table()
            return

        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            criteria_mapping = {
                "Order ID": "OrderID",
                "Customer Name": "CustomerName",
                "Destination": "Destination",
                "Status": "Status",
                "Payment Status": "AmountStatus",
                "Order Date": "OrderDate"
            }
            field = criteria_mapping.get(search_by, "OrderID")
            cursor.execute(f"""
                        SELECT rowid, * FROM Orders
                        WHERE LOWER({field}) LIKE ?
                    """, (f'%{search_term}%',))

            filtered_orders = cursor.fetchall()
            conn.close()
            load_data_into_table(filtered_orders)

        except Exception as e:
            messagebox.showerror("Error", f"Search failed: {e}")

    def load_data_into_table(data=None):
        order_table.delete(*order_table.get_children())
        orders = data if data is not None else load_orders()

        for order in orders:
            try:
                # Calculate remaining amount
                total_amount = float(order[12])  # Adjusted index for database columns
                paid_amount = float(order[13])
                remaining = "{:.2f}".format(total_amount - paid_amount)

                # Insert values into treeview
                order_table.insert('', 'end', values=(*order, remaining))
            except (ValueError, IndexError) as e:
                print(f"Error processing order: {e}")
                continue

    def clear_form():
        for widget in (order_id_entry,order_name_entry, customer_entry, contact_entry,order_date_entry,
                       pickup_entry, destination_entry,region_entry, distance_entry,gst_entry,weight_entry,
                       total_amount_entry, paid_amount_entry):
            widget.delete(0, tk.END)
        status_combo.set("")
        payment_combo.set("")

    def export_to_pdf():
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib.units import inch
            from datetime import datetime

            orders = load_orders()
            if not orders:
                messagebox.showwarning("Warning", "No data to export")
                return

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"exports/ordersdata_{timestamp}.pdf"

            page_size = landscape(letter)

            doc = SimpleDocTemplate(filename,
                                    pagesize=page_size,
                                    leftMargin=0.25 * inch,
                                    rightMargin=0.25 * inch,
                                    topMargin=0.25 * inch,
                                    bottomMargin=0.25 * inch)

            elements = []

            # Remove first column from headers
            modified_columns = order_columns[1:]

            # Create table data without first column
            table_data = [modified_columns]
            for order in orders:
                table_data.append([str(x) for x in order[1:]])  # Skip first column

            page_width = page_size[0] - (0.5 * inch)
            num_columns = len(modified_columns)

            # Adjust column widths to better fit content
            col_widths = []
            for i in range(num_columns):
                max_width = max(len(str(row[i])) for row in table_data) * 10  # Increased multiplier from 7 to 10
                min_width = 70  # Increased minimum width from 55 to 70
                col_widths.append(min(max(max_width, min_width), page_width / num_columns))

            table = Table(table_data, colWidths=col_widths, repeatRows=1)

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),  # Reduced font size from 9 to 8
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Increased padding from 8 to 12
                ('TOPPADDING', (0, 0), (-1, 0), 12),  # Added top padding
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('WORDWRAP', (0, 0), (-1, -1), True),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),  # Increased from 4 to 6
                ('RIGHTPADDING', (0, 0), (-1, -1), 6)  # Increased from 4 to 6
            ]))

            elements.append(table)
            doc.build(elements)
            messagebox.showinfo("Success", f"PDF exported successfully as {filename}!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export PDF: {e}")

    def export_to_excel():
        """Export order data to Excel with formatting"""
        try:
            from datetime import datetime

            orders = load_orders()
            if not orders:
                messagebox.showwarning("Warning", "No data to export")
                return

            # Create filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"exports/ordersdata_{timestamp}.xlsx"

            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Add headers (skip first column)
            for col, header in enumerate(order_columns[1:], 1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Add data (skip first column)
            for row, order in enumerate(orders, 2):
                for col, value in enumerate(order[1:], 1):
                    sheet.cell(row=row, column=col, value=value)

            workbook.save(filename)
            messagebox.showinfo("Success", f"Excel file exported successfully as {filename}!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export to Excel: {e}")

    def calculate_statistics():
        """Calculate and display comprehensive order statistics with visualization"""
        try:
            orders = load_orders()
            if not orders:
                messagebox.showwarning("Warning", "No data available")
                return

            # Calculate basic metrics
            total_orders = len(orders)
            total_amount = sum(float(order[12]) for order in orders if order[11])
            pending_orders = len([o for o in orders if o[9] == "Pending"])
            delivered_orders = len([o for o in orders if o[9] == "Delivered"])
            avg_order_value = total_amount / total_orders if total_orders > 0 else 0
            delivery_rate = (delivered_orders / total_orders * 100) if total_orders > 0 else 0

            # Create statistics window
            stats_window = tk.Toplevel(root)
            stats_window.title("Order Statistics Dashboard")
            stats_window.geometry("600x400")
            stats_window.resizable(False, False)

            # Style configuration
            style = ttk.Style()
            style.configure("Stats.TLabel", padding=5, font=('Helvetica', 10))

            # Create main frame
            main_frame = ttk.Frame(stats_window, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Statistics display
            stats_frame = ttk.LabelFrame(main_frame, text="Key Metrics", padding="10")
            stats_frame.pack(fill=tk.X, padx=5, pady=5)

            metrics = [
                ("Total Orders:", f"{total_orders:,}"),
                ("Total Revenue:", f"/-{total_amount:,.2f}"),
                ("Average Order Value:", f"/-{avg_order_value:,.2f}"),
                ("Pending Orders:", f"{pending_orders:,}"),
                ("Delivered Orders:", f"{delivered_orders:,}"),
                ("Delivery Rate:", f"{delivery_rate:.1f}%")
            ]

            # Create grid layout for metrics
            for i, (label, value) in enumerate(metrics):
                row = i // 2
                col = i % 2 * 2
                ttk.Label(stats_frame, text=label, style="Stats.TLabel").grid(row=row, column=col, sticky='e', padx=5)
                ttk.Label(stats_frame, text=value, style="Stats.TLabel").grid(row=row, column=col + 1, sticky='w',
                                                                              padx=5)
            export_frame = ttk.LabelFrame(main_frame, text="Export Options", padding="10")
            export_frame.pack(fill=tk.X, padx=5, pady=5)

            # Add radio buttons for export type
            export_type = tk.StringVar(value="pdf")
            ttk.Radiobutton(export_frame, text="PDF", value="pdf", variable=export_type).pack(side=tk.LEFT, padx=10)
            ttk.Radiobutton(export_frame, text="Excel", value="xlsx", variable=export_type).pack(side=tk.LEFT, padx=10)
            ttk.Radiobutton(export_frame, text="CSV", value="csv", variable=export_type).pack(side=tk.LEFT, padx=10)

            button_frame = ttk.Frame(main_frame)
            button_frame.pack(pady=10)

            # Add refresh button
            ttk.Button(button_frame, text="Refresh",command=lambda: calculate_statistics()).pack(side=tk.LEFT, padx=5)

            # Modified export button
            ttk.Button(button_frame, text="Export",command=lambda: export_statistics(metrics, export_type.get())).pack(side=tk.LEFT, padx=5)

            # Add print button
            ttk.Button(button_frame, text="Print",command=lambda: print_statistics(stats_frame)).pack(side=tk.LEFT, padx=5)


        except Exception as e:

            messagebox.showerror("Error", f"Failed to calculate statistics: {str(e)}")

            logging.error(f"Statistics calculation error: {str(e)}")

    def export_statistics(metrics, format_type):
        """Export statistics to format type in a fixed directory"""
        try:

            # Generate filename with timestamp
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(f"exports/order_statistics_{timestamp}.{format_type}")

            if format_type == "csv":
                with open(filename, 'w', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerow(["Metric", "Value"])
                    writer.writerows(metrics)

            elif format_type == "xlsx":
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Metric", "Value"])
                for metric in metrics:
                    ws.append(metric)
                wb.save(filename)

            elif format_type == "pdf":
                doc = SimpleDocTemplate(filename, pagesize=letter)
                elements = []
                data = [["Metric", "Value"]] + metrics
                table = Table(data)
                elements.append(table)
                doc.build(elements)

            messagebox.showinfo("Success", f"Statistics exported successfully to:\n{filename}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export statistics: {str(e)}")


    def print_statistics(frame):
        """Print the statistics"""
        try:
            # Create a temporary file for printing
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
            with open(temp_file.name, 'w') as f:
                f.write("Order Statistics Dashboard\n\n")
                for child in frame.winfo_children():
                    if isinstance(child, ttk.Label):
                        f.write(f"{child.cget('text')}\n")

            # Open print dialog
            os.startfile(temp_file.name, "print")

        except Exception as e:
            messagebox.showerror("Print Error", f"Failed to print statistics: {str(e)}")

    def print_selected_order():
        """Print the details of the selected order"""
        selected = order_table.selection()
        if not selected:
            messagebox.showerror("Error", "Please select an order to print.")
            return

        try:
            # Get selected order details
            values = order_table.item(selected[0])['values']

            # Create temporary file for printing
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
            with open(temp_file.name, 'w') as f:
                # Write header
                f.write("TruckFlow Solutions - Order Details\n")
                f.write("=" * 50 + "\n\n")

                # Write order details
                order_details = [
                    ("Order ID:", values[1]),
                    ("Order Name:", values[2]),
                    ("Customer Name:", values[3]),
                    ("Contact:", values[4]),
                    ("Pickup Location:", values[5]),
                    ("Destination:", values[6]),
                    ("Region:", values[7]),
                    ("Distance:", values[8]),
                    ("Status:", values[9]),
                    ("Weight(KG):", values[10]),
                    ("GST%:", values[11]),
                    ("Total Amount:", f"/-{values[12]}"),
                    ("Paid Amount:", f"/-{values[13]}"),
                    ("Payment Status:", values[14]),
                    ("Remaining Amount:", f"/-{values[15]}"),
                    ("Order Date:", values[16])
                ]

                for label, value in order_details:
                    f.write(f"{label:<20} {value}\n")

            # Open print dialog
            os.startfile(temp_file.name, "print")
            messagebox.showinfo("Success", "Order Recipt sent to printer")

        except Exception as e:
            messagebox.showerror("Print Error", f"Failed to print order: {str(e)}")

    def on_select_order(event):
        selected_item = order_table.selection()
        if selected_item:
            values = order_table.item(selected_item[0])['values']
            order_id_entry.delete(0, tk.END)
            order_id_entry.insert(0, values[1])
            order_name_entry.delete(0, tk.END)
            order_name_entry.insert(0, values[2])
            customer_entry.delete(0, tk.END)
            customer_entry.insert(0, values[3])
            contact_entry.delete(0, tk.END)
            contact_entry.insert(0, values[4])
            pickup_entry.delete(0, tk.END)
            pickup_entry.insert(0, values[5])
            destination_entry.delete(0, tk.END)
            destination_entry.insert(0, values[6])
            region_entry.delete(0, tk.END)
            region_entry.insert(0, values[7])
            distance_entry.delete(0, tk.END)
            distance_entry.insert(0, values[8])
            status_combo.set(values[9])
            weight_entry.delete(0, tk.END)
            weight_entry.insert(0, values[10])
            gst_entry.delete(0, tk.END)
            gst_entry.insert(0, values[11])
            total_amount_entry.delete(0, tk.END)
            total_amount_entry.insert(0, values[12])
            paid_amount_entry.delete(0, tk.END)
            paid_amount_entry.insert(0, values[13])
            payment_combo.set(values[14])
            order_date_entry.delete(0,tk.END)
            order_date_entry.insert(0,values[16])
    def validate_contact(P):
        return all(c.isdigit() or c in '+-() ' for c in P)
    def calculate_total_with_gst(*args):
        try:
            amount = float(total_amount_entry.get() or 0)
            gst_percent = float(gst_entry.get() or 0)
            total_with_gst = amount + (amount * gst_percent / 100)
            total_amount_entry.delete(0, tk.END)
            total_amount_entry.insert(0, f"{total_with_gst:.2f}")
        except ValueError:
            pass

    # Root window
    root = ThemedTk(theme="arc")
    root.title("Order Management System")
    root.geometry("1000x700")
    root.minsize(1024, 700)
    root.configure(bg='#f0f0f0')

    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Header frame
    header_frame = tk.Frame(root, bg="#1a237e", height=100)
    header_frame.grid(row=0, column=0, sticky="nsew")
    header_frame.grid_columnconfigure(0, weight=1)

    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a separate frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)
    # Navigation bar
    nav_frame = tk.Frame(header_frame, bg="#1a237e")
    nav_frame.pack(fill=tk.X, pady=5)
    ttk.Button(nav_frame, text="← Back", command=go_back).pack(side=tk.LEFT, padx=20)

    tk.Label(
        nav_frame,
        text="Order Management System",
        font=("Helvetica", 16, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.LEFT, padx=20)

    # Main container
    main_container = ttk.Frame(root, padding="10")
    main_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

    # Search frame
    search_frame = ttk.LabelFrame(main_container, text="Search", padding="5")
    search_frame.pack(fill=tk.X, pady=(0, 10))
    search_criteria = ttk.Combobox(search_frame,
                                   values=["Order ID", "Customer Name", "Destination", "Status", "Payment Status"],
                                   width=15)
    search_criteria.set("Order ID")
    search_criteria.pack(side=tk.LEFT, padx=5)

    search_entry = ttk.Entry(search_frame)
    search_entry.pack(side=tk.LEFT, padx=5)
    ttk.Button(search_frame, text="Search", command=search_orders).pack(side=tk.LEFT, padx=5)

    # Form
    form_frame = ttk.LabelFrame(main_container, text="Order Details")
    form_frame.pack(fill=tk.X, padx=10, pady=10)

    fields = [
        ("Order ID", order_id_entry := ttk.Entry(form_frame)),
        ("Order Date",
         order_date_entry := DateEntry(form_frame, width=18, background='darkblue', foreground='white', borderwidth=2,
                                       date_pattern='yyyy-mm-dd')),
        ("Order Name", order_name_entry := ttk.Entry(form_frame)),
        ("Customer Name", customer_entry := ttk.Entry(form_frame)),
        ("Contact", contact_entry := ttk.Entry(form_frame)),
        ("Pickup Location", pickup_entry := ttk.Entry(form_frame)),
        ("Destination", destination_entry := ttk.Entry(form_frame)),
        ("Region", region_entry := ttk.Entry(form_frame)),
        ("Distance", distance_entry := ttk.Entry(form_frame)),
        ("Weight(KG)", weight_entry := ttk.Entry(form_frame)),
        ("GST(%)", gst_entry := ttk.Entry(form_frame)),
        ("Status", status_combo := ttk.Combobox(form_frame, width=18,
                                                values=["Pending", "In Transit", "Delivered", "Cancelled"])),
        ("Total Amount", total_amount_entry := ttk.Entry(form_frame)),
        ("Paid Amount", paid_amount_entry := ttk.Entry(form_frame)),
        ("Payment Status",
         payment_combo := ttk.Combobox(form_frame, width=18, values=["Pending", "Partial", "Completed"]))
    ]

    for i, (label, widget) in enumerate(fields):
        ttk.Label(form_frame, text=label).grid(row=i // 3, column=(i % 3) * 2, padx=5, pady=5, sticky='e')
        widget.grid(row=i // 3, column=(i % 3) * 2 + 1, padx=5, pady=5, sticky='w')
    # Buttons
    buttons_frame = ttk.Frame(main_container)
    buttons_frame.pack(fill=tk.X, padx=10, pady=10)

    ttk.Button(buttons_frame, text="Add Order", command=add_order).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Update Order", command=update_selected_order).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Delete Order", command=delete_selected_order).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Clear", command=clear_form).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Refresh", command=refresh_treeview).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Export PDF", command=export_to_pdf).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Export Excel", command=export_to_excel).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Statistics", command=calculate_statistics).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Print Order", command=print_selected_order).pack(side=tk.LEFT, padx=5)


    # Table
    table_frame = ttk.Frame(main_container)
    table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    table_scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL)
    table_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    order_columns = [
        "Row",
        "OrderID",
        "OrderName",
        "CustomerName",
        "Contact",
        "Pickup",
        "Destination",
        "Region",
        "Distance",
        "Status",
        "Weight(KG)",
        "GST%",
        "TotalAmount",
        "PaidAmount",
        "AmountStatus",
        "RemainingAmount",
        "Order Date"
    ]

    # Add horizontal scrollbar
    h_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
    h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    order_table = ttk.Treeview(table_frame, columns=order_columns, show="headings", yscrollcommand=table_scrollbar.set,
                               xscrollcommand=h_scrollbar.set, height=15)

    # Configure both scrollbars
    table_scrollbar.config(command=order_table.yview)
    h_scrollbar.config(command=order_table.xview)

    # Adjust column widths and make them resizable
    for col in order_columns:
        order_table.heading(col, text=col)
        order_table.column(col, width=100, minwidth=50)  # Set minimum width

    order_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    table_scrollbar.config(command=order_table.yview)

    # Bind table events

    contact_validation = root.register(validate_contact)
    contact_entry.config(validate='key', validatecommand=(contact_validation, '%P'))

    # Add automatic GST calculation

    gst_entry.bind('<FocusOut>', calculate_total_with_gst)

    order_table.bind("<<TreeviewSelect>>", on_select_order)

    # Load initial data
    load_data_into_table()
    footer_frame = tk.Frame(root, bg="#1a237e", height=30)
    footer_frame.grid(row=2, column=0, sticky="ew")

    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.BOTTOM, pady=10)

    root.mainloop()



# Main Dispatch Management GUI
def dispatch_management_gui(user_role=""):
    import datetime
    from datetime import datetime

    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()
        elif user_role == "Dispatcher":
            dispatcher_dashboard()

    # Logging configuration
    logging.basicConfig(filename='dispatch_system.log', level=logging.ERROR)

    def log_error(message):
        logging.error(f"{datetime.datetime.now()}: {message}")

    def is_valid_estimated_time(dispatch_time, estimated_time):
        try:
            dispatch_time_obj = datetime.datetime.strptime(dispatch_time, "%Y-%m-%d %H:%M")
            estimated_time_obj = datetime.datetime.strptime(estimated_time, "%Y-%m-%d %H:%M")
            return estimated_time_obj > dispatch_time_obj
        except ValueError:
            return False

    def is_driver_compliant(driver_id):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT LicenseExpiry
                    FROM Drivers
                    WHERE DriverID = ?
                """, (driver_id,))
                result = cursor.fetchone()
                conn.close()

                if result:
                    expiry_date = datetime.strptime(result[0], "%Y-%m-%d")
                    return expiry_date > datetime.now()
                return False
        except Exception as e:
            messagebox.showerror("Error", f"Failed to check driver compliance: {e}")
            log_error(f"Failed to check driver compliance: {e}")
            return False



    def create_db_connection():
        try:
            conn = sqlite3.connect('TTMS.db')
            return conn
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to connect to database: {e}")
            return None

    def load_dispatch_data():
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT d.DispatchID, d.OrderID, d.DriverID, d.TruckID,
                           d.DispatchTime, d.Status, d.EstimatedDeliveryTime
                    FROM Dispatch d
                    ORDER BY d.DispatchTime DESC
                """)
                dispatches = cursor.fetchall()
                conn.close()
                return dispatches
            return []
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load dispatch data: {e}")
            log_error("Error",f"Failed to load dispatch:{e}")
            return []

    def update_dispatch_status(dispatch_id, new_status):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                        UPDATE Dispatch
                        SET Status = ?
                        WHERE DispatchID = ?
                    """, (new_status, dispatch_id))

                cursor.execute("""
                        UPDATE Orders
                        SET Status = ?
                        WHERE OrderID = (
                            SELECT OrderID FROM Dispatch WHERE DispatchID = ?
                        )
                    """, (new_status, dispatch_id))

                conn.commit()
                conn.close()

                for item in dispatch_table.get_children():
                    if dispatch_table.item(item)['values'][0] == dispatch_id:
                        dispatch_table.set(item, "Status", new_status)
                        update_progress_bar(item, new_status)
                        break
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update status: {e}")
            log_error(f"Failed to update status: {e}")


    def delete_dispatch(dispatch_id):  # Added dispatch_id parameter
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()

                # Get driver and truck IDs before deletion
                cursor.execute("""
                        SELECT DriverID, TruckID, OrderID
                        FROM Dispatch
                        WHERE DispatchID = ?
                    """, (dispatch_id,))
                result = cursor.fetchone()

                if result is None:
                    messagebox.showerror("Error", "Dispatch not found")
                    return

                driver_id, truck_id, order_id = result

                # Delete the dispatch
                cursor.execute("DELETE FROM Dispatch WHERE DispatchID = ?", (dispatch_id,))

                # Update related statuses
                cursor.execute("UPDATE Orders SET Status = 'Pending' WHERE OrderID = ?", (order_id,))
                cursor.execute("UPDATE Drivers SET Status = 'Available' WHERE DriverID = ?", (driver_id,))
                cursor.execute("UPDATE Trucks SET Status = 'Operational' WHERE TruckID = ?", (truck_id,))

                conn.commit()
                messagebox.showinfo("Success", "Dispatch deleted successfully")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            if conn:
                conn.close()

        load_data_into_table()

    # Modify the button configuration section to include dispatch_id retrieval
    def delete_selected_dispatch():
        selected_item = dispatch_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a dispatch to delete")
            return
        dispatch_id = dispatch_table.item(selected_item)['values'][0]
        delete_dispatch(dispatch_id)

    def create_progress_column():
        dispatch_table["dispatch_columns"] = dispatch_columns + ("Progress",)
        dispatch_table.heading("Progress", text="Progress")
        dispatch_table.column("Progress", width=150, anchor="center")

    def update_progress_bar(item_id, status):
        progress_values = {
            "Pending": 0,
            "In Transit": 33,
            "Out for Delivery": 66,
            "Delivered": 100,
            "Completed": 100
        }
        print(f"Received item_id: {item_id}, status: {status}")  # Additional debug output
        progress = progress_values.get(status, 0)
        print(f"Updating progress for item {item_id} with status '{status}' to {progress}%")  # Debug output
        dispatch_table.set(item_id, "Progress", f"{progress}%")
        return progress

    def update_selected_dispatch():
        selected_item = dispatch_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a dispatch to update")
            return

        status_options = ["In Transit", "Out for Delivery", "Delivered", "Completed"]
        status_combo = ttk.Combobox(form_frame, values=status_options, width=30)
        status_combo.grid(row=6, column=1, padx=10, pady=5)
        ttk.Button(form_frame, text="Update", command=lambda: update_status(selected_item)).grid(row=6, column=0,
                                                                                                 padx=10, pady=5)
        def update_status(selected_item):
            new_status = status_combo.get()
            if new_status and new_status in status_options:
                values = dispatch_table.item(selected_item)['values']
                dispatch_id = values[0]  # DispatchID
                driver_id = values[2]  # DriverID is in 3rd column
                truck_id = values[3]  # TruckID is in 4th column

                update_dispatch_status(dispatch_id, new_status)

                if new_status in ["Delivered", "Completed"]:
                    update_driver_status(driver_id, "Available")
                    update_truck_status(truck_id, "Operational")
                update_progress_bar(selected_item, new_status)  # Removed [0]
                load_data_into_table()

    def save_dispatch(dispatch_data):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO Dispatch (OrderID, DriverID, TruckID, DispatchTime, Status, EstimatedDeliveryTime)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, dispatch_data)
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "Dispatch assignment saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save dispatch: {e}")
            log_error(f"Failed to save dispatch: {e}")

    def check_truck_compatibility(truck_id, order_id):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()

                # Get truck details
                cursor.execute("""
                    SELECT WeightCapacity, Permit
                    FROM Trucks
                    WHERE TruckID = ?
                """, (truck_id,))
                truck_data = cursor.fetchone()
                if not truck_data:
                    return False
                truck_capacity, truck_permit = truck_data

                # Get order details
                cursor.execute("""
                    SELECT Weight, Region
                    FROM Orders
                    WHERE OrderID = ?
                """, (order_id,))
                order_data = cursor.fetchone()
                if not order_data:
                    return False
                order_weight, order_region = order_data

                conn.close()

                if float(truck_capacity) < float(order_weight):
                    messagebox.showerror("Error", "Order weight exceeds truck capacity!")
                    return False

                if truck_permit != order_region:
                    messagebox.showerror("Error", "Truck permit does not match order region!")
                    return False

                return True

        except Exception as e:
            messagebox.showerror("Error", f"Failed to check truck compatibility: {e}")
            return False

    def update_truck_odometer(truck_id, distance):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE Trucks
                    SET Odometer = Odometer + ?
                    WHERE TruckID = ?
                """, (float(distance), truck_id))
                conn.commit()
                conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update truck odometer: {e}")
            log_error(f"Failed to update truck odometer: {e}")

    def get_order_distance(order_id):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT Distance
                    FROM Orders
                    WHERE OrderID = ?
                """, (order_id,))
                result = cursor.fetchone()
                conn.close()
                return float(result[0]) if result else 0
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get order distance: {e}")
            return 0

    def get_available_drivers():
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT DriverID, Name
                    FROM Drivers
                    WHERE Status = 'Available'
                """)
                drivers = cursor.fetchall()
                conn.close()
                return [f"{driver[0]} - {driver[1]}" for driver in drivers]
            return []
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load available drivers: {e}")
            log_error(f"Failed to load available drivers: {e}")
            return []

    def get_available_trucks():
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT TruckID, Model
                    FROM Trucks
                    WHERE Status = 'Operational'
                """)
                trucks = cursor.fetchall()
                conn.close()
                return [f"{truck[0]} - {truck[1]}" for truck in trucks]
            return []
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load available trucks: {e}")
            log_error(f"Failed to load available trucks: {e}")
            return []

    def update_truck_status(truck_id, new_status):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE Trucks
                    SET Status = ?
                    WHERE TruckID = ?
                """, (new_status, truck_id))
                conn.commit()
                conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update truck status: {e}")
            log_error(f"Failed to update truck status: {e}")

    def get_pending_orders():
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT OrderID, CustomerName
                    FROM Orders
                    WHERE Status = 'Pending'
                """)
                orders = cursor.fetchall()
                conn.close()
                return [f"{order[0]} - {order[1]}" for order in orders]
            return []
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load pending orders: {e}")
            log_error(f"Failed to load pending orders: {e}")
            return []

    def update_order_status(order_id, new_status):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE Orders
                    SET Status = ?
                    WHERE OrderID = ?
                """, (new_status, order_id))
                conn.commit()
                conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update order status: {e}")
            log_error(f"Failed to update order status: {e}")

    def update_driver_status(driver_id, new_status):
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE Drivers
                    SET Status = ?,
                        Trip = CASE
                            WHEN ? = 'On Trip' THEN Trip + 1
                            ELSE Trip
                        END
                    WHERE DriverID = ?
                """, (new_status, new_status, driver_id))
                conn.commit()
                conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update driver status: {e}")
            log_error(f"Failed to update driver status: {e}")

    def refresh_combos():
        driver_combo.set("")
        truck_combo.set("")
        order_combo.set("")

    def load_data_into_table():
        dispatch_table.delete(*dispatch_table.get_children())
        dispatches = load_dispatch_data()
        for dispatch in dispatches:
            item = dispatch_table.insert('', 'end', values=dispatch)
            update_progress_bar(item, dispatch[5])  # Changed from dispatch[4] to dispatch[5]

    def show_loading_indicator():
        loading_label = ttk.Label(root, text="Loading...", font=("Helvetica", 12))
        loading_label.pack()
        root.update_idletasks()

    def assign_dispatch():
        if not order_combo.get() or not driver_combo.get() or not truck_combo.get():
            messagebox.showerror("Error", "Please select all required fields!")
            return

        order_id = order_combo.get().split(" - ")[0]
        driver_id = driver_combo.get().split(" - ")[0]
        truck_id = truck_combo.get().split(" - ")[0]

        if not check_truck_compatibility(truck_id, order_id):
            return

        dispatch_time = datetime.now().strftime("%Y-%m-%d %H:%M")

        # Use the separated date and time fields
        estimated_date_str = estimated_date.get()
        estimated_time_str = estimated_time.get() or "00:00"  # Default to midnight if no time entered
        estimated_delivery = f"{estimated_date_str} {estimated_time_str}"

        dispatch_data = [
            order_id,
            driver_id,
            truck_id,
            dispatch_time,
            "In Transit",
            estimated_delivery  # Use the combined date/time string
        ]

        order_distance = get_order_distance(order_id)
        update_truck_odometer(truck_id, order_distance)

        save_dispatch(dispatch_data)
        update_order_status(dispatch_data[0], "In Transit")
        update_driver_status(dispatch_data[1], "On Trip")
        update_truck_status(dispatch_data[2], "In Use")
        load_data_into_table()
        load_all_data()
        refresh_combos()

    def load_all_data():
        # Clear existing items
        order_tree.delete(*order_tree.get_children())
        driver_tree.delete(*driver_tree.get_children())
        truck_tree.delete(*truck_tree.get_children())

        # Load new data
        for order in load_orders_data():
            order_tree.insert('', 'end', values=(
                order[0],  # OrderID
                order[1],  # CustomerName
                order[2],  # Weight
                order[3],  # Distance
                order[4]  # Region
            ))

        for driver in load_drivers_data():
            driver_tree.insert('', 'end', values=(
                driver[0],  # DriverID
                driver[1],  # Name
                driver[2]  # Status
            ))

            for truck in load_trucks_data():
                truck_tree.insert('', 'end', values=(
                    truck[0],  # TruckID
                    truck[1],  # Model
                    truck[2],  # Status
                    truck[3],  # Capacity
                    truck[4]  # Permit
                ))

    def on_order_select(event):
        selected_item = order_tree.selection()
        if selected_item:
            values = order_tree.item(selected_item)['values']
            order_combo.set(f"{values[0]} - {values[1]}")

    def on_driver_select(event):
        selected_item = driver_tree.selection()
        if selected_item:
            values = driver_tree.item(selected_item)['values']
            driver_combo.set(f"{values[0]} - {values[1]}")

    def on_truck_select(event):
        selected_item = truck_tree.selection()
        if selected_item:
            values = truck_tree.item(selected_item)['values']
            truck_combo.set(f"{values[0]} - {values[1]}")

    # Add these functions just before the GUI setup code:

    def load_orders_data():
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT OrderID, CustomerName, Weight, Distance, Region
                    FROM Orders
                    WHERE Status = 'Pending'
                """)
                orders = cursor.fetchall()
                conn.close()
                return orders
            return []
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load orders data: {e}")
            log_error(f"Failed to load orders data: {e}")
            return []

    def load_drivers_data():
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT DriverID, Name, Status
                    FROM Drivers
                    WHERE Status = 'Available'
                """)
                drivers = cursor.fetchall()
                conn.close()
                return drivers
            return []
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load drivers:{e}")
            log_error(f"Failed to loareturn []:{e}")

    def load_trucks_data():
        try:
            conn = create_db_connection()
            if conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT TruckID, Model, Status, WeightCapacity, Permit
                    FROM Trucks
                    WHERE Status = 'Operational'
                """)
                trucks = cursor.fetchall()
                conn.close()
                return trucks
            return []
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load trucks data: {e}")
            log_error("Error",f"Failed to load trucks:{e}")
            return []

    def print_selected_dispatch():
        """Print selected dispatch details"""
        selected_item = dispatch_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a dispatch to print")
            return

        try:
            # Get selected dispatch details
            values = dispatch_table.item(selected_item)['values']

            # Create temporary file for printing
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
            with open(temp_file.name, 'w') as f:
                # Write header
                f.write("TruckFlow Solutions - Dispatch Details\n")
                f.write("=" * 50 + "\n\n")

                # Write dispatch details
                dispatch_details = [
                    ("Order ID:", values[0]),
                    ("Driver:", values[1]),
                    ("Truck:", values[2]),
                    ("Dispatch Time:", values[3]),
                    ("Status:", values[4]),
                    ("Estimated Delivery:", values[5]),
                    ("Progress:", values[6])
                ]

                for label, value in dispatch_details:
                    f.write(f"{label:<20} {value}\n")

            # Send to default printer
            os.startfile(temp_file.name, "print")
            messagebox.showinfo("Success", "Dispatch report sent to printer")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to print dispatch: {str(e)}")
            log_error(f"Failed to print dispatch: {str(e)}")

    def export_dispatches():
        """Export dispatch data to various formats"""
        try:
            # Get all dispatch data
            dispatches = load_dispatch_data()
            if not dispatches:
                messagebox.showerror("Error", "No dispatch data to export")
                return

            # Create exports directory if it doesn't exist

            # Get timestamp for filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Create export format selection window
            export_window = tk.Toplevel(root)
            export_window.title("Export Dispatches")
            export_window.geometry("300x150")

            def export_to_format(format_type):
                try:
                    if format_type == "excel":
                        filename = f"exports/dispatches_{timestamp}.xlsx"
                        df = pd.DataFrame(dispatches, columns=["Dispatch ID", "Order ID", "Driver", "Truck", "Dispatch Time", "Status", "Estimated Delivery Time"])
                        df.to_excel(filename, index=False)

                    elif format_type == "csv":
                        filename = f"exports/dispatches_{timestamp}.csv"
                        with open(filename, 'w', newline='') as f:
                            writer = csv.writer(f)
                            writer.writerow(
                                ["Order ID", "Driver", "Truck", "Dispatch Time", "Status", "Estimated Delivery Time"])
                            writer.writerows(dispatches)

                    elif format_type == "pdf":
                        filename = f"exports/dispatches_{timestamp}.pdf"
                        doc = SimpleDocTemplate(filename, pagesize=letter)
                        elements = []

                        # Add header and data to table
                        data = [["Order ID", "Driver", "Truck", "Dispatch Time", "Status", "Estimated Delivery Time"]]
                        data.extend(dispatches)

                        table = Table(data)
                        table.setStyle(TableStyle([
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))

                        elements.append(table)
                        doc.build(elements)

                    messagebox.showinfo("Success", f"Dispatches exported successfully to {filename}")
                    export_window.destroy()

                except Exception as e:
                    messagebox.showerror("Error", f"Failed to export: {str(e)}")
                    log_error(f"Failed to export: {str(e)}")

            # Export format buttons
            ttk.Button(export_window, text="Export to Excel", command=lambda: export_to_format("excel")).pack(pady=10)
            ttk.Button(export_window, text="Export to CSV", command=lambda: export_to_format("csv")).pack(pady=10)
            ttk.Button(export_window, text="Export to PDF", command=lambda: export_to_format("pdf")).pack(
                pady=10)  # Fixed pack()

        except Exception as e:  # Properly formatted exception handling
            messagebox.showerror("Error", f"Failed to prepare export: {str(e)}")
            log_error(f"Failed to prepare export: {str(e)}")

    def create_scrolled_tree(parent, columns, height=5, col_widths=None):
        container = ttk.Frame(parent)
        tree = ttk.Treeview(container, columns=columns, show="headings", height=height)

        # Scrollbars
        y_scroll = ttk.Scrollbar(container, orient=tk.VERTICAL, command=tree.yview)
        x_scroll = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=tree.xview)

        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # Set custom column widths if provided
        if col_widths:
            for col, width in zip(columns, col_widths):
                tree.column(col, width=width, anchor="center")

        return tree, container

    root = ThemedTk(theme="arc")
    root.title("Dispatch Management System")
    root.geometry("1024x768")
    root.minsize(1024, 600)

    # Configure root grid weights for responsiveness
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Header
    header_frame = tk.Frame(root, bg="#1a237e", height=100)  # Reduced from 150 to 100
    header_frame.grid(row=0, column=0, sticky="ew")

    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a separate frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)  # Navigation bar
    nav_frame = tk.Frame(header_frame, bg="#1a237e")
    nav_frame.pack(fill=tk.X, pady=5)
    ttk.Button(nav_frame, text="← Back", command=go_back).pack(side=tk.LEFT, padx=20)
    tk.Label(
        nav_frame,
        text="Dispatch Management System",
        font=("Helvetica", 16, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.LEFT, padx=20)

    # Main container
    main_container = ttk.Frame(root)
    main_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
    main_container.grid_rowconfigure(0, weight=1)  # Top section
    main_container.grid_rowconfigure(1, weight=2)  # Bottom section gets more space
    main_container.grid_columnconfigure(0, weight=1)  # Form
    main_container.grid_columnconfigure(1, weight=2)  # Trees

    # Left side - Form
    form_frame = ttk.LabelFrame(main_container, text="Dispatch Assignment")
    form_frame.grid(row=0, column=0, sticky="nw", padx=5, pady=5)

    # Replace the DateEntry with separate date and time inputs
    def on_focus_in(event):
        if estimated_time.get() == 'HH:MM':
            estimated_time.delete(0, 'end')
            estimated_time.configure(foreground='black')

    def on_focus_out(event):
        if estimated_time.get() == '':
            estimated_time.insert(0, 'HH:MM')
            estimated_time.configure(foreground='gray')

    fields = [
        ("Order", order_combo := ttk.Combobox(form_frame, values=get_pending_orders(), width=27)),
        ("Driver", driver_combo := ttk.Combobox(form_frame, values=get_available_drivers(), width=27)),
        ("Truck", truck_combo := ttk.Combobox(form_frame, values=get_available_trucks(), width=27)),
        ("Estimated Delivery Date", (estimated_date := DateEntry(form_frame, width=13,
                                                                 background='darkblue', foreground='white',
                                                                 borderwidth=2,
                                                                 date_pattern='yyyy-mm-dd'),
                                     estimated_time := ttk.Entry(form_frame, width=10)))
    ]

    for i, (label_text, widget) in enumerate(fields):
        ttk.Label(form_frame, text=label_text).grid(row=i, column=0, padx=10, pady=8, sticky="e")
        if isinstance(widget, tuple):
            widget[0].grid(row=i, column=1, padx=10, pady=8, sticky="w")
            widget[1].grid(row=i, column=1, padx=(7, 0), pady=8, sticky="e")  # Adjusted padx and sticky
        else:
            widget.grid(row=i, column=1, columnspan=2, padx=10, pady=8, sticky="w")

    # Configure placeholder for time entry
    estimated_time.insert(0, 'HH:MM')
    estimated_time.configure(foreground='gray')
    estimated_time.bind('<FocusIn>', on_focus_in)
    estimated_time.bind('<FocusOut>', on_focus_out)

    button_frame = ttk.Frame(form_frame)
    button_frame.grid(row=len(fields), column=0, columnspan=2, pady=15)

    # Create two rows of buttons (3 buttons per row)
    for idx, (text, command) in enumerate([
        ("Assign Dispatch", assign_dispatch),
        ("Update Status", update_selected_dispatch),
        ("Delete Dispatch", delete_dispatch),
        ("Print Dispatch", print_selected_dispatch),
        ("Export", export_dispatches),
        ("Refresh", load_data_into_table)
    ]):
        row = idx // 3  # Determines the row (0 or 1)
        col = idx % 3  # Determines the column (0, 1, or 2)
        ttk.Button(button_frame, text=text, command=command).grid(
            row=row, column=col, padx=5, pady=5, sticky="ew"
        )

    # Configure grid columns to be evenly spaced
    for i in range(3):
        button_frame.grid_columnconfigure(i, weight=1)

    # Create tree container frames with scrollbars

    # Right side trees section
    trees_frame = ttk.Frame(main_container)
    trees_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
    trees_frame.grid_columnconfigure(0, weight=1)
    trees_frame.grid_rowconfigure((0, 1), weight=1)

    # Drivers Tree
    drivers_frame = ttk.LabelFrame(trees_frame, text="Available Drivers")
    drivers_frame.grid(row=0, column=0, sticky="nsew", pady=5)
    driver_tree, driver_container = create_scrolled_tree(
        drivers_frame,
        ("ID", "Name", "Status"),
        height=4,
        col_widths=[80, 150, 100]
    )
    driver_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    driver_tree.bind('<<TreeviewSelect>>', on_driver_select)

    # Trucks Tree
    trucks_frame = ttk.LabelFrame(trees_frame, text="Available Trucks")
    trucks_frame.grid(row=1, column=0, sticky="nsew", pady=5)
    truck_tree, truck_container = create_scrolled_tree(
        trucks_frame,
        ("ID", "Model", "Status", "Capacity", "Permit"),
        height=4,
        col_widths=[80, 120, 100, 100, 100]
    )
    truck_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    truck_tree.bind('<<TreeviewSelect>>', on_truck_select)

    # Bottom section
    bottom_frame = ttk.Frame(main_container)
    bottom_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=10)
    bottom_frame.grid_columnconfigure((0, 1), weight=1)
    bottom_frame.grid_rowconfigure(0, weight=1)

    dispatch_frame = ttk.LabelFrame(bottom_frame, text="Dispatch Table")
    dispatch_frame.grid(row=0, column=0, sticky="nsew", padx=5)
    dispatch_columns = ("Dispatch ID", "Order ID", "Driver", "Truck", "Dispatch Time", "Status", "Estimated Delivery Time", "Progress")
    dispatch_table, dispatch_container = create_scrolled_tree(
        dispatch_frame,
        dispatch_columns,
        height=8,
        col_widths=[80, 120, 120, 150, 100, 150, 100]
    )
    dispatch_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    # Orders Table
    orders_frame = ttk.LabelFrame(bottom_frame, text="Pending Orders")
    orders_frame.grid(row=0, column=1, sticky="nsew", padx=5)
    order_columns = ("ID", "Customer", "Destination", "Weight", "Region")
    order_tree, order_container = create_scrolled_tree(
        orders_frame,
        order_columns,
        height=8,
        col_widths=[80, 150, 150, 100, 120]
    )
    order_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    order_tree.bind('<<TreeviewSelect>>', on_order_select)

    # Configure column widths for all trees
    for tree in [driver_tree, truck_tree, dispatch_table, order_tree]:
        for col in tree["columns"]:
            tree.column(col, width=100, anchor="center")
            tree.heading(col, text=col, anchor='center')

    load_all_data()
    load_data_into_table()
    footer_frame = tk.Frame(root, bg="#1a237e", height=30)
    footer_frame.grid(row=2, column=0, sticky="ew")

    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.BOTTOM, pady=10)

    root.mainloop()


def accounts_management_gui(user_role=""):
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()
        elif user_role == "Accountant":
            accountant_dashboard()

    def load_financial_data():
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute('''SELECT Date, Type, Amount, Description, PaymentMode
                             FROM Financials
                             ORDER BY Date DESC''')
            data = cursor.fetchall()
            conn.close()
            return data
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load financial data: {e}")
            return []

    def save_transaction(transaction_data):
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute('''INSERT INTO Financials (Date, Type, Amount, Description, PaymentMode)
                             VALUES (?, ?, ?, ?, ?)''', transaction_data)
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Transaction recorded successfully!")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save transaction: {e}")
            return False

    def generate_invoice(order_id):
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()
            cursor.execute('''SELECT OrderID, CustomerName, TotalAmount, GST, AmountStatus
                                FROM Orders WHERE OrderID = ?''', (order_id,))
            invoice_data = cursor.fetchone()
            conn.close()

            if invoice_data:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"Invoice_{order_id}_{timestamp}.txt"

                with open(filename, 'w') as f:
                    f.write("=== TRANSPORT MANAGEMENT SYSTEM ===\n")
                    f.write(f"Invoice Date: {datetime.now().strftime('%Y-%m-%d')}\n")
                    f.write(f"Order ID: {invoice_data[0]}\n")
                    f.write(f"Customer: {invoice_data[1]}\n")
                    f.write(f"Amount: /-{invoice_data[2]}\n")
                    f.write(f"GST: {invoice_data[3]}%\n")
                    f.write(f"Status: {invoice_data[4]}\n")

                messagebox.showinfo("Success", f"Invoice generated: {filename}")
            else:
                messagebox.showwarning("Warning", "Order ID not found.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate invoice: {e}")
    def record_expense():
        try:
            expense_type = expense_type_combo.get()
            amount = float(expense_amount_entry.get())
            description = expense_desc_entry.get()
            payment_types=epayment_type_combo.get()
            date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            transaction_data = [date, expense_type, amount, description, payment_types]
            if save_transaction(transaction_data):
                refresh_data()
                clear_expense_fields()
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount.")

    def record_payment():
        try:
            order_id = payment_order_entry.get()
            amount = float(payment_amount_entry.get())
            payment_type = payment_type_combo.get()
            date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            transaction_data = [date, "Order Payment", amount, f"Order ID: {order_id}", payment_type]
            if save_transaction(transaction_data):
                update_order_payment_status(order_id, amount)
                refresh_data()
                clear_payment_fields()
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid amount.")

    def update_order_payment_status(order_id, paid_amount):
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Get current order details
            cursor.execute('''SELECT TotalAmount, PaidAmount, RemainingAmount
                             FROM Orders WHERE OrderID = ?''', (order_id,))
            order_data = cursor.fetchone()

            if order_data:
                total_amount, current_paid, current_remaining = order_data
                new_remaining = float(current_remaining) - paid_amount
                new_paid = float(current_paid) + paid_amount

                # Determine new status
                status = "Completed" if new_paid >= float(total_amount) else "Partial"

                # Update order
                cursor.execute('''UPDATE Orders
                                SET PaidAmount = ?,
                                    RemainingAmount = ?,
                                    AmountStatus = ?
                                WHERE OrderID = ?''',
                             (new_paid, new_remaining, status, order_id))

            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update payment status: {e}")

    def clear_expense_fields():
        expense_type_combo.set('')
        expense_amount_entry.delete(0, tk.END)
        expense_desc_entry.delete(0, tk.END)

    def clear_payment_fields():
        payment_order_entry.delete(0, tk.END)
        payment_amount_entry.delete(0, tk.END)
        payment_type_combo.set('')

    def load_data_into_table():
        for item in transaction_table.get_children():
            transaction_table.delete(item)
        transactions = load_financial_data()
        for transaction in transactions:
            transaction_table.insert('', 'end', values=transaction)

    def calculate_totals():
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Calculate total payments
            cursor.execute('''SELECT SUM(Amount) FROM Financials
                             WHERE Type = 'Order Payment' ''')
            total_payments = cursor.fetchone()[0] or 0

            # Calculate total expenses
            cursor.execute('''SELECT SUM(Amount) FROM Financials
                             WHERE Type != 'Order Payment' ''')
            total_expenses = cursor.fetchone()[0] or 0

            balance = total_payments - total_expenses
            conn.close()
            return total_payments, total_expenses, balance
        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate totals: {e}")
            return 0, 0, 0

    def update_totals_display():
        total_payments, total_expenses, balance = calculate_totals()
        payments_label.config(text=f"Total Payments: /-{total_payments:.2f}")
        expenses_label.config(text=f"Total Expenses: /-{total_expenses:.2f}")
        balance_label.config(text=f"Balance: /-{balance:.2f}")

    # Modify the refresh_data function to include totals update:
    def refresh_data():
        load_data_into_table()
        update_totals_display()

    def print_transaction():
        """Print selected transaction details"""
        selected_item = transaction_table.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a transaction to print")
            return

        try:
            # Get selected transaction details
            item = transaction_table.item(selected_item[0])
            transaction_data = item['values']

            # Create temporary file for printing
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
            with open(temp_file.name, 'w') as f:
                # Write header
                f.write("=== TRUCKFLOW SOLUTIONS ===\n")
                f.write("Transaction Receipt\n")
                f.write("-" * 40 + "\n")

                transaction_details = [
                    ("Date:", transaction_data[0]),
                    ("Type:", transaction_data[1]),
                    ("Amount:", f"/-{transaction_data[2]}"),
                    ("Description:", transaction_data[3]),
                    ("Payment Method:", transaction_data[4])
                ]

                for label, value in transaction_details:
                    f.write(f"{label:<20} {value}\n")

                # Open print dialog
            os.startfile(temp_file.name, "print")
            messagebox.showinfo("Success", "Transaction sent to printer")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to print transaction: {e}")

    def export_to_pdf():
        """Export financial report to PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"exports/Financial_Report_{timestamp}.pdf"

            doc = SimpleDocTemplate(filename, pagesize=letter)
            elements = []

            # Add title
            styles = getSampleStyleSheet()
            elements.append(Paragraph("Financial Report", styles['Title']))

            # Get transaction data
            data = [["Date", "Type", "Amount", "Description", "Payment Method"]]
            for item in transaction_table.get_children():
                row = transaction_table.item(item)['values']
                data.append(row)

            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            elements.append(table)
            doc.build(elements)
            messagebox.showinfo("Success", f"Report exported as {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export report: {e}")

    def filter_transactions():
        """Filter transactions by date range and type"""
        filter_window = tk.Toplevel(root)
        filter_window.title("Filter Transactions")
        filter_window.geometry("400x300")

        ttk.Label(filter_window, text="Start Date:").pack(pady=5)
        start_date = DateEntry(filter_window, width=12, background='darkblue', foreground='white', borderwidth=2)
        start_date.pack(pady=5)
        ttk.Label(filter_window, text="End Date:").pack(pady=5)
        end_date = DateEntry(filter_window, width=12, background='darkblue', foreground='white', borderwidth=2)
        end_date.pack(pady=5)
        ttk.Label(filter_window, text="Transaction Type:").pack(pady=5)
        type_var = tk.StringVar()
        type_combo = ttk.Combobox(filter_window, textvariable=type_var, values=["All", "Order Payment"] + expense_types)
        type_combo.pack(pady=5)
        type_combo.set("All")

        def apply_filter():
            start = start_date.get_date()
            end = end_date.get_date()
            trans_type = type_var.get()

            for item in transaction_table.get_children():
                transaction_table.delete(item)

            data = load_financial_data()
            for transaction in data:
                trans_date = datetime.strptime(transaction[0], "%Y-%m-%d %H:%M:%S").date()
                if start <= trans_date <= end:
                    if trans_type == "All" or trans_type == transaction[1]:
                        transaction_table.insert('', 'end', values=transaction)

            filter_window.destroy()

        ttk.Button(filter_window, text="Apply Filter", command=apply_filter).pack(pady=20)

    root = ThemedTk(theme="arc")
    root.title("Accounts & Financial Management")
    root.geometry("1024x700")
    root.minsize(1024, 600)

    # Configure grid layout for root
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Header frame
    # Reduce overall header height
    header_frame = tk.Frame(root, bg="#1a237e", height=70)  # Reduced from 100 to 70
    header_frame.grid(row=0, column=0, sticky="nsew")
    header_frame.grid_columnconfigure(0, weight=1)

    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")
    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo with smaller dimensions
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((50, 50))
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Compact company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=5)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo
    logo_label.pack(side=tk.LEFT, padx=10)

    # Company name and details frame with center alignment
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)  # Changed to expand and fill

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 14, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(anchor="center")  # Added anchor="center"

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 8),
        bg="#1a237e",
        fg="white",
    ).pack(anchor="center")  # Added anchor="center"

    # Navigation bar with reduced spacing
    nav_frame = tk.Frame(header_frame, bg="#1a237e")
    nav_frame.pack(fill=tk.X, pady=2)  # Reduced pady from 5 to 2
    ttk.Button(nav_frame, text="← Back", command=go_back).pack(side=tk.LEFT, padx=10)  # Reduced padx from 20 to 10

    tk.Label(
        nav_frame,
        text="Accounts & Financial Management",
        font=("Helvetica", 12, "bold"),  # Reduced font size from 16 to 12
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.LEFT, padx=10)  # Reduced padx from 20 to 10

    # Financial summary labels
    summary_frame = tk.Frame(header_frame, bg="#1a237e")
    summary_frame.pack(side=tk.RIGHT, padx=20)

    # Base style with common properties
    common_label_style = {"font": ("Helvetica", 10, "bold"), "bg": "#1a237e"}

    # Payments label (white)
    payments_label = tk.Label(summary_frame, text="Total Payments: /-0.00", **common_label_style, fg="#FFFFFF")
    payments_label.pack(side=tk.LEFT, padx=10)

    # Expenses label (light red)
    expenses_label = tk.Label(summary_frame, text="Total Expenses: /-0.00", **common_label_style, fg="#FF8A80")
    expenses_label.pack(side=tk.LEFT, padx=10)

    # Balance label (light green)
    balance_label = tk.Label(summary_frame, text="Balance: /-0.00", **common_label_style, fg="#B2FF59")
    balance_label.pack(side=tk.LEFT, padx=10)

    # Main container
    main_container = ttk.Frame(root)
    main_container.grid(row=1, column=0, sticky="nsew")
    main_container.grid_columnconfigure(0, weight=1, minsize=400)  # Forms column
    main_container.grid_columnconfigure(1, weight=2, minsize=600)  # Table column
    main_container.grid_rowconfigure(0, weight=1)  # Make row expandable

    # Left-side form container
    # Left-side form container
    forms_frame = ttk.Frame(main_container)
    forms_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
    forms_frame.grid_columnconfigure(0, weight=1)  # Allow horizontal stretch
    forms_frame.grid_rowconfigure(0, weight=1)  # Allow vertical stretch

    # Expense form
    expense_frame = ttk.LabelFrame(forms_frame, text="Record Expense")
    expense_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)

    expense_types = ["Fuel", "Maintenance", "Salary", "Insurance", "Other"]
    ttk.Label(expense_frame, text="Type:").grid(row=0, column=0, padx=5, pady=5)
    expense_type_combo = ttk.Combobox(expense_frame, width=18, values=expense_types)
    expense_type_combo.grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(expense_frame, text="Amount:").grid(row=1, column=0, padx=5, pady=5)
    expense_amount_entry = ttk.Entry(expense_frame)
    expense_amount_entry.grid(row=1, column=1, padx=5, pady=5)

    ttk.Label(expense_frame, text="Description:").grid(row=2, column=0, padx=5, pady=5)
    expense_desc_entry = ttk.Entry(expense_frame)
    expense_desc_entry.grid(row=2, column=1, padx=5, pady=5)

    payment_types = ["Cash", "Credit", "Bank Transfer"]
    ttk.Label(expense_frame, text="Payment Type:").grid(row=3, column=0, padx=5, pady=5)
    epayment_type_combo = ttk.Combobox(expense_frame, width=18, values=payment_types)
    epayment_type_combo.grid(row=3, column=1, padx=5, pady=5)

    ttk.Button(expense_frame, text="Record Expense", command=record_expense).grid(row=4, column=0, columnspan=2,
                                                                                  pady=10)

    # Payment form
    payment_frame = ttk.LabelFrame(forms_frame, text="Record Payment")
    payment_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=5)

    ttk.Label(payment_frame, text="Order ID:").grid(row=0, column=0, padx=5, pady=5)
    payment_order_entry = ttk.Entry(payment_frame)
    payment_order_entry.grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(payment_frame, text="Amount:").grid(row=1, column=0, padx=5, pady=5)
    payment_amount_entry = ttk.Entry(payment_frame)
    payment_amount_entry.grid(row=1, column=1, padx=5, pady=5)

    ttk.Label(payment_frame, text="Payment Type:").grid(row=2, column=0, padx=5, pady=5)
    payment_type_combo = ttk.Combobox(payment_frame, width=18, values=payment_types)
    payment_type_combo.grid(row=2, column=1, padx=5, pady=5)

    ttk.Button(payment_frame, text="Record Payment", command=record_payment).grid(row=3, column=0, columnspan=2,
                                                                                  pady=10)

    # Invoice generation
    invoice_frame = ttk.LabelFrame(forms_frame, text="Generate Invoice")
    invoice_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=5)

    ttk.Label(invoice_frame, text="Order ID:").grid(row=0, column=0, padx=5, pady=5)
    invoice_order_entry = ttk.Entry(invoice_frame)
    invoice_order_entry.grid(row=0, column=1, padx=5, pady=5)

    ttk.Button(invoice_frame, text="Generate Invoice",
               command=lambda: generate_invoice(invoice_order_entry.get())).grid(row=1, column=0, columnspan=2, pady=10)

    # Adjust weights for rows in forms_frame
    forms_frame.grid_rowconfigure(0, weight=1)
    forms_frame.grid_rowconfigure(1, weight=1)
    forms_frame.grid_rowconfigure(2, weight=1)

    # Right-side table container
    table_frame = ttk.Frame(main_container)
    table_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
    table_frame.grid_columnconfigure(0, weight=1)
    table_frame.grid_rowconfigure(0, weight=1)

    # Add these buttons to the table_frame
    button_frame = ttk.Frame(table_frame)
    button_frame.grid(row=1, column=0, sticky="ew", pady=5)

    ttk.Button(button_frame, text="Print Selected", command=print_transaction).pack(side=tk.LEFT, padx=5)
    ttk.Button(button_frame, text="Export to PDF", command=export_to_pdf).pack(side=tk.LEFT, padx=5)
    ttk.Button(button_frame, text="Filter", command=filter_transactions).pack(side=tk.LEFT, padx=5)

    table_container = ttk.Frame(table_frame)
    table_container.grid(row=0, column=0, sticky="nsew")
    table_container.grid_columnconfigure(0, weight=1)
    table_container.grid_rowconfigure(0, weight=1)

    scrollbar = ttk.Scrollbar(table_container)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    columns = ("Date", "Type", "Amount", "Description", "Category")
    transaction_table = ttk.Treeview(table_container, columns=columns, show="headings", yscrollcommand=scrollbar.set)
    for col in columns:
        transaction_table.heading(col, text=col, anchor="center")
        transaction_table.column(col, anchor="center", width=100)  # Set minimum column width
    transaction_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=transaction_table.yview)

    # Footer frame
    footer_frame = tk.Frame(root, bg="#1a237e", height=30)
    footer_frame.grid(row=2, column=0, sticky="nsew")
    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(pady=10)

    # Refresh data on load
    refresh_data()

    # Start the application
    root.mainloop()

def reports_analytics_gui(user_role=""):
    def go_back():
        root.destroy()
        if user_role == "Admin":
            admin_dashboard()
        elif user_role == "Manager":
            manager_dashboard()

    def export_to_excel(data, filename):
        try:
            # Create a new workbook and select active sheet
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # Only proceed if there is data to export
            if not data:
                messagebox.showinfo("Warning", "No data to export!")
                return

            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)

            # Write data
            for row, record in enumerate(data, 2):
                for col, key in enumerate(headers, 1):
                    sheet.cell(row=row, column=col, value=record[key])

            # Save with timestamp
            export_filename = f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            workbook.save(export_filename)
            workbook.close()

            messagebox.showinfo("Success", f"Data exported successfully to {export_filename}!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {e}")
    def export_to_pdf(data, filename):
        try:
            pdf = FPDF(format='A4')  # Explicitly specify format
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()

            # Set font before writing
            pdf.set_font("helvetica", size=11)

            # Add headers
            headers = list(data[0].keys())
            for header in headers:
                pdf.cell(40, 10, str(header), 1)
            pdf.ln()

            # Add data
            for row in data:
                for key in headers:
                    pdf.cell(40, 10, str(row[key]), 1)
                pdf.ln()

            # Save with timestamp
            export_filename = f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            pdf.output(export_filename)
            messagebox.showinfo("Success", f"PDF exported successfully as {export_filename}!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export PDF: {e}")

    def export_to_csv(data, filename):
        try:
            import csv
            export_filename = f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

            with open(export_filename, 'w', newline='') as csvfile:
                if not data:
                    messagebox.showinfo("Warning", "No data to export!")
                    return

                # Write headers
                fieldnames = list(data[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()

                # Write data
                for row in data:
                    writer.writerow(row)

            messagebox.showinfo("Success", f"Data exported successfully to {export_filename}!")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export CSV: {e}")

    def generate_fleet_report():
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Query to get truck usage data
            query = """
            SELECT 
                t.TruckID,
                t.Model,
                t.Status,
                COUNT(d.DispatchID) as TotalTrips,
                t.Odometer
            FROM Trucks t
            LEFT JOIN Dispatch d ON t.TruckID = d.TruckID
            GROUP BY t.TruckID
            """

            cursor.execute(query)
            results = cursor.fetchall()

            fleet_data = []
            for row in results:
                fleet_data.append({
                    "Truck ID": row[0],
                    "Model": row[1],
                    "Status": row[2],
                    "Total Trips": row[3],
                    "Odometer": row[4]
                })

            update_fleet_chart(fleet_data)
            conn.close()
            return fleet_data

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate fleet report: {e}")
            return []

    def generate_driver_performance():
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Query to get driver performance data
            query = """
            SELECT 
                d.DriverID,
                d.Name,
                COUNT(CASE WHEN di.Status = 'Delivered' THEN 1 END) as CompletedTrips,
                d.Status
            FROM Drivers d
            LEFT JOIN Dispatch di ON d.DriverID = di.DriverID
            GROUP BY d.DriverID
            """

            cursor.execute(query)
            results = cursor.fetchall()

            driver_data = []
            for row in results:
                driver_data.append({
                    "Driver ID": row[0],
                    "Name": row[1],
                    "Completed Trips": row[2],
                    "Status": row[3]
                })

            update_driver_chart(driver_data)
            conn.close()
            return driver_data

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate driver performance: {e}")
            return []

    def generate_financial_summary():
        try:
            conn = sqlite3.connect('TTMS.db')
            cursor = conn.cursor()

            # Query to calculate total revenue from orders
            revenue_query = """
            SELECT COALESCE(SUM(TotalAmount), 0) 
            FROM Orders
            """

            # Query to calculate total expenses
            expenses_query = """
            SELECT COALESCE(SUM(Amount), 0)
            FROM Financials 
            WHERE Type != 'Order Payment'
            """

            cursor.execute(revenue_query)
            total_revenue = cursor.fetchone()[0]

            cursor.execute(expenses_query)
            total_expenses = cursor.fetchone()[0]

            net_profit = total_revenue - total_expenses

            # Create chart data dictionary
            chart_data = {
                "Total Revenue": total_revenue,
                "Total Expenses": total_expenses,
                "Net Profit": net_profit
            }

            # Update chart
            update_financial_chart(chart_data)

            # Create list format for export
            financial_data = [{
                "Category": key,
                "Amount": value
            } for key, value in chart_data.items()]

            conn.close()
            return financial_data

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate financial summary: {e}")
            print(f"Debug - Financial Summary Error: {str(e)}")
            return []

    def update_financial_chart(data):
        try:
            # Clear previous chart
            for widget in financial_chart_frame.winfo_children():
                widget.destroy()

            # Create new bar chart for financial summary
            fig = Figure(figsize=(5, 3))
            ax = fig.add_subplot(111)

            categories = list(data.keys())
            values = list(data.values())

            ax.bar(categories, values, color=['green', 'red', 'blue'])
            ax.set_title("Financial Summary")
            ax.set_ylabel("Amount (/-)")

            # Rotate x-axis labels for better readability
            plt.setp(ax.get_xticklabels(), rotation=45)

            canvas = FigureCanvasTkAgg(fig, financial_chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        except Exception as e:
            print(f"Debug - Chart Update Error: {str(e)}")  # Add this for debugging

    def update_fleet_chart(data):
        # Clear previous chart
        for widget in fleet_chart_frame.winfo_children():
            widget.destroy()

        # Create new bar chart for fleet usage
        fig = Figure(figsize=(5, 3))
        ax = fig.add_subplot(111)

        trucks = [d["Truck ID"] for d in data]
        trips = [d["Total Trips"] for d in data]

        ax.bar(trucks, trips)
        ax.set_title("Fleet Usage")
        ax.set_xlabel("Truck ID")
        ax.set_ylabel("Total Trips")

        canvas = FigureCanvasTkAgg(fig, fleet_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def update_driver_chart(data):
        # Clear previous chart
        for widget in driver_chart_frame.winfo_children():
            widget.destroy()

        # Create new bar chart for driver performance
        fig = Figure(figsize=(5, 3))
        ax = fig.add_subplot(111)

        drivers = [d["Driver ID"] for d in data]
        completed_trips = [d["Completed Trips"] for d in data]

        ax.bar(drivers, completed_trips)
        ax.set_title("Driver Performance")
        ax.set_xlabel("Driver ID")
        ax.set_ylabel("Completed Trips")

        canvas = FigureCanvasTkAgg(fig, driver_chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    # Create main window
    root = ThemedTk(theme="arc")
    root.title("Reports & Analytics")
    root.geometry("1000x600")
    root.minsize(1024, 600)

    # Configure root grid weights
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Header
    # Replace the existing header section with:
    # Header frame
    header_frame = tk.Frame(root, bg="#1a237e", height=100)
    header_frame.grid(row=0, column=0, sticky="nsew")
    header_frame.grid_columnconfigure(0, weight=1)



    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")

    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a separate frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)
    # Navigation bar
    nav_frame = tk.Frame(header_frame, bg="#1a237e")
    nav_frame.pack(fill=tk.X, pady=5)
    ttk.Button(nav_frame, text="← Back", command=go_back).pack(side=tk.LEFT, padx=20)

    tk.Label(
        nav_frame,
        text="Reports & Analytics",
        font=("Helvetica", 16, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.LEFT, padx=20)

    # Add this before root.mainloop():

    # Main container
    main_container = ttk.Frame(root)
    main_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
    main_container.grid_columnconfigure(1, weight=1)  # Make display panel expand
    main_container.grid_rowconfigure(0, weight=1)  # Make container expand vertically

    control_panel = ttk.Frame(main_container)
    control_panel.grid(row=0, column=0, sticky="ns", padx=5)

    # Report generation buttons
    buttons = [
        ("Generate Fleet Performance", generate_fleet_report),
        ("Generate Driver Performance", generate_driver_performance),
        ("Generate Financial Summary", generate_financial_summary),
        ("Export Fleet Data", lambda: export_data(generate_fleet_report(), "fleet_report")),
        ("Export Driver Data", lambda: export_data(generate_driver_performance(), "driver_report")),
        ("Export Financial Data", lambda: export_data(generate_financial_summary(), "financial_report"))
    ]

    for idx, (text, command) in enumerate(buttons):
        ttk.Button(control_panel, text=text, command=command).grid(
            row=idx, column=0, pady=5, padx=5, sticky="ew")

    export_frame = ttk.LabelFrame(control_panel, text="Export Options")
    export_frame.grid(row=len(buttons) + 1, column=0, pady=10, padx=5, sticky="ew")

    export_format = tk.StringVar(value="excel")
    ttk.Radiobutton(export_frame, text="Excel", variable=export_format, value="excel").pack()
    ttk.Radiobutton(export_frame, text="PDF", variable=export_format, value="pdf").pack()
    ttk.Radiobutton(export_frame, text="CSV", variable=export_format, value="csv").pack()

    # Modify export button command
    def export_data(data, filename):
        format_type = export_format.get()
        if format_type == "excel":
            export_to_excel(data, filename)
        elif format_type == "pdf":
            export_to_pdf(data, filename)
        elif format_type == "csv":

            export_to_csv(data, filename)

    # Right panel for charts and data display
    display_panel = ttk.Frame(main_container)
    display_panel.grid(row=0, column=1, sticky="nsew", padx=5)

    # Configure display panel grid
    display_panel.grid_columnconfigure(0, weight=1)
    display_panel.grid_columnconfigure(1, weight=1)
    display_panel.grid_rowconfigure(0, weight=1)
    display_panel.grid_rowconfigure(1, weight=1)

    # Create chart frames with grid layout
    driver_chart_frame = ttk.LabelFrame(display_panel, text="Driver Performance")
    driver_chart_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    driver_chart_frame.grid_columnconfigure(0, weight=1)
    driver_chart_frame.grid_rowconfigure(0, weight=1)

    fleet_chart_frame = ttk.LabelFrame(display_panel, text="Fleet Usage")
    fleet_chart_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")
    fleet_chart_frame.grid_columnconfigure(0, weight=1)
    fleet_chart_frame.grid_rowconfigure(0, weight=1)

    financial_chart_frame = ttk.LabelFrame(display_panel, text="Financial Summary")
    financial_chart_frame.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")
    financial_chart_frame.grid_columnconfigure(0, weight=1)
    financial_chart_frame.grid_rowconfigure(0, weight=1)

    # Generate initial reports
    generate_fleet_report()
    generate_driver_performance()
    generate_financial_summary()
    # Footer frame
    footer_frame = tk.Frame(root, bg="#1a237e", height=30)
    footer_frame.grid(row=2, column=0, sticky="nsew")
    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(pady=10)

    root.mainloop()
def user_management_gui():
    def go_back():
        root.destroy()
        admin_dashboard()

    import sqlite3
    from sqlite3 import Error

    def create_connection():
        """Create a database connection"""
        conn = None
        try:
            conn = sqlite3.connect('TTMS.db', timeout=10)  # Add timeout
            return conn
        except Error as e:
            messagebox.showerror("Database Error", f"Error connecting to database: {e}")
        return conn

    def load_users():
        """Load user data from database"""
        conn = create_connection()
        if conn is not None:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM Users")
                users = cursor.fetchall()
                conn.close()
                return users
            except Error as e:
                messagebox.showerror("Error", f"Failed to load users: {e}")
                return []
        return []

    def save_user(user_data):
        conn = create_connection()
        if conn is not None:
            try:
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO Users (Username, Password, Role, FullName, Contact, Address, CNIC, Email)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, user_data)

                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "User added successfully!")
                return True
            except Error as e:
                messagebox.showerror("Error", f"Failed to save user: {e}")
                return False
        return False

    def update_user(user_id, updated_data):
        """Update existing user data"""
        conn = create_connection()
        if conn is not None:
            try:
                cursor = conn.cursor()
                cursor.execute("""
                        UPDATE Users
                        SET Username=?, Password=?, Role=?, FullName=?, Contact=?, Address=?, CNIC=?, Email=?
                        WHERE UserID=?
                    """, (*updated_data, user_id))
                conn.commit()
                messagebox.showinfo("Success", "User updated successfully!")
                return True
            except Error as e:
                messagebox.showerror("Error", f"Failed to update user: {e}")
                return False
            finally:
                conn.close()  # Ensure connection is always closed
        return False

    def delete_user(user_id):
        """Delete a user from database"""
        conn = create_connection()
        if conn is not None:
            try:
                cursor = conn.cursor()
                cursor.execute("DELETE FROM Users WHERE UserID=?", (user_id,))
                conn.commit()
                conn.close()
                messagebox.showinfo("Success", "User deleted successfully!")
                return True
            except Error as e:
                messagebox.showerror("Error", f"Failed to delete user: {e}")
                return False
        return False
    def add_user():
        if not all([fullname_entry.get(), username_entry.get(), password_entry.get(),
                    userrole_combo.get(), contact_entry.get(), address_entry.get(),
                    cnic_entry.get(), email_entry.get()]):
            messagebox.showerror("Error", "All fields are required!")
            return

        user_data = [
            username_entry.get(),
            password_entry.get(),
            userrole_combo.get(),
            fullname_entry.get(),
            contact_entry.get(),
            address_entry.get(),
            cnic_entry.get(),
            email_entry.get()
        ]

        if save_user(user_data):
            load_data_into_table()
            clear_form()

    def update_selected_user():
        selected_item = user_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a user to update")
            return

        values = user_table.item(selected_item)["values"]
        user_id = values[0]  # UserID is now the first column

        updated_data = [
            username_entry.get(),
            password_entry.get(),
            userrole_combo.get(),
            fullname_entry.get(),
            contact_entry.get(),
            address_entry.get(),
            cnic_entry.get(),
            email_entry.get()
        ]

        if update_user(user_id, updated_data):
            load_data_into_table()
            clear_form()

    def delete_selected_user():
        selected_item = user_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a user to delete")
            return

        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this user?"):
            values = user_table.item(selected_item)["values"]
            user_id = values[0]  # UserID is now the first column

            if delete_user(user_id):
                load_data_into_table()
                clear_form()

    def clear_form():
        for widget in [fullname_entry, username_entry, password_entry,
                       contact_entry, address_entry, cnic_entry, email_entry]:
            widget.delete(0, tk.END)
        userrole_combo.set("")

    def search_users(search_text, search_by):
        """Search users in database"""
        conn = create_connection()
        if conn is not None:
            try:
                cursor = conn.cursor()
                field_map = {
                    "Username": "Username",
                    "User Role": "Role",
                    "Contact": "Contact",
                    "CNIC": "CNIC",
                    "Full Name": "FullName"
                }

                field = field_map.get(search_by, "Username")
                cursor.execute(f"SELECT * FROM Users WHERE {field} LIKE ?", (f"%{search_text}%",))
                users = cursor.fetchall()
                conn.close()
                return users
            except Error as e:
                messagebox.showerror("Error", f"Failed to search users: {e}")
                return []
        return []

    def load_data_into_table(data=None):
        user_table.delete(*user_table.get_children())
        users = data if data is not None else load_users()
        for user in users:
            # user[0] is UserID, no need to add extra index
            user_table.insert("", "end", values=user)

    # Create main window
    root = ThemedTk(theme="arc")
    root.title("User Management System")
    root.geometry("1000x600")
    root.minsize(1024, 600)
    # Replace the existing header section with:
    # Configure root grid weights
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)

    # Header frame
    header_frame = tk.Frame(root, bg="#1a237e", height=100)
    header_frame.grid(row=0, column=0, sticky="nsew")
    header_frame.grid_columnconfigure(0, weight=1)



    # Then use it for your icon paths:
    icon_path = resource_path("icons/logo1.png")

    root.iconphoto(False, tk.PhotoImage(file=icon_path))

    # Load and resize logo for header
    logo_image = Image.open(icon_path)
    logo_image = logo_image.resize((80, 80))  # Adjust size as needed
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Modify company_info_frame layout
    company_info_frame = tk.Frame(header_frame, bg="#1a237e")
    company_info_frame.pack(fill=tk.X, pady=10)

    # Add logo to company info frame
    logo_label = tk.Label(company_info_frame, image=logo_photo, bg="#1a237e")
    logo_label.image = logo_photo  # Keep a reference!
    logo_label.pack(side=tk.LEFT, padx=20)

    # Company name and details in a separate frame
    company_text_frame = tk.Frame(company_info_frame, bg="#1a237e")
    company_text_frame.pack(expand=True, fill=tk.BOTH)

    tk.Label(
        company_text_frame,
        text="TruckFlow Solutions",
        font=("Helvetica", 18, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)

    tk.Label(
        company_text_frame,
        text="📍 Logistics Avenue, Transport City, Islamabad  |  📞 (+92) 333-5130-796  |  📧 info@truckflow.com",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(expand=True)
    # Navigation bar
    nav_frame = tk.Frame(header_frame, bg="#1a237e")
    nav_frame.pack(fill=tk.X, pady=5)
    ttk.Button(nav_frame, text="← Back", command=go_back).pack(side=tk.LEFT, padx=20)

    tk.Label(
        nav_frame,
        text="User Management System",
        font=("Helvetica", 16, "bold"),
        bg="#1a237e",
        fg="white",
    ).pack(side=tk.LEFT, padx=20)


    # Main container
    main_container = ttk.Frame(root, padding="10")
    main_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)

    # Search frame
    search_frame = ttk.LabelFrame(main_container, text="Search", padding="5")
    search_frame.pack(fill=tk.X, pady=(0, 10))

    search_criteria = ttk.Combobox(search_frame,
                                   values=["Username", "User Rolee", "Contact", "CNIC", "Full Name"])
    search_criteria.set("Username")
    search_criteria.pack(side=tk.LEFT, padx=5)

    search_entry = ttk.Entry(search_frame)
    search_entry.pack(side=tk.LEFT, padx=5)

    ttk.Button(search_frame, text="Search", command=search_users).pack(side=tk.LEFT, padx=5)

    # Form frame
    form_frame = ttk.LabelFrame(main_container, text="User Information", padding="10")
    form_frame.pack(fill=tk.X, pady=(0, 10))

    # Form fields
    fields = [
        ("Full Name:", fullname_entry := ttk.Entry(form_frame)),
        ("Username:", username_entry := ttk.Entry(form_frame)),
        ("Password:", password_entry := ttk.Entry(form_frame, show="*")),
        ("User Role:", userrole_combo := ttk.Combobox(form_frame,
                                                      values=["Admin", "Dispatcher", "Manager",
                                                              "Accountant", "Driver"])),
        ("Contact:", contact_entry := ttk.Entry(form_frame)),
        ("Address:", address_entry := ttk.Entry(form_frame)),
        ("CNIC:", cnic_entry := ttk.Entry(form_frame)),
        ("Email:", email_entry := ttk.Entry(form_frame))
    ]

    # Create grid layout for form fields
    for i, (label, widget) in enumerate(fields):
        ttk.Label(form_frame, text=label).grid(row=i // 2, column=(i % 2) * 2, padx=5, pady=5, sticky='e')
        widget.grid(row=i // 2, column=(i % 2) * 2 + 1, padx=5, pady=5, sticky='w')

    # Buttons frame
    buttons_frame = ttk.Frame(main_container)
    buttons_frame.pack(fill=tk.X, pady=(0, 10))

    ttk.Button(buttons_frame, text="Add User", command=add_user).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Update User", command=update_selected_user).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Delete User", command=delete_selected_user).pack(side=tk.LEFT, padx=5)
    ttk.Button(buttons_frame, text="Clear Form", command=clear_form).pack(side=tk.LEFT, padx=5)

    # Table frame
    table_frame = ttk.Frame(main_container)
    table_frame.pack(fill=tk.BOTH, expand=True)

    # Scrollbars
    y_scrollbar = ttk.Scrollbar(table_frame)
    y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    x_scrollbar = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
    x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    # Table
    columns = ("UserID", "Username", "Password", "Role", "FullName", "Contact",
               "Address", "CNIC", "Email")

    user_table = ttk.Treeview(table_frame, columns=columns, show="headings",
                              yscrollcommand=y_scrollbar.set,
                              xscrollcommand=x_scrollbar.set)

    # Configure scrollbars
    y_scrollbar.config(command=user_table.yview)
    x_scrollbar.config(command=user_table.xview)

    # Configure columns
    for col in columns:
        user_table.heading(col, text=col,anchor='center')
        user_table.column(col, width=100,anchor='center')

    user_table.pack(fill=tk.BOTH, expand=True)

    # Bind selection event
    def on_select(event):
        selected = user_table.selection()
        if selected:
            values = user_table.item(selected[0])['values']
            # Values indexes should match database schema
            username_entry.delete(0, tk.END)
            username_entry.insert(0, values[1])  # Username
            password_entry.delete(0, tk.END)
            password_entry.insert(0, values[2])  # Password
            userrole_combo.set(values[3])  # Role
            fullname_entry.delete(0, tk.END)
            fullname_entry.insert(0, values[4])  # FullName
            contact_entry.delete(0, tk.END)
            contact_entry.insert(0, values[5])  # Contact
            address_entry.delete(0, tk.END)
            address_entry.insert(0, values[6])  # Address
            cnic_entry.delete(0, tk.END)
            cnic_entry.insert(0, values[7])  # CNIC
            email_entry.delete(0, tk.END)
            email_entry.insert(0, values[8])  # Email

    user_table.bind('<<TreeviewSelect>>', on_select)

    # Load initial data
    load_data_into_table()

    # Footer frame
    footer_frame = tk.Frame(root, bg="#1a237e", height=30)
    footer_frame.grid(row=2, column=0, sticky="nsew")
    tk.Label(
        footer_frame,
        text="© 2024 TruckFlow Solutions. All rights reserved.",
        font=("Helvetica", 10),
        bg="#1a237e",
        fg="white",
    ).pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    login_window(window_size=(300, 400))  # Larger window
